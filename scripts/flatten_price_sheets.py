"""Aplana los listados de precios (Excel) a CSV commiteables en carpinteria/data/.

Por qué existe: los listados maestros viven en Excel locales (Windows del dueño,
98 MB el modelo grande de molduras) con fórmulas/links externos que NO sobreviven
ni a Railway ni a una subida directa a Google Sheets. La app solo necesita los
NÚMEROS finales, así que aplanamos los valores cacheados a un CSV chico que sí va
al repo y funciona igual en Windows, Mac y Linux (Railway).

Flujo de actualización de precios:
  1. El dueño actualiza su Excel como siempre.
  2. Alguien corre:  uv run python scripts/flatten_price_sheets.py
  3. Commitear los CSV regenerados.

Fuentes:
  - molduras_catalog.csv  <- "Molduras." (o "FINAL") del listado de molduras.
  - wood_datos.csv        <- "Datos" del cotizador de madera (sirve muebles + molduras).
"""
from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = REPO_ROOT / "carpinteria" / "data"

# Ubicaciones conocidas de los Excel maestros (Windows del dueño + Downloads en Mac).
MOLDURAS_CANDIDATES = [
    Path(r"C:\Users\Peluca\Documents\La casa del Carpintero\Licitaciones 2026\Listado de precios Molduras 26.5.26.xlsx"),
    Path.home() / "Downloads" / "Listado de precios Molduras 26.5.26.xlsx",
]
WOOD_CANDIDATES = [
    Path(r"C:\Users\Peluca\Documents\La casa del Carpintero\Cotizador_Madera_V2_Corregido.xlsx"),
    Path(r"C:\Users\Peluca\Downloads\Cotizador_Madera_V2_Corregido (2).xlsx"),
    Path.home() / "Downloads" / "Cotizador_Madera_V2_Corregido (2).xlsx",
    Path.home() / "Downloads" / "Cotizador_Madera_V2_Corregido.xlsx",
]

MOLDURAS_SHEETS = ("FINAL", "Molduras.", "Molduras. (2)")
WOOD_SHEET = "Datos"

MOLDURAS_HEADER = [
    "code", "family", "description",
    "width_mm", "height_mm", "price_meter_iva", "price_varilla_iva",
]
WOOD_HEADER = [
    "id", "species", "features",
    "thickness_in", "length_m", "width_in", "price_uyu", "supplier",
]


def _first_existing(candidates: list[Path], override: str | None) -> Path | None:
    if override:
        p = Path(override)
        return p if p.exists() else None
    return next((p for p in candidates if p.exists()), None)


def _num(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


def _txt(value: object) -> str:
    return str(value or "").strip()


def flatten_molduras(path: Path) -> list[list[object]]:
    """Lee la primera solapa de catálogo con datos y devuelve filas A-G válidas."""
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_order = [s for s in MOLDURAS_SHEETS if s in wb.sheetnames] + [
        s for s in wb.sheetnames if s not in MOLDURAS_SHEETS
    ]
    for sheet_name in sheet_order:
        ws = wb[sheet_name]
        rows: list[list[object]] = []
        for row in ws.iter_rows(values_only=True):
            code = _txt(row[0] if len(row) > 0 else "")
            family = _txt(row[1] if len(row) > 1 else "")
            description = _txt(row[2] if len(row) > 2 else "")
            width = _num(row[3] if len(row) > 3 else None)
            height = _num(row[4] if len(row) > 4 else None)
            meter = _num(row[5] if len(row) > 5 else None)
            varilla = _num(row[6] if len(row) > 6 else None)
            if code and family and description and width and height and meter and varilla:
                rows.append([code, family, description, width, height,
                             round(meter, 4), round(varilla, 4)])
        if rows:
            print(f"  molduras: '{sheet_name}' -> {len(rows)} filas")
            return rows
    return []


def flatten_wood(path: Path) -> list[list[object]]:
    """Lee la solapa 'Datos' (desde fila 4) y devuelve filas con madera y precio."""
    wb = load_workbook(path, data_only=True, read_only=True)
    if WOOD_SHEET not in wb.sheetnames:
        return []
    ws = wb[WOOD_SHEET]
    rows: list[list[object]] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx < 4:
            continue
        species = _txt(row[1] if len(row) > 1 else "")
        price = _num(row[9] if len(row) > 9 else None)
        if not species or not price:
            continue
        rows.append([
            _txt(row[0] if len(row) > 0 else ""),
            species,
            _txt(row[2] if len(row) > 2 else ""),
            _num(row[3] if len(row) > 3 else None) or "",
            _num(row[5] if len(row) > 5 else None) or "",
            _num(row[7] if len(row) > 7 else None) or "",
            round(price, 4),
            _txt(row[11] if len(row) > 11 else ""),
        ])
    print(f"  madera: 'Datos' -> {len(rows)} filas")
    return rows


def _write_csv(path: Path, header: list[str], rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(header)
        writer.writerows(rows)
    print(f"  -> escrito {path.relative_to(REPO_ROOT)} ({len(rows)} filas)")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--molduras", help="Ruta al Excel de molduras (override)")
    parser.add_argument("--wood", help="Ruta al Excel de madera (override)")
    args = parser.parse_args()

    molduras_path = _first_existing(MOLDURAS_CANDIDATES, args.molduras)
    wood_path = _first_existing(WOOD_CANDIDATES, args.wood)

    if molduras_path:
        print(f"Molduras: {molduras_path}")
        rows = flatten_molduras(molduras_path)
        if rows:
            _write_csv(DATA_DIR / "molduras_catalog.csv", MOLDURAS_HEADER, rows)
    else:
        print("Molduras: no encontré el Excel fuente (uso --molduras), salteo.")

    if wood_path:
        print(f"Madera: {wood_path}")
        rows = flatten_wood(wood_path)
        if rows:
            _write_csv(DATA_DIR / "wood_datos.csv", WOOD_HEADER, rows)
    else:
        print("Madera: no encontré el Excel fuente (uso --wood), salteo.")


if __name__ == "__main__":
    main()
