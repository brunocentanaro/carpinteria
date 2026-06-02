from __future__ import annotations

import json
import sys

from dotenv import load_dotenv

load_dotenv()


def _payment_days_from_text(text: str) -> int | None:
    import re
    if not text:
        return None
    matches = [int(m.group(1)) for m in re.finditer(r"(\d{1,3})\s*d[ií]as?", text.lower())]
    if matches:
        return max(matches)
    return None


def _default_shipping_unit(destination: str, quantity: int) -> float:
    from carpinteria.settings import (
        DEFAULT_BID_DESTINATION,
        LABOR_DAY_HOURS,
        SHIPPING_UNLOAD_DAY_PRICE_UYU,
        SHIPPING_UNLOAD_EMPLOYEES,
        SHIPPING_UNLOAD_HOURS,
    )
    if not destination:
        destination = DEFAULT_BID_DESTINATION
    from carpinteria.shipping import DEFAULT_SHIPPING_RATES
    dest = destination.lower()
    for key, price in DEFAULT_SHIPPING_RATES.items():
        if key in dest or dest in key:
            unload_total = 0.0
            if "montevideo" in dest or "mvd" in dest:
                unload_total = SHIPPING_UNLOAD_EMPLOYEES * SHIPPING_UNLOAD_HOURS / max(LABOR_DAY_HOURS, 1) * SHIPPING_UNLOAD_DAY_PRICE_UYU
            return round((price + unload_total) / max(1, int(quantity or 1)), 2)
    return 0.0


def _effective_payment_days(payment_days: int | None, payment_terms: str = "") -> int:
    from carpinteria.settings import DEFAULT_BID_PAYMENT_DAYS
    return payment_days or _payment_days_from_text(payment_terms) or DEFAULT_BID_PAYMENT_DAYS


def _effective_destination(destination: str = "") -> str:
    from carpinteria.settings import DEFAULT_BID_DESTINATION
    return destination or DEFAULT_BID_DESTINATION


def _payment_delay_percent(days: int) -> float:
    from carpinteria.settings import PAYMENT_DELAY_TIERS
    if days <= 0:
        return 0.0
    for max_days, pct in PAYMENT_DELAY_TIERS:
        if days <= max_days:
            return float(pct) / 100
    return float(PAYMENT_DELAY_TIERS[-1][1]) / 100


def _guarantee_percent(text: str) -> float:
    import re
    norm = _norm_export_text(text)
    match = re.search(r"(\d+(?:[.,]\d+)?)\s*%", norm)
    if match:
        return float(match.group(1).replace(",", ".")) / 100
    if "fiel" in norm or "cumplimiento" in norm:
        return 0.05
    return 0.0


ADDITIONAL_SERVICE_LABELS = {
    "rectification": "Rectificacion de medidas",
    "installation": "Colocacion",
    "painting": "Pintura",
    "varnishing": "Barniz",
    "polishing": "Lustre",
}


def _additional_services_from_data(data: dict) -> dict[str, bool]:
    raw = data.get("additional_services") or {}
    return {
        key: bool(raw.get(key))
        for key in ADDITIONAL_SERVICE_LABELS
    }


def _additional_services_text(services: dict[str, bool]) -> str:
    selected = [
        label
        for key, label in ADDITIONAL_SERVICE_LABELS.items()
        if services.get(key)
    ]
    return ", ".join(selected) if selected else "No incluidos"


def _get_tc() -> tuple[float, str]:
    try:
        from carpinteria.exchange_rate import fetch_bcu_usd
        return fetch_bcu_usd()
    except Exception:
        return 40.0, "fallback"


def _norm_export_text(value: object) -> str:
    import unicodedata
    raw = str(value or "").lower()
    raw = unicodedata.normalize("NFKD", raw)
    raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
    return raw


def _count_piece_labels(pieces: list[dict], *needles: str) -> int:
    total = 0
    for piece in pieces:
        label = _norm_export_text(piece.get("label", ""))
        if all(needle in label for needle in needles):
            total += int(piece.get("quantity", 1) or 1)
    return total


def _is_drawer_base_piece(piece: dict) -> bool:
    label = _norm_export_text(piece.get("label", ""))
    return "caj" in label and ("base" in label or "parte de abajo" in label)


def _parse_moldura_request_text(text: str) -> dict | None:
    import re
    normalized = _norm_export_text(text)
    dim = re.search(r"\b(\d+(?:[.,]\d+)?)\s*(?:mm)?\s*x\s*(\d+(?:[.,]\d+)?)\s*(?:mm)?\b", normalized)
    if not dim:
        return None
    material = None
    if any(token in normalized for token in ("pino", "nac", "nacional", " pn")):
        material = "pino"
    elif any(token in normalized for token in ("euca", "eucalipto", "eucaliptus", "imp", "importado")):
        material = "euca"
    family = None
    for candidate in ("liston", "barrote", "zocalo", "contravidrio", "media cana", "cuadro", "montante", "moldura"):
        if candidate in normalized:
            family = candidate
            break
    if family is None and "picado" in normalized:
        family = "montante"
    if family == "contravidrio":
        model_match = re.search(r"\b(?:n|no|num|numero|nro|nro\.)\s*[°º.]?\s*(\d{1,3})\b", normalized)
        if model_match:
            family = f"contravidrio {model_match.group(1)}"
    if family == "media cana":
        model_match = re.search(r"\b(?:n|no|num|numero|nro|nro\.)\s*[°º.]?\s*([a-z0-9-]+)\b", normalized)
        if model_match:
            family = f"media cana {model_match.group(1)}"
    if family == "cuadro":
        model_match = re.search(r"\b(?:n|no|num|numero|nro|nro\.)\s*[°º.]?\s*([a-z0-9-]+)\b", normalized)
        if model_match:
            family = f"cuadro {model_match.group(1)}"
    qty_match = re.search(r"\b(\d+(?:[.,]\d+)?)\s*(?:varillas?|listones?|barrotes?|zocalos?|molduras?|metros?|mts?|m)\b", normalized)
    quantity = float(qty_match.group(1).replace(",", ".")) if qty_match else 1.0
    unit = "metro" if re.search(r"\b(?:metros?|mts?|m)\b", normalized) and "3.3" not in normalized else "varilla"
    return {
        "width_mm": float(dim.group(1).replace(",", ".")),
        "height_mm": float(dim.group(2).replace(",", ".")),
        "material": material,
        "family": family,
        "quantity": quantity,
        "unit": unit,
        "include_iva": True,
    }


def _make_plan_images(q: dict, out_dir: str) -> list[tuple[str, str]]:
    from pathlib import Path
    import re
    from PIL import Image as PILImage, ImageDraw, ImageFont

    pieces = list((q.get("decomposition") or {}).get("pieces") or [])
    dims = q.get("dimensions") or {}
    width = float(dims.get("width_mm") or q.get("width_mm") or 0)
    height = float(dims.get("height_mm") or q.get("height_mm") or 0)
    depth = float(dims.get("depth_mm") or q.get("depth_mm") or 0)
    if not width:
        width = max((float(p.get("width_mm") or 0) for p in pieces), default=380)
    if not height:
        height = max((float(p.get("height_mm") or 0) for p in pieces), default=600)
    if not depth:
        depth = max((float(p.get("height_mm") or 0) for p in pieces if "lateral" in _norm_export_text(p.get("label"))), default=460)

    text = _norm_export_text(f"{q.get('item_name', '')} {q.get('notes', '')}")
    drawer_count = _count_piece_labels(pieces, "caj", "frente")
    door_count = _count_piece_labels(pieces, "puerta")
    shelf_count = _count_piece_labels(pieces, "estante")
    has_wheels = "rueda" in text or "movil" in text or any("RUEDA" in str(h.get("code", "")) for h in q.get("hardware_lines", []))
    has_lock = "cerradura" in text or "traba" in text or any("CERR" in str(h.get("code", "")) for h in q.get("hardware_lines", []))
    stacked_drawers = "cajonera" in text or (drawer_count > 1 and door_count == 0 and shelf_count == 0)

    try:
        font_title = ImageFont.truetype("arial.ttf", 18)
        font = ImageFont.truetype("arial.ttf", 12)
    except Exception:
        font_title = ImageFont.load_default()
        font = ImageFont.load_default()

    bg = "#f8fafc"
    wood = "#fed7aa"
    wood_light = "#fff7ed"
    stroke = "#92400e"
    ink = "#334155"

    def canvas(title: str) -> tuple[PILImage.Image, ImageDraw.ImageDraw]:
        img = PILImage.new("RGB", (360, 260), bg)
        d = ImageDraw.Draw(img)
        d.text((180, 14), title, fill="#0f172a", font=font_title, anchor="ma")
        return img, d

    def dims_label(d: ImageDraw.ImageDraw, box: tuple[int, int, int, int], w_label: str, h_label: str) -> None:
        x1, y1, x2, y2 = box
        d.line((x1, y2 + 18, x2, y2 + 18), fill=ink, width=1)
        d.line((x1, y2 + 12, x1, y2 + 24), fill=ink, width=1)
        d.line((x2, y2 + 12, x2, y2 + 24), fill=ink, width=1)
        d.text(((x1 + x2) / 2, y2 + 25), w_label, fill=ink, font=font, anchor="ma")
        d.line((x1 - 18, y1, x1 - 18, y2), fill=ink, width=1)
        d.line((x1 - 24, y1, x1 - 12, y1), fill=ink, width=1)
        d.line((x1 - 24, y2, x1 - 12, y2), fill=ink, width=1)
        d.text((x1 - 34, (y1 + y2) / 2), h_label, fill=ink, font=font, anchor="mm")

    prefix = re.sub(r"[^A-Za-z0-9_-]+", "_", str(q.get("item_code") or "item"))[:24]

    def save(img: PILImage.Image, name: str) -> str:
        path = str(Path(out_dir) / f"{prefix}_{name}.png")
        img.save(path)
        return path

    images: list[tuple[str, str]] = []

    front, d = canvas("Frente")
    box = (75, 55, 285, 205)
    d.rounded_rectangle(box, radius=4, fill=wood_light, outline=stroke, width=3)
    if drawer_count:
        rows = drawer_count if stacked_drawers else 1
        gap = 8
        row_h = (box[3] - box[1] - gap * (rows + 1)) / rows
        for i in range(rows):
            y = box[1] + gap + i * (row_h + gap)
            d.rectangle((box[0] + 8, y, box[2] - 8, y + row_h), fill=wood, outline=stroke, width=2)
            d.ellipse(((box[0] + box[2]) / 2 - 3, y + row_h / 2 - 3, (box[0] + box[2]) / 2 + 3, y + row_h / 2 + 3), fill=stroke)
            if has_lock:
                lx = box[2] - 28
                ly = y + row_h / 2
                d.ellipse((lx - 5, ly - 5, lx + 5, ly + 5), fill="#f8fafc", outline=ink, width=2)
                d.rectangle((lx - 1, ly + 4, lx + 2, ly + 10), fill=ink)
    elif door_count:
        for i in range(door_count):
            x = box[0] + i * ((box[2] - box[0]) / door_count)
            d.line((x, box[1], x, box[3]), fill=stroke, width=2)
    if has_wheels:
        for x in (box[0] + 35, box[2] - 35):
            d.line((x, box[3], x, box[3] + 18), fill=ink, width=3)
            d.ellipse((x - 8, box[3] + 12, x + 8, box[3] + 28), fill="#94a3b8", outline=ink, width=2)
    dims_label(d, box, f"{width:.0f}mm", f"{height:.0f}mm")
    images.append(("H2", save(front, "plano_frente")))

    side, d = canvas("Costado")
    box = (95, 55, 240, 205)
    d.polygon([(box[0], box[1]), (box[2], box[1]), (box[2] + 35, box[1] + 25), (box[0] + 35, box[1] + 25)], fill=wood, outline=stroke)
    d.polygon([(box[2], box[1]), (box[2] + 35, box[1] + 25), (box[2] + 35, box[3]), (box[2], box[3] - 25)], fill="#fdba74", outline=stroke)
    d.rectangle((box[0], box[1], box[2], box[3]), fill=wood_light, outline=stroke, width=3)
    if has_wheels:
        for x in (box[0] + 20, box[2] - 20, box[2] + 15, box[2] + 35):
            d.ellipse((x - 6, box[3] + 6, x + 6, box[3] + 18), fill="#94a3b8", outline=ink, width=2)
    dims_label(d, box, f"{depth:.0f}mm", f"{height:.0f}mm")
    images.append(("L2", save(side, "plano_costado")))

    top, d = canvas("Planta")
    box = (70, 85, 290, 180)
    d.rounded_rectangle(box, radius=3, fill=wood_light, outline=stroke, width=3)
    d.rectangle((box[0] + 18, box[1] + 16, box[2] - 18, box[3] - 16), fill=wood, outline="#d97706", width=1)
    d.line(((box[0] + box[2]) / 2, box[1], (box[0] + box[2]) / 2, box[3]), fill=stroke, width=1)
    dims_label(d, box, f"{width:.0f}mm", f"{depth:.0f}mm")
    images.append(("P2", save(top, "plano_planta")))

    return images


def handle_prices() -> dict:
    from dataclasses import asdict
    from carpinteria.catalog import ProductCatalog

    catalog = ProductCatalog.from_activa()
    tc, fecha = _get_tc()
    placas = catalog.filter(tipo_producto="PLACA")
    cantos = catalog.filter(tipo_producto="CANTO")
    return {
        "placas": [asdict(p) for p in placas],
        "cantos": [asdict(p) for p in cantos],
        "exchange_rate": {"buy": tc, "sell": tc},
        "tc_source": f"BCU {fecha}",
    }


def handle_quote(data: dict) -> dict:
    from carpinteria.catalog import ProductCatalog
    from carpinteria.calculator import calculate_quotation
    from carpinteria.schemas import CutPiece
    from carpinteria.shipping import default_shipping_provider

    pieces = [CutPiece(**p) for p in data["pieces"]]
    catalog = ProductCatalog.from_activa()
    tc, _ = _get_tc()

    destination = _effective_destination(data.get("destination", ""))
    shipping = default_shipping_provider() if destination else None

    payment_days = data.get("payment_days")
    if isinstance(payment_days, int) and payment_days <= 0:
        payment_days = None
    payment_days = _effective_payment_days(payment_days)

    q = calculate_quotation(
        pieces=pieces,
        catalog=catalog,
        tc=tc,
        material=data["material"],
        thickness_mm=float(data["thickness_mm"]),
        color=data["color"],
        boards_needed=data.get("boards_needed") or None,
        edge_banding_name=data.get("edge_banding_name") or None,
        payment_days=payment_days,
        shipping_provider=shipping,
        shipping_units=int(data.get("quantity") or 1),
        destination=destination,
    )
    return q.model_dump()


def handle_analyze(data: dict) -> dict:
    from carpinteria.vision import analyze_cutting_plan

    results = analyze_cutting_plan(data["image_path"])
    return {"plans": [r.model_dump() for r in results]}


def handle_analyze_pliego(data: dict) -> dict:
    from carpinteria.pliego import analyze_pliego

    return analyze_pliego(data["file_paths"])


def handle_quote_item(data: dict) -> dict:
    from carpinteria.catalog import ProductCatalog
    from carpinteria.calculator import calculate_quotation
    from carpinteria.hardware_catalog import DEFAULT_HARDWARE_PRICES_UYU, get_by_code
    from carpinteria.pliego import decompose_furniture
    from carpinteria.schemas import CutPiece, QuotationLine
    from carpinteria.shipping import default_shipping_provider

    item = data["item"]
    missing: list[str] = []

    material = item.get("material", "")
    mat_lower = material.lower()
    if "melam" in mat_lower or "mdf" in mat_lower or "fibro" in mat_lower:
        material = "melamínico"
    if not material:
        missing.append("Material de la placa (ej: melamínico, MDF)")

    thickness = item.get("thickness_mm", 0)
    if not thickness:
        missing.append("Espesor de placa en mm (ej: 18)")

    color = data.get("color", "")
    if not color:
        missing.append("Color de la placa (ej: blanco, gris humo)")

    dims = item.get("dimensions") or {}
    has_dims = dims.get("width_mm") or dims.get("height_mm")
    has_desc = bool(item.get("description"))
    if not has_dims and not has_desc:
        missing.append("Dimensiones del mueble (ancho, alto, profundidad en mm) o descripcion con medidas")

    if missing:
        return {
            "error": f"Faltan datos para cotizar {item.get('code', '?')} ({item.get('name', '?')}):",
            "missing_inputs": missing,
        }

    catalog = ProductCatalog.from_activa()
    tc, tc_fecha = _get_tc()

    placa_match = catalog.find_placa(material, float(thickness), color)
    if placa_match is None:
        available = sorted({
            f"{p.material or p.familia} {(p.espesor_mm or 0):.0f}mm — {p.nombre}"
            for p in catalog.filter(tipo_producto="PLACA")
        })
        return {
            "error": f"No se encontro placa para {item.get('code', '?')}: {material} {thickness}mm color '{color}'",
            "missing_inputs": [
                f"Placa {material} {thickness}mm en color '{color}' no existe en la lista de precios.",
                "Opciones disponibles:",
                *[f"  - {a}" for a in available[:8]],
            ],
        }

    try:
        decomposition = decompose_furniture(item)
    except Exception as e:
        return {
            "error": f"No se pudo descomponer el mueble {item.get('code', '?')}: {e}",
            "missing_inputs": ["Revisa que la descripcion del mueble tenga suficiente detalle (dimensiones, puertas, cajones, etc.)"],
        }

    pieces = [CutPiece(**p) for p in decomposition.get("pieces", [])]
    if not pieces:
        return {
            "error": f"No se pudieron extraer piezas para {item.get('code', '?')}",
            "missing_inputs": ["La IA no pudo descomponer este mueble en piezas. Intenta con una descripcion mas detallada."],
        }

    eb_name = item.get("edge_banding") or None
    canto = catalog.find_canto(eb_name or color)
    warnings: list[str] = []
    if not canto:
        warnings.append(f"No se encontro canto para '{eb_name or color}'. Se cotiza sin canto.")

    destination = _effective_destination(data.get("destination", ""))
    shipping = default_shipping_provider() if destination else None

    payment_days = data.get("payment_days")
    if isinstance(payment_days, int) and payment_days <= 0:
        payment_days = None
    payment_days = _effective_payment_days(payment_days, str((data.get("general_specs") or {}).get("payment_terms", "")))

    # Hardware: agent already chose curated codes; we just attach the
    # user-provided price (if any). Hardware is part of inputs before profit,
    # but operational surcharges still use only placa+canto as base.
    hardware_prices: dict = data.get("hardware_prices") or {}
    hw_lines = []
    pending_hardware: list[dict] = []
    quote_hw_lines: list[QuotationLine] = []
    for hw in decomposition.get("hardware", []):
        code = (hw.get("code") or "").strip()
        if not code:
            continue
        spec = get_by_code(code)
        if spec is None:
            warnings.append(f"Herraje fuera de catálogo: {code}")
            continue
        try:
            qty = int(hw.get("quantity") or 0)
        except (TypeError, ValueError):
            qty = 0
        if qty <= 0:
            continue
        unit_price = float(hardware_prices.get(code, 0) or DEFAULT_HARDWARE_PRICES_UYU.get(code, 0) or 0)
        line = {
            "code": spec.code,
            "concept": f"Herraje: {spec.name}",
            "category": spec.category,
            "quantity": qty,
            "unit": spec.unit,
            "unit_price": round(unit_price, 2),
            "subtotal": round(qty * unit_price, 2),
        }
        hw_lines.append(line)
        quote_hw_lines.append(QuotationLine(
            concept=line["concept"],
            quantity=qty,
            unit=spec.unit,
            unit_price=round(unit_price, 2),
            subtotal=line["subtotal"],
        ))
        if unit_price <= 0:
            pending_hardware.append({
                "code": spec.code,
                "name": spec.name,
                "category": spec.category,
                "unit": spec.unit,
                "quantity": qty,
            })

    q = calculate_quotation(
        pieces=pieces,
        catalog=catalog,
        tc=tc,
        material=material,
        thickness_mm=float(thickness),
        color=color,
        edge_banding_name=eb_name,
        payment_days=payment_days,
        shipping_provider=shipping,
        shipping_units=int(item.get("quantity", 1) or 1),
        destination=destination,
        extra_input_lines=quote_hw_lines,
    )

    result = q.model_dump()
    result["_tc"] = tc
    result["_tc_source"] = f"BCU {tc_fecha}"
    result["_payment_days"] = payment_days or 0
    result["item_code"] = item.get("code", "")
    result["item_name"] = item.get("name", "")
    result["item_quantity"] = int(item.get("quantity", 1))
    result["dimensions"] = item.get("dimensions") or {}
    result["decomposition"] = decomposition
    result["hardware_lines"] = hw_lines
    result["pending_hardware"] = pending_hardware
    if warnings:
        result["warnings"] = warnings

    return result


def handle_export_excel(data: dict) -> dict:
    import shutil
    import tempfile
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter as cl

    quotes = data.get("quotes", [])
    if not quotes:
        return {"error": "No hay cotizaciones para exportar"}
    additional_services = _additional_services_from_data(data)

    wb = openpyxl.Workbook()
    image_tmp = tempfile.mkdtemp(prefix="planos_")
    hf = Font(bold=True, size=11)
    hfw = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    pfill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    efill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bdr = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    money = '#,##0.00'
    pct_fmt = '0%'

    def hdr(ws, r, cols):
        for i, t in enumerate(cols, 1):
            c = ws.cell(row=r, column=i, value=t)
            c.font = hfw
            c.fill = hfill
            c.border = bdr

    ws_r = wb.active
    ws_r.title = "Resumen"
    hdr(ws_r, 1, ["Codigo", "Mueble", "Cant", "Precio Unit.", "Total (Cant x Unit)", "Estado"])

    total_refs = []
    for i, q in enumerate(quotes):
        r = i + 2
        code = q.get("item_code", "")
        has_error = q.get("has_error", False)

        ws_r.cell(row=r, column=1, value=code).border = bdr
        ws_r.cell(row=r, column=2, value=q.get("item_name", "")).border = bdr
        ws_r.cell(row=r, column=3, value=q.get("item_quantity", 1)).border = bdr

        if has_error:
            ws_r.cell(row=r, column=4, value=0).border = bdr
            ws_r.cell(row=r, column=5, value=0).border = bdr
            missing = q.get("missing_inputs") or [q.get("notes", "Error")]
            ws_r.cell(row=r, column=6, value="FALTA: " + "; ".join(str(m) for m in missing)).border = bdr
            for c in range(1, 7):
                ws_r.cell(row=r, column=c).fill = efill
        else:
            sheet_name = code[:31]
            ws_r.cell(row=r, column=4).border = bdr
            ws_r.cell(row=r, column=4, value=f"='{sheet_name}'!B3")
            ws_r.cell(row=r, column=4).number_format = money
            ws_r.cell(row=r, column=5, value=f"=C{r}*D{r}").border = bdr
            ws_r.cell(row=r, column=5).number_format = money
            ws_r.cell(row=r, column=6, value="OK").border = bdr
            total_refs.append(f"E{r}")

    tr = len(quotes) + 2
    ws_r.cell(row=tr, column=4, value="TOTAL").font = Font(bold=True, size=13)
    if total_refs:
        ws_r.cell(row=tr, column=5, value=f"={'+'.join(total_refs)}")
    ws_r.cell(row=tr, column=5).font = Font(bold=True, size=13)
    ws_r.cell(row=tr, column=5).number_format = money

    services_row = tr + 2
    ws_r.cell(row=services_row, column=1, value="SERVICIOS ADICIONALES").font = hf
    ws_r.cell(row=services_row, column=1).fill = pfill
    ws_r.cell(row=services_row, column=2, value=_additional_services_text(additional_services)).border = bdr
    ws_r.merge_cells(start_row=services_row, start_column=2, end_row=services_row, end_column=6)

    ws_r.column_dimensions["A"].width = 10
    ws_r.column_dimensions["B"].width = 40
    ws_r.column_dimensions["C"].width = 8
    ws_r.column_dimensions["D"].width = 18
    ws_r.column_dimensions["E"].width = 18
    ws_r.column_dimensions["F"].width = 50

    for q in quotes:
        code = q.get("item_code", "?")
        ws = wb.create_sheet(title=code[:31])

        if q.get("has_error"):
            ws.cell(row=1, column=1, value=f"{code} - {q.get('item_name', '')}").font = Font(bold=True, size=14)
            ws.cell(row=2, column=1, value="NO SE PUDO COTIZAR").font = Font(bold=True, size=12, color="FF0000")
            rr = 4
            for m in q.get("missing_inputs") or [q.get("notes", "Error")]:
                ws.cell(row=rr, column=1, value=f"  - {m}")
                rr += 1
            ws.column_dimensions["A"].width = 80
            continue

        # A1: titulo, B1: label parametros
        ws.cell(row=1, column=1, value=f"{code} - {q.get('item_name', '')}").font = Font(bold=True, size=14)

        # Row 2: parametros editables
        ws.cell(row=2, column=1, value="TOTAL UNITARIO →").font = Font(bold=True, size=13)
        # B3 will hold the final total formula, set later
        ws.cell(row=2, column=2).font = Font(bold=True, size=13)
        ws.cell(row=2, column=2).number_format = money

        # Row 4: Parametros
        ws.cell(row=4, column=1, value="PARAMETROS (editalos para recalcular)").font = Font(bold=True, size=11)
        ws.cell(row=4, column=1).fill = pfill
        param_labels = [
            ("A5", "TC (tipo cambio)"), ("A6", "% Cortes base"), ("A7", "Cortes base max"),
            ("A8", "Mano de obra (jornales)"), ("A9", "% Maquinaria"), ("A10", "% Merma"),
            ("A11", "% Ganancia"), ("A12", "% Recargo financiero"),
            ("A13", "Flete base UYU"), ("A14", "Empleados descarga"), ("A15", "Horas descarga"),
            ("A16", "Jornal descarga UYU"), ("A17", "Cantidad entrega"),
        ]
        decomp = q.get("decomposition", {})
        pieces = [p for p in decomp.get("pieces", []) if not _is_drawer_base_piece(p)]
        n_cuts = sum(p.get("quantity", 1) * 2 for p in pieces)

        try:
            for anchor, image_path in _make_plan_images(q, image_tmp):
                img = XLImage(image_path)
                img.width = 220
                img.height = 160
                ws.add_image(img, anchor)
            for col_letter in ("H", "I", "J", "L", "M", "N", "P", "Q", "R"):
                ws.column_dimensions[col_letter].width = 13
            for row_idx in range(2, 11):
                ws.row_dimensions[row_idx].height = 19
        except Exception:
            pass

        from carpinteria.settings import (
            CUTS_PERCENT, CUTS_BASE_MAX, LABOR_DAY_HOURS, LABOR_PERCENT,
            MACHINERY_PERCENT, WASTE_PERCENT, PROFIT_PERCENT,
            SHIPPING_UNLOAD_DAY_PRICE_UYU, SHIPPING_UNLOAD_EMPLOYEES, SHIPPING_UNLOAD_HOURS,
        )
        flete_unit = 0.0
        for line in q.get("lines", []):
            if "flete" in str(line.get("concept", "")).lower():
                flete_unit = float(line.get("subtotal", 0) or 0)
        if flete_unit <= 0:
            flete_unit = _default_shipping_unit(str(q.get("_destination", "") or ""), int(q.get("item_quantity", 1) or 1))
        payment_days_export = int(q.get("_payment_days", 0) or 0)
        if payment_days_export <= 0:
            payment_days_export = _payment_days_from_text(str(q.get("_payment_terms", "") or "")) or 0
        if payment_days_export <= 0:
            rec_fin_pct = 0.0
        elif payment_days_export <= 30:
            rec_fin_pct = 0.05
        elif payment_days_export <= 45:
            rec_fin_pct = 0.08
        elif payment_days_export <= 60:
            rec_fin_pct = 0.10
        else:
            rec_fin_pct = 0.13

        param_vals = {
            "B5": q.get("_tc", 40),
            "B6": CUTS_PERCENT / 100,
            "B7": CUTS_BASE_MAX,
            "B8": n_cuts / 64,
            "B9": MACHINERY_PERCENT / 100,
            "B10": WASTE_PERCENT / 100,
            "B11": PROFIT_PERCENT / 100,
            "B12": rec_fin_pct,
            "B13": 0,
            "B14": SHIPPING_UNLOAD_EMPLOYEES,
            "B15": SHIPPING_UNLOAD_HOURS,
            "B16": SHIPPING_UNLOAD_DAY_PRICE_UYU,
            "B17": int(q.get("item_quantity", 1) or 1),
        }
        destination_for_flete = str(q.get("_destination", "") or "")
        from carpinteria.shipping import DEFAULT_SHIPPING_RATES
        for key, price in DEFAULT_SHIPPING_RATES.items():
            dest_norm = destination_for_flete.lower()
            if key in dest_norm or dest_norm in key:
                param_vals["B13"] = price
                break
        if flete_unit > 0 and param_vals["B13"] <= 0:
            param_vals["B13"] = flete_unit * max(1, int(q.get("item_quantity", 1) or 1))

        for cell_ref, label in param_labels:
            ws[cell_ref] = label
            ws[cell_ref].font = hf
        for cell_ref, val in param_vals.items():
            ws[cell_ref] = val
            ws[cell_ref].border = bdr
            if "%" in (dict(param_labels).get("A" + cell_ref[1:], "")):
                ws[cell_ref].number_format = pct_fmt
            elif cell_ref in ("B13", "B16"):
                ws[cell_ref].number_format = money

        # Row 19: Piezas de placa
        r = 19
        ws.cell(row=r, column=1, value="PIEZAS DE PLACA").font = Font(bold=True, size=11)
        ws.cell(row=r, column=1).fill = pfill
        r += 1
        hdr(ws, r, ["Pieza", "Ancho mm", "Alto mm", "Cant", "Cantos", "Area m2", "Canto metros"])
        r += 1
        piece_start = r
        for p in pieces:
            w = p.get("width_mm", 0)
            h = p.get("height_mm", 0)
            qty = p.get("quantity", 1)
            edges = p.get("edge_sides", [])
            ws.cell(row=r, column=1, value=p.get("label", "")).border = bdr
            ws.cell(row=r, column=2, value=w).border = bdr
            ws.cell(row=r, column=3, value=h).border = bdr
            ws.cell(row=r, column=4, value=qty).border = bdr
            ws.cell(row=r, column=5, value=", ".join(edges) if edges else "sin canto").border = bdr
            ws.cell(row=r, column=6, value=f"=B{r}*C{r}*D{r}/1000000").border = bdr
            ws.cell(row=r, column=6).number_format = '0.0000'
            ws.cell(row=r, column=7, value=f"=2*(B{r}+C{r})*D{r}/1000").border = bdr
            ws.cell(row=r, column=7).number_format = '0.00'
            r += 1
        piece_end = r - 1

        ws.cell(row=r, column=5, value="TOTAL").font = hf
        ws.cell(row=r, column=6, value=f"=SUM(F{piece_start}:F{piece_end})")
        ws.cell(row=r, column=6).font = hf
        ws.cell(row=r, column=6).number_format = '0.0000'
        area_total_cell = f"F{r}"
        ws.cell(row=r, column=7, value=f"=SUM(G{piece_start}:G{piece_end})")
        ws.cell(row=r, column=7).font = hf
        ws.cell(row=r, column=7).number_format = '0.00'
        canto_total_cell = f"G{r}"
        r += 2

        # Insumos: placas
        ws.cell(row=r, column=1, value="INSUMOS").font = Font(bold=True, size=11)
        ws.cell(row=r, column=1).fill = pfill
        r += 1
        hdr(ws, r, ["Insumo", "Cantidad", "Unidad", "Precio USD", "Precio UYU (=USD*TC)", "Subtotal UYU"])
        r += 1
        insumo_start = r

        for line in q.get("lines", []):
            concept = line.get("concept", "").lower()
            if "placa" in concept or "canto" in concept:
                ws.cell(row=r, column=1, value=line["concept"]).border = bdr
                ws.cell(row=r, column=2, value=line["quantity"]).border = bdr
                ws.cell(row=r, column=3, value=line["unit"]).border = bdr
                usd_price = line["unit_price"] / param_vals["B5"] if param_vals["B5"] else 0
                ws.cell(row=r, column=4, value=round(usd_price, 4)).border = bdr
                ws.cell(row=r, column=4).number_format = '0.0000'
                ws.cell(row=r, column=5, value=f"=D{r}*B5").border = bdr
                ws.cell(row=r, column=5).number_format = money
                if "placa" in concept:
                    board_meta = ((q.get("metadata") or {}).get("selected_placa") or {})
                    board_area = float(board_meta.get("area_m2") or 0)
                    contingency = float(board_meta.get("partial_contingency_percent") or 7.5) / 100
                    if board_area > 0:
                        ws.cell(row=r, column=6, value=f"=MIN(E{r},{area_total_cell}/{board_area}*E{r}*(1+{contingency}))").border = bdr
                    else:
                        ws.cell(row=r, column=6, value=f"=B{r}*E{r}").border = bdr
                else:
                    ws.cell(row=r, column=6, value=f"=B{r}*E{r}").border = bdr
                ws.cell(row=r, column=6).number_format = money
                r += 1

        for hw in q.get("hardware_lines", []):
            ws.cell(row=r, column=1, value=hw["concept"]).border = bdr
            ws.cell(row=r, column=2, value=hw["quantity"]).border = bdr
            ws.cell(row=r, column=3, value="unidad").border = bdr
            ws.cell(row=r, column=4, value="").border = bdr
            ws.cell(row=r, column=5, value=hw["unit_price"]).border = bdr
            ws.cell(row=r, column=5).number_format = money
            ws.cell(row=r, column=6, value=f"=B{r}*E{r}").border = bdr
            ws.cell(row=r, column=6).number_format = money
            r += 1
        insumo_end = r - 1

        ws.cell(row=r, column=5, value="TOTAL INSUMOS").font = hf
        ws.cell(row=r, column=6, value=f"=SUM(F{insumo_start}:F{insumo_end})")
        ws.cell(row=r, column=6).font = hf
        ws.cell(row=r, column=6).number_format = money
        insumos_total_cell = f"F{r}"
        r += 2

        # Material subtotal (solo placas+cantos, sin herrajes) para recargos
        placa_canto_rows = []
        for rr2 in range(insumo_start, insumo_end + 1):
            val = ws.cell(row=rr2, column=1).value or ""
            if "placa" in val.lower() or "canto" in val.lower():
                placa_canto_rows.append(f"F{rr2}")
        mat_formula = "+".join(placa_canto_rows) if placa_canto_rows else "0"

        ws.cell(row=r, column=1, value="RECARGOS OPERATIVOS (sobre costo placas+cantos)").font = Font(bold=True, size=11)
        ws.cell(row=r, column=1).fill = pfill
        r += 1
        hdr(ws, r, ["Concepto", "", "", "Porcentaje", "", "Monto UYU"])
        r += 1

        ws.cell(row=r, column=1, value=f"Cortes ({n_cuts} cortes)").border = bdr
        ws.cell(row=r, column=4, value=f"=B6*{n_cuts}/B7").border = bdr
        ws.cell(row=r, column=4).number_format = '0.0%'
        ws.cell(row=r, column=6, value=f"=({mat_formula})*D{r}").border = bdr
        ws.cell(row=r, column=6).number_format = money
        cortes_row = r
        r += 1

        ws.cell(row=r, column=1, value="Mano de obra").border = bdr
        ws.cell(row=r, column=4, value="=B8").border = bdr
        ws.cell(row=r, column=4).number_format = '0.0000'
        ws.cell(row=r, column=6, value=f"=D{r}*2500").border = bdr
        ws.cell(row=r, column=6).number_format = money
        mo_row = r
        r += 1

        ws.cell(row=r, column=1, value="Maquinaria").border = bdr
        ws.cell(row=r, column=4, value="=B9").border = bdr
        ws.cell(row=r, column=4).number_format = pct_fmt
        ws.cell(row=r, column=6, value=f"=({mat_formula})*D{r}").border = bdr
        ws.cell(row=r, column=6).number_format = money
        maq_row = r
        r += 1

        ws.cell(row=r, column=1, value="Merma").border = bdr
        ws.cell(row=r, column=4, value="=B10").border = bdr
        ws.cell(row=r, column=4).number_format = pct_fmt
        ws.cell(row=r, column=6, value=f"=({mat_formula})*D{r}").border = bdr
        ws.cell(row=r, column=6).number_format = money
        merma_row = r
        r += 2

        # Subtotal
        ws.cell(row=r, column=5, value="SUBTOTAL").font = hf
        ws.cell(row=r, column=6, value=f"={insumos_total_cell}+F{cortes_row}+F{mo_row}+F{maq_row}+F{merma_row}")
        ws.cell(row=r, column=6).font = hf
        ws.cell(row=r, column=6).number_format = money
        subtotal_cell = f"F{r}"
        r += 1

        ws.cell(row=r, column=5, value="GANANCIA").font = hf
        ws.cell(row=r, column=6, value=f"={subtotal_cell}*B11")
        ws.cell(row=r, column=6).number_format = money
        ganancia_cell = f"F{r}"
        r += 1

        ws.cell(row=r, column=5, value="TOTAL BASE")
        ws.cell(row=r, column=6, value=f"={subtotal_cell}+{ganancia_cell}")
        ws.cell(row=r, column=6).number_format = money
        total_base_cell = f"F{r}"
        r += 1

        ws.cell(row=r, column=5, value="REC. FINANCIERO")
        ws.cell(row=r, column=6, value=f"={total_base_cell}*B12")
        ws.cell(row=r, column=6).number_format = money
        fin_cell = f"F{r}"
        r += 1

        ws.cell(row=r, column=5, value="FLETE")
        ws.cell(row=r, column=6, value=f"=(B13+(B14*B15/{LABOR_DAY_HOURS}*B16))/MAX(1,B17)")
        ws.cell(row=r, column=6).number_format = money
        flete_cell = f"F{r}"
        r += 1

        ws.cell(row=r, column=5, value="SERVICIOS ADICIONALES")
        ws.cell(row=r, column=6, value=_additional_services_text(additional_services))
        ws.cell(row=r, column=6).border = bdr
        r += 1

        ws.cell(row=r, column=5, value="TOTAL UNITARIO").font = Font(bold=True, size=14)
        ws.cell(row=r, column=6, value=f"={total_base_cell}+{fin_cell}+{flete_cell}")
        ws.cell(row=r, column=6).font = Font(bold=True, size=14)
        ws.cell(row=r, column=6).number_format = money
        total_cell = f"F{r}"

        ws["B2"] = f"={total_cell}"
        ws["B2"].number_format = money
        ws["B3"] = f"={total_cell}"
        ws["B3"].number_format = money

        ws.column_dimensions["A"].width = 55
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 15

    fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="cotizacion_")
    wb.save(path)
    shutil.rmtree(image_tmp, ignore_errors=True)
    return {"excel_path": path}


def handle_export_molduras_excel(data: dict) -> dict:
    import os
    import tempfile
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    from carpinteria.molduras_prices import (
        ESTIMATED_MOLDURA_SURCHARGE,
        IVA_RATE,
        SCALE_THRESHOLD_VARILLAS,
        VARILLA_LENGTH_M,
        _family_cost,
        _family_group,
        _material_kind,
        estimate_price_from_conversor,
        quote_price,
    )

    requests = list(data.get("items") or [])
    if not requests:
        parsed = _parse_moldura_request_text(str(data.get("text") or ""))
        if parsed:
            requests = [parsed]
    if not requests:
        return {"error": "No hay molduras para exportar"}

    commercial = data.get("commercial") or {}
    general_specs = data.get("general_specs") or {}
    additional_services = _additional_services_from_data(data)
    payment_days = _effective_payment_days(
        commercial.get("payment_days", data.get("payment_days")),
        str(general_specs.get("payment_terms") or ""),
    )
    destination = _effective_destination(str(commercial.get("destination") or data.get("destination") or general_specs.get("delivery_location") or ""))
    delivery_days = general_specs.get("delivery_days")
    guarantee_text = str(general_specs.get("performance_guarantee") or "")
    guarantee_pct = _guarantee_percent(guarantee_text)
    payment_pct = _payment_delay_percent(payment_days)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"

    hf = Font(bold=True)
    title_font = Font(bold=True, size=15)
    small_font = Font(size=9)
    hfw = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    section_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    warn_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    total_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    bdr = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    money = '#,##0.00'

    def hdr(sheet, row: int, labels: list[str], start_col: int = 1) -> None:
        for i, label in enumerate(labels, start_col):
            cell = sheet.cell(row=row, column=i, value=label)
            cell.font = hfw
            cell.fill = hfill
            cell.border = bdr
            cell.alignment = Alignment(horizontal="center")

    def box(sheet, row: int, col: int, value="", fill=None, bold: bool = False):
        cell = sheet.cell(row=row, column=col, value=value)
        cell.border = bdr
        if fill:
            cell.fill = fill
        if bold:
            cell.font = hf
        return cell

    def quoted_sheet(title: str) -> str:
        return "'" + title.replace("'", "''") + "'"

    def sheet_safe_name(base: str, idx: int) -> str:
        clean = "".join(ch for ch in base if ch not in r'[]:*?/\\').strip()[:24]
        return f"{clean} {idx}"[:31]

    def profile_kind(family: str, description: str) -> str:
        import re
        text = _norm_export_text(f"{family} {description}")
        if "barrote" in text:
            return "barrote"
        if "zocal" in text:
            return "zocalo"
        if "contravidrio" in text:
            model = re.search(r"n[°º]?\s*\.?\s*(\d+)", text)
            if model and model.group(1) in {"1", "2", "31", "113", "137", "180", "229", "299", "410", "411", "412"}:
                return f"contravidrio-{model.group(1)}"
            return "contravidrio"
        if "contramarco" in text:
            if "finger" in text or "7x" in text or "9x" in text:
                return "contramarco-finger-n2" if "n2" in text else "contramarco-finger-n1"
            if re.search(r"\b6\s*x", text):
                return "contramarco-canal"
            return "contramarco-nariz"
        if "media cana" in text:
            model = re.search(r"n[°º]?\s*\.?\s*([a-z0-9-]+)", text)
            if model:
                value = model.group(1)
                if value in {"17", "38", "39", "50", "101", "101-a", "101a", "z-4", "z-6", "7-6"}:
                    return "media-cana-alta"
                if value in {"28", "30", "32", "33", "34", "35", "36"}:
                    return "media-cana-larga"
                if value in {"113", "229", "48", "4"}:
                    return "media-cana-cuadrada"
            return "media-cana-larga" if "x" in text and re.search(r"\b(5|6|8|9|10|11)\s*x\s*(1[7-9]|2[0-9]|3[0-9])", text) else "media-cana-cuadrada"
        if "cuadro" in text:
            model = re.search(r"n[Â°Âº]?\s*\.?\s*([a-z0-9-]+)", text)
            if model:
                value = model.group(1).replace("-a", "a")
                if value in {"118", "131", "133", "203", "213a", "224", "232", "234", "45", "53", "57", "60", "63", "95", "z-1"}:
                    return f"cuadro-{value}"
            return "cuadro"
        if "montante" in text:
            return "montante"
        if "liston" in text or "tabla" in text:
            return "liston"
        return "moldura"

    def product_title(item, material: str) -> str:
        import re
        text = _norm_export_text(f"{item.code} {item.family} {item.description} {material}")
        mat = ""
        if "impc" in text:
            mat = " PINO NACIONAL"
        elif "imp" in text or "euca" in text:
            mat = " EUCALYPTUS"
        elif "nac" in text or "pino" in text:
            mat = " PINO NACIONAL"
        model = re.search(r"n[°º]?\s*\.?\s*(\d+)", text)
        if "contravidrio" in text:
            no = f" No. {model.group(1)}" if model else ""
            return f"CONTRAVIDRIO {item.width_mm:g}x{item.height_mm:g}mm LA VARILLA 3.30mts{no}{mat}"
        if "contramarco" in text:
            finger = ""
            if "finger" in text:
                finger = f" Finger {'N2' if 'n2' in text else 'N1'}"
            length = "3.05mts" if finger else "3.30mts"
            return f"CONTRAMARCO {item.width_mm:g}x{item.height_mm:g}MM LA VARILLA {length}{finger}{mat}"
        if "media cana" in text:
            no = f" No. {model.group(1)}" if model else ""
            return f"MEDIA CAÑA {item.width_mm:g}x{item.height_mm:g}mm LA VARILLA 3.30mts{no}{mat}"
        if "cuadro" in text:
            model = re.search(r"n[Â°Âº]?\s*\.?\s*([a-z0-9-]+)", text)
            no = f" No. {model.group(1)}" if model else ""
            return f"MOLDURA P/CUADRO{no} LA VARILLA 3.30mts{mat}"
        if "montante" in text:
            return f"MONTANTE {item.width_mm:g}x{item.height_mm:g}MM LA VARILLA 3.30mts{mat}"
        return f"{item.family} {item.width_mm:g}x{item.height_mm:g}mm{mat}"

    image_paths: list[str] = []

    def add_profile_image(sheet, family: str, description: str, width_mm: float, height_mm: float) -> None:
        try:
            from PIL import Image, ImageDraw, ImageFont
            from openpyxl.drawing.image import Image as XLImage
        except Exception:
            sheet.cell(row=15, column=10, value=f"Perfil: {profile_kind(family, description)}")
            return

        kind = profile_kind(family, description)
        img = Image.new("RGB", (360, 230), "white")
        draw = ImageDraw.Draw(img)
        ink = "#8B3F10"
        dim = "#5F6B73"
        fill = "#FFF3E3"
        hatch = "#C98A54"

        def paste_hatch(mask):
            layer = Image.new("RGB", img.size, fill)
            layer_draw = ImageDraw.Draw(layer)
            for x in range(-120, 420, 16):
                layer_draw.line((x, 220, x + 210, 10), fill=hatch, width=2)
            img.paste(layer, mask=mask)

        def hatch_polygon(points):
            mask = Image.new("L", img.size, 0)
            ImageDraw.Draw(mask).polygon(points, fill=255)
            paste_hatch(mask)
            draw.line(points + [points[0]], fill=ink, width=4)

        def hatch_ellipse(box):
            mask = Image.new("L", img.size, 0)
            ImageDraw.Draw(mask).ellipse(box, fill=255)
            paste_hatch(mask)
            draw.ellipse(box, outline=ink, width=4)

        if kind == "barrote":
            body = [(164, 48), (308, 72), (308, 142), (164, 164)]
            draw.polygon(body, fill="#FFE2B7", outline=ink)
            draw.line((190, 55, 308, 74), fill=hatch, width=2)
            draw.line((194, 157, 308, 140), fill=hatch, width=2)
            hatch_ellipse((80, 44, 194, 158))
            draw.arc((80, 44, 194, 158), -82, 82, fill=ink, width=3)
        elif kind == "zocalo":
            pts = [(142, 22), (188, 28), (204, 60), (192, 92), (218, 125), (220, 194), (126, 194), (112, 150), (132, 112), (113, 88), (132, 58)]
            hatch_polygon(pts)
        elif kind == "contramarco-nariz":
            draw.polygon([(82, 112), (302, 46), (322, 66), (104, 146)], fill="#FFE2B7", outline=ink)
            draw.polygon([(78, 86), (298, 22), (322, 44), (100, 122)], fill="#FFF3E3", outline=ink)
            draw.line((116, 72, 250, 33), fill=ink, width=4)
            draw.line((132, 91, 270, 48), fill=hatch, width=4)
            draw.line((172, 78, 292, 43), fill=hatch, width=3)
            nose = [(78, 86), (58, 97), (63, 120), (100, 122), (88, 104)]
            hatch_polygon(nose)
        elif kind == "contramarco-canal":
            draw.polygon([(84, 122), (304, 52), (324, 74), (104, 154)], fill="#FFE2B7", outline=ink)
            draw.polygon([(80, 94), (300, 30), (324, 52), (102, 132)], fill="#FFF3E3", outline=ink)
            draw.line((128, 84, 252, 48), fill=ink, width=4)
            draw.line((142, 102, 270, 62), fill=ink, width=4)
            nose = [(80, 94), (60, 106), (66, 132), (102, 132), (90, 112)]
            hatch_polygon(nose)
        elif kind in {"contramarco-finger-n1", "contramarco-finger-n2"}:
            draw.polygon([(84, 122), (304, 54), (324, 76), (106, 154)], fill="#FFE2B7", outline=ink)
            draw.polygon([(80, 96), (300, 32), (324, 54), (102, 132)], fill="#FFF3E3", outline=ink)
            if kind == "contramarco-finger-n2":
                draw.line((128, 84, 238, 50), fill=ink, width=4)
                draw.line((156, 96, 262, 62), fill=hatch, width=3)
            else:
                draw.line((130, 90, 258, 52), fill=hatch, width=4)
            nose = [(80, 96), (61, 108), (67, 132), (102, 132), (90, 114)]
            hatch_polygon(nose)
        elif kind == "contravidrio-113":
            draw.polygon([(86, 90), (290, 42), (312, 64), (106, 134)], fill="#FFE2B7", outline=ink)
            nose = [(86, 90), (62, 106), (70, 136), (106, 134), (94, 112)]
            hatch_polygon(nose)
            draw.line((122, 90, 236, 56), fill=hatch, width=3)
        elif kind == "contravidrio-137":
            draw.polygon([(90, 68), (140, 40), (312, 72), (312, 132), (88, 132)], fill="#FFE2B7", outline=ink)
            hatch_ellipse((56, 52, 138, 136))
            draw.rectangle((96, 94, 312, 132), fill="#FFE2B7", outline=ink)
            draw.line((148, 62, 300, 88), fill=hatch, width=3)
        elif kind in {"contravidrio-229", "contravidrio-299"}:
            pts = [(74, 148), (300, 148), (300, 96), (188, 104), (128, 148)]
            hatch_polygon(pts)
        elif kind == "contravidrio-1":
            pts = [(82, 150), (302, 150), (238, 58), (136, 58)]
            hatch_polygon(pts)
        elif kind == "contravidrio-2":
            pts = [(76, 150), (308, 150), (244, 48), (138, 48), (138, 82), (104, 82)]
            hatch_polygon(pts)
        elif kind == "contravidrio-31":
            pts = [(76, 146), (306, 146), (306, 102), (134, 102), (134, 70), (76, 70)]
            hatch_polygon(pts)
        elif kind == "contravidrio-180":
            pts = [(78, 150), (300, 150), (300, 104), (250, 72), (134, 72), (78, 104)]
            hatch_polygon(pts)
        elif kind in {"contravidrio-410", "contravidrio-411", "contravidrio-412"}:
            pts = [(76, 150), (306, 150), (264, 66), (126, 66)]
            hatch_polygon(pts)
        elif kind == "contravidrio":
            pts = [(108, 166), (230, 166), (196, 54), (136, 54)]
            hatch_polygon(pts)
        elif kind == "media-cana":
            mask = Image.new("L", img.size, 0)
            mask_draw = ImageDraw.Draw(mask)
            mask_draw.pieslice((94, 44, 236, 186), 180, 360, fill=255)
            mask_draw.rectangle((94, 115, 236, 186), fill=255)
            paste_hatch(mask)
            draw.pieslice((94, 44, 236, 186), 180, 360, outline=ink, width=4)
            draw.line((94, 115, 236, 115), fill=ink, width=4)
        elif kind == "media-cana-cuadrada":
            mask = Image.new("L", img.size, 0)
            md = ImageDraw.Draw(mask)
            md.pieslice((92, 50, 244, 202), 180, 360, fill=255)
            md.rectangle((92, 126, 244, 202), fill=255)
            paste_hatch(mask)
            draw.pieslice((92, 50, 244, 202), 180, 360, outline=ink, width=4)
            draw.line((92, 126, 244, 126), fill=ink, width=4)
        elif kind == "media-cana-larga":
            pts = [(70, 168), (300, 168), (260, 96), (212, 70), (148, 70), (96, 96)]
            hatch_polygon(pts)
            draw.arc((118, 48, 242, 172), 190, 350, fill=ink, width=4)
        elif kind == "media-cana-alta":
            pts = [(78, 178), (286, 178), (286, 64), (254, 64), (226, 100), (116, 126), (78, 126)]
            hatch_polygon(pts)
            draw.arc((92, 54, 226, 188), 205, 340, fill=ink, width=4)
            draw.line((96, 154, 274, 154), fill=hatch, width=3)
            draw.line((108, 166, 282, 166), fill=hatch, width=3)
        elif kind == "liston":
            mask = Image.new("L", img.size, 0)
            max_dim = max(float(width_mm or 1), float(height_mm or 1), 1.0)
            rect_w = max(42, min(190, 190 * float(width_mm or 1) / max_dim))
            rect_h = max(42, min(125, 125 * float(height_mm or 1) / max_dim))
            cx, cy = 168, 106
            box = (
                int(cx - rect_w / 2),
                int(cy - rect_h / 2),
                int(cx + rect_w / 2),
                int(cy + rect_h / 2),
            )
            ImageDraw.Draw(mask).rectangle(box, fill=255)
            paste_hatch(mask)
            draw.rectangle(box, outline=ink, width=4)
        elif kind == "montante":
            box = (64, 82, 304, 140)
            draw.rectangle(box, fill="#FFF3E3", outline=ink, width=4)
            for i in range(4):
                x = 88 + i * 48
                cutout = Image.new("RGBA", (56, 34), (255, 255, 255, 0))
                cutout_draw = ImageDraw.Draw(cutout)
                cutout_draw.rounded_rectangle((6, 10, 50, 24), radius=7, fill=fill, outline=ink, width=4)
                for hx in range(12, 48, 8):
                    cutout_draw.line((hx, 24, hx + 16, 10), fill=hatch, width=2)
                cutout = cutout.rotate(-28, expand=True)
                img.paste(cutout, (x - 8, 92), cutout)
        elif kind.startswith("cuadro-") or kind == "cuadro":
            if kind in {"cuadro-203", "cuadro-224", "cuadro-232"}:
                pts = [(72, 170), (128, 46), (154, 86), (238, 102), (298, 98), (298, 128), (252, 128), (252, 150), (208, 150), (208, 170)]
            elif kind == "cuadro-45":
                pts = [(76, 172), (76, 70), (122, 34), (234, 34), (302, 72), (232, 100), (212, 138), (302, 138), (302, 172)]
            elif kind == "cuadro-57":
                pts = [(78, 170), (116, 90), (164, 58), (208, 108), (292, 92), (292, 130), (232, 146), (204, 170)]
            elif kind == "cuadro-z1":
                pts = [(82, 178), (82, 38), (118, 38), (118, 86), (210, 146), (284, 146), (284, 174), (182, 174), (182, 178)]
            elif kind == "cuadro-234":
                pts = [(86, 176), (86, 50), (124, 50), (124, 92), (210, 126), (294, 126), (294, 154), (220, 154), (220, 176)]
            else:
                pts = [(78, 174), (78, 54), (112, 28), (150, 34), (168, 76), (210, 88), (294, 88), (294, 120), (252, 120), (252, 154), (190, 154), (190, 174)]
            hatch_polygon(pts)
        else:
            pts = [(112, 34), (220, 34), (198, 70), (220, 104), (190, 128), (215, 176), (112, 176)]
            hatch_polygon(pts)

        draw.line((36, 202, 300, 202), fill=dim, width=3)
        draw.line((36, 194, 36, 210), fill=dim, width=3)
        draw.line((300, 194, 300, 210), fill=dim, width=3)
        draw.line((34, 30, 34, 188), fill=dim, width=3)
        draw.line((26, 30, 42, 30), fill=dim, width=3)
        draw.line((26, 188, 42, 188), fill=dim, width=3)
        width_label = f"{width_mm:g} mm"
        height_label = f"{height_mm:g} mm"
        label = f"Ø {max(width_mm, height_mm):g} mm" if kind == "barrote" else width_label
        draw.text((142, 204), label, fill=dim)
        if kind != "barrote":
            rotated = Image.new("RGBA", (92, 24), (255, 255, 255, 0))
            rdraw = ImageDraw.Draw(rotated)
            rdraw.text((2, 3), height_label, fill=dim)
            rotated = rotated.rotate(90, expand=True)
            img.paste(rotated, (3, 78), rotated)
        draw.text((245, 28), kind.replace("-", " "), fill=dim)

        fd, path = tempfile.mkstemp(suffix=".png", prefix="perfil_moldura_")
        os.close(fd)
        img.save(path)
        image_paths.append(path)
        xl_img = XLImage(path)
        xl_img.width = 270
        xl_img.height = 172
        sheet.add_image(xl_img, "J4")

    def material_hint(raw: dict, item) -> str:
        explicit = str(raw.get("material") or "").strip()
        if explicit:
            return explicit
        code = str(getattr(item, "code", "") or "").upper()
        if "IMPC" in code:
            return "pino"
        return "euca" if "IMP" in code else "pino"

    def write_conversor_sheet(sheet, q, raw: dict, idx: int) -> dict[str, str]:
        item = q.item
        material = material_hint(raw, item)
        group = _family_group(raw.get("family") or item.family, item.description)
        cost = _family_cost(group, material)
        breakdown = q.breakdown or {}
        if not breakdown:
            model_item = estimate_price_from_conversor(
                item.width_mm,
                item.height_mm,
                material=material,
                family=raw.get("family") or item.family,
            )
            breakdown = getattr(model_item, "_breakdown", {}) if model_item is not None else {}

        thick_cm = min(item.width_mm, item.height_mm) / 10
        wide_cm = max(item.width_mm, item.height_mm) / 10
        requested_length_m = float(raw.get("length_m") or raw.get("requested_length_m") or VARILLA_LENGTH_M)
        setup_average = max(SCALE_THRESHOLD_VARILLAS, float(q.quantity or 1))
        surcharge = float(breakdown.get("recargo_modelo_no_listado") or (ESTIMATED_MOLDURA_SURCHARGE - 1)) if q.estimated else 0.0

        sheet.cell(row=1, column=1, value=f"Conversor molduras - {item.family}").font = title_font
        sheet.cell(row=2, column=1, value="Esta hoja queda editable: podes tocar los campos amarillos y Excel recalcula el precio.").font = small_font
        sheet.cell(row=4, column=10, value="PERFIL / SECCION").font = hf
        sheet.cell(row=4, column=10).fill = section_fill
        add_profile_image(sheet, raw.get("family") or item.family, item.description, item.width_mm, item.height_mm)

        sheet.cell(row=4, column=1, value="MEDIDAS DE MOLDURA A REALIZAR").font = hf
        sheet.cell(row=4, column=1).fill = section_fill
        for row, label, value in (
            (5, "Espesor / alto cm", thick_cm),
            (6, "Ancho cm", wide_cm),
            (7, "Largo varilla m", requested_length_m),
            (8, "Cantidad pedida", q.quantity),
            (9, "Unidad pedida", q.unit),
            (10, "Fuente del precio", q.source),
            (11, "Precio varilla listado", item.price_varilla_iva if q.iva_included else item.price_varilla_without_iva),
            (12, "Precio metro listado", item.price_meter_iva if q.iva_included else item.price_meter_without_iva),
            (13, "Titulo asociado", product_title(item, material)),
        ):
            box(sheet, row, 1, label)
            c = box(sheet, row, 2, value, input_fill)
            if isinstance(value, (int, float)):
                c.number_format = "0.00"
            if row in (11, 12):
                c.number_format = money

        sheet.cell(row=4, column=4, value="TABLA PARA COTIZACION").font = hf
        sheet.cell(row=4, column=4).fill = section_fill
        hdr(sheet, 5, ["Madera", "Espesor pulg", "Ancho pulg", "Largo m", "Precio tabla UYU"], start_col=4)
        wood_values = [
            breakdown.get("madera") or ("Pino" if _material_kind(material) == "pino" else "Eucaliptus"),
            breakdown.get("tabla_espesor_pulg") or "",
            breakdown.get("tabla_ancho_pulg") or "",
            breakdown.get("tabla_largo_m") or VARILLA_LENGTH_M,
            breakdown.get("precio_tabla_uyu") or "",
        ]
        for col, value in enumerate(wood_values, 4):
            cell = box(sheet, 6, col, value, input_fill)
            if col == 8:
                cell.number_format = money

        sheet.cell(row=8, column=4, value="CALCULO DE MATERIA PRIMA").font = hf
        sheet.cell(row=8, column=4).fill = section_fill
        calc_rows = [
            ("Espesor tabla cm", "=E6*2.25"),
            ("Ancho tabla cm", "=F6*2.25"),
            ("Molduras por tabla sin merma", "=(E9/B5)*(E10/B6)"),
            ("Merma", 0.35),
            ("Molduras por tabla con merma", "=E11*(1-E12)"),
            ("MP por varilla", "=H6/E13"),
        ]
        for pos, (label, value) in enumerate(calc_rows, 9):
            box(sheet, pos, 4, label)
            cell = box(sheet, pos, 5, value, input_fill if label == "Merma" else None)
            if pos in (12,):
                cell.number_format = "0%"
            if pos == 14:
                cell.number_format = money

        sheet.cell(row=17, column=1, value="PARAMETROS").font = hf
        sheet.cell(row=17, column=1).fill = section_fill
        finish_coats = {
            "painting": 1 if additional_services.get("painting") else 0,
            "varnishing": 3 if additional_services.get("varnishing") else 0,
            "polishing": 5 if additional_services.get("polishing") else 0,
        }
        params = [
            ("Costo dia MO", 2300),
            ("Horas por dia", 8),
            ("Costo hora MO", "=B18/B19"),
            ("Seteos promedio", setup_average),
            ("Maquinaria", 0.10),
            ("IVA", IVA_RATE - 1),
            ("Recargo modelo no listado", surcharge),
            ("Umbral escala varillas", SCALE_THRESHOLD_VARILLAS),
            ("Recargo corte/metro especial", 0),
            ("Manos pintura", finish_coats["painting"]),
            ("Manos barniz", finish_coats["varnishing"]),
            ("Manos lustre", finish_coats["polishing"]),
            ("Minutos por m2 por mano", 15),
            ("Rinde litro m2", 8),
            ("Precio terminacion litro UYU", 0),
        ]
        for offset, (label, value) in enumerate(params, 18):
            box(sheet, offset, 1, label)
            cell = box(sheet, offset, 2, value, input_fill if not isinstance(value, str) else None)
            if label in {"Maquinaria", "IVA", "Recargo modelo no listado", "Recargo corte/metro especial"}:
                cell.number_format = "0%"
            if label.startswith("Costo"):
                cell.number_format = money
            if label == "Precio terminacion litro UYU":
                cell.number_format = money

        sheet.cell(row=17, column=4, value="COSTOS POR TIPO").font = hf
        sheet.cell(row=17, column=4).fill = section_fill
        hdr(sheet, 18, ["Tipo", "Minutos MO", "Seteo dias", "Ganancia"], start_col=4)
        for offset, fam in enumerate(("Listones / Tablas", "Molduras", "Barrotes"), 19):
            fam_cost = _family_cost(fam, material)
            values = [fam, fam_cost.minutes, fam_cost.setup_days, fam_cost.profit_percent]
            for col, value in enumerate(values, 4):
                cell = box(sheet, offset, col, value, input_fill if fam == group else None)
                if col == 7:
                    cell.number_format = "0%"

        product_row = 37
        sheet.cell(row=product_row - 2, column=1, value="PRODUCTO").font = hf
        sheet.cell(row=product_row - 2, column=1).fill = section_fill
        hdr(sheet, product_row - 1, [
            "Producto", "Tipo", "MP x varilla", "MO min", "MO x varilla", "Seteo x varilla",
            "Maquinaria", "Costo total", "% Ganancia", "Ganancia", "Precio sin IVA",
            "Manos terminacion", "Area 1 mano m2", "Area total m2", "Litros terminacion",
            "MO terminacion", "Insumos terminacion", "Unitario sin IVA", "IVA unitario",
            "Unitario + IVA", "Total sin IVA", "IVA total", "Total + IVA",
        ])
        product_name = f"{item.family} {item.width_mm:g}x{item.height_mm:g} mm"
        values = [
            product_name,
            group,
            "=E14",
            f'=VLOOKUP(B{product_row},$D$19:$G$21,2,FALSE)',
            f"=(D{product_row}/60)*$B$20",
            f'=(VLOOKUP(B{product_row},$D$19:$G$21,3,FALSE)*$B$18)/$B$21',
            f"=(C{product_row}+E{product_row})*$B$22",
            f"=C{product_row}+E{product_row}+F{product_row}+G{product_row}",
            f'=VLOOKUP(B{product_row},$D$19:$G$21,4,FALSE)',
            f"=H{product_row}*I{product_row}",
            f"=(H{product_row}+J{product_row})*(1+$B$24)*($B$7/{VARILLA_LENGTH_M})",
            "=$B$27+$B$28+$B$29",
            f"=2*($B$5+$B$6)/100*$B$7",
            f"=L{product_row}*M{product_row}",
            f"=N{product_row}/$B$31",
            f"=N{product_row}*$B$30/60*$B$20",
            f"=O{product_row}*$B$32",
            f'=IF($B$9="metro",K{product_row}/$B$7*(1+$B$26),K{product_row}*(1+$B$26))+P{product_row}+Q{product_row}',
            f"=R{product_row}*$B$23",
            f"=R{product_row}+S{product_row}",
            f"=R{product_row}*$B$8",
            f"=S{product_row}*$B$8",
            f"=T{product_row}*$B$8",
        ]
        for col, value in enumerate(values, 1):
            cell = box(sheet, product_row, col, value)
            if col in (3, 5, 6, 7, 8, 10, 11, 16, 17, 18, 19, 20, 21, 22, 23):
                cell.number_format = money
            if col in (9,):
                cell.number_format = "0%"
            if col in (13, 14, 15):
                cell.number_format = "0.00"
        for col in (18, 19, 20, 21, 22, 23):
            sheet.cell(row=product_row, column=col).fill = total_fill
            sheet.cell(row=product_row, column=col).font = hf

        sheet.cell(row=product_row + 3, column=1, value="NOTAS").font = hf
        notes = [
            "Unitario modelo muestra el precio que surge del conversor para controlar el razonamiento.",
            "El total usa el unitario modelo; no hay columna de precio cotizado duplicada.",
            "El IVA se calcula al final: unitario sin IVA, IVA unitario, unitario + IVA y totales separados.",
            "No se aplica recargo 20% por defecto por cambios de largo o cortes; consultar y completar B26 si corresponde.",
            "Pintura suma 1 mano; barniz suma 3 manos; lustre suma 5 manos.",
            "El insumo de terminacion se calcula por area: 1 litro cada 8 m2. La mano de obra usa 15 min por m2 por mano.",
            f"El largo solicitado ({requested_length_m:g} m) se calcula proporcional al precio/listado base de {VARILLA_LENGTH_M:g} m.",
            "Si supera 20 varillas, revisar opcion de escala.",
        ]
        if q.estimated:
            notes.insert(0, "No disponemos de esa moldura en stock/listado; este es precio estimativo con conversor.")
        else:
            notes.insert(0, "Producto encontrado en listado: se usa como referencia; el total se arma con el modelo/conversor editable.")
        for offset, note in enumerate(notes, product_row + 4):
            sheet.cell(row=offset, column=1, value=note)
            sheet.merge_cells(start_row=offset, start_column=1, end_row=offset, end_column=8)

        state_row = product_row + len(notes) + 7
        sheet.cell(row=state_row, column=1, value="CONDICIONES LICITACION ESTATAL / SERVICIOS").font = hf
        sheet.cell(row=state_row, column=1).fill = section_fill
        shipping_unit = _default_shipping_unit(destination, int(float(q.quantity or 1)))
        state_lines = [
            ("Lugar de entrega", destination),
            ("Plazo de entrega", f"{delivery_days} dias" if delivery_days else "A confirmar en pliego"),
            ("Plazo de pago", f"{payment_days} dias"),
            ("% recargo financiero", payment_pct),
            ("Garantia fiel cumplimiento", guarantee_text or "No indicada"),
            ("% garantia aplicado", guarantee_pct),
            ("Flete/descarga unitario", shipping_unit),
            ("Rectificacion de medidas", "Incluida" if additional_services.get("rectification") else "No incluida"),
            ("Colocacion", "Incluida" if additional_services.get("installation") else "No incluida"),
            ("Servicios terminacion", _additional_services_text(additional_services)),
            ("Mantenimiento oferta", f"{general_specs.get('offer_maintenance_days')} dias" if general_specs.get("offer_maintenance_days") else ""),
            ("Otras condiciones", general_specs.get("other_conditions") or ""),
        ]
        for offset, (label, value) in enumerate(state_lines, state_row + 1):
            box(sheet, offset, 1, label)
            cell = box(sheet, offset, 2, value, input_fill if label.startswith("%") or "Flete" in label else None)
            if label.startswith("%"):
                cell.number_format = "0%"
            if "Flete" in label:
                cell.number_format = money
        extras_row = state_row + len(state_lines) + 3
        extras = [
            ("Total productos + IVA", f"=W{product_row}"),
            ("Recargo financiero", f"=W{product_row}*B{state_row + 4}"),
            ("Garantia fiel cumplimiento", f"=W{product_row}*B{state_row + 6}"),
            ("Flete/descarga total", f"=B{state_row + 7}*$B$8"),
            ("TOTAL LICITACION", f"=B{extras_row}+B{extras_row+1}+B{extras_row+2}+B{extras_row+3}"),
        ]
        for offset, (label, value) in enumerate(extras, extras_row):
            box(sheet, offset, 1, label, total_fill if "TOTAL" in label else None, bold=True)
            cell = box(sheet, offset, 2, value, total_fill if "TOTAL" in label else None, bold=True)
            cell.number_format = money

        for col in range(1, 24):
            letter = openpyxl.utils.get_column_letter(col)
            sheet.column_dimensions[letter].width = 18
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["D"].width = 22
        sheet.freeze_panes = "A36"

        return {
            "unit_without_iva_cell": f"R{product_row}",
            "unit_iva_cell": f"S{product_row}",
            "unit_cell": f"T{product_row}",
            "total_without_iva_cell": f"U{product_row}",
            "total_iva_cell": f"V{product_row}",
            "total_cell": f"W{product_row}",
        }

    ws.cell(row=1, column=1, value="Cotizacion de molduras").font = title_font
    ws.cell(row=2, column=1, value="Resumen de items. Las condiciones y el armado editable quedan en la segunda solapa/conversor.").font = small_font
    hdr(ws, 4, [
        "Codigo", "Tipo", "Descripcion", "Medida", "Material", "Cant.", "Unidad",
        "Unitario sin IVA", "IVA unitario", "Unitario + IVA",
        "Total sin IVA", "IVA total", "Total + IVA", "Hoja", "Nota",
    ])

    total_without_iva_refs = []
    total_iva_refs = []
    total_refs = []
    for idx, raw in enumerate(requests, 1):
        q = quote_price(
            width_mm=float(raw.get("width_mm") or raw.get("width") or 0),
            height_mm=float(raw.get("height_mm") or raw.get("height") or 0),
            quantity=float(raw.get("quantity") or 1),
            material=raw.get("material"),
            family=raw.get("family"),
            unit=str(raw.get("unit") or "varilla"),
            include_iva=bool(raw.get("include_iva", True)),
        )
        r = idx + 4
        if q is None:
            ws.cell(row=r, column=1, value="SIN REFERENCIA").fill = warn_fill
            ws.cell(row=r, column=11, value="No se encontro referencia suficiente para estimar").fill = warn_fill
            continue

        item = q.item
        kind_label = "EUCA" if _material_kind(material_hint(raw, item)) == "euca" else "PN"
        conv = wb.create_sheet(sheet_safe_name(f"Conversor {kind_label}", idx))
        refs = write_conversor_sheet(conv, q, raw, idx)
        measure = f"{item.width_mm:g}x{item.height_mm:g} mm"
        note_parts = []
        if q.estimated:
            note_parts.append("No stock/listado: precio estimativo")
        if q.scale_hint:
            note_parts.append("Revisar escala")
        total_without_iva_refs.append(f"K{r}")
        total_iva_refs.append(f"L{r}")
        total_refs.append(f"M{r}")
        unit_without_iva_expr = f"{quoted_sheet(conv.title)}!{refs['unit_without_iva_cell']}"
        unit_iva_expr = f"{quoted_sheet(conv.title)}!{refs['unit_iva_cell']}"
        unit_expr = f"{quoted_sheet(conv.title)}!{refs['unit_cell']}"
        total_without_iva_expr = f"{quoted_sheet(conv.title)}!{refs['total_without_iva_cell']}"
        total_iva_expr = f"{quoted_sheet(conv.title)}!{refs['total_iva_cell']}"
        total_expr = f"{quoted_sheet(conv.title)}!{refs['total_cell']}"
        length_m = float(raw.get("length_m") or raw.get("requested_length_m") or VARILLA_LENGTH_M)
        if abs(length_m - VARILLA_LENGTH_M) > 0.001:
            note_parts.append(f"Largo solicitado {length_m:g} m")
            note_parts.append("Sin recargo 20% automatico")
        unit_without_iva_ref = f"={unit_without_iva_expr}"
        unit_iva_ref = f"={unit_iva_expr}"
        unit_ref = f"={unit_expr}"
        total_without_iva_ref = f"={total_without_iva_expr}"
        total_iva_ref = f"={total_iva_expr}"
        total_ref = f"={total_expr}"
        note = " | ".join(note_parts)
        values = [
            item.code, item.family, item.description, measure,
            raw.get("material") or "", q.quantity, q.unit,
            unit_without_iva_ref, unit_iva_ref, unit_ref,
            total_without_iva_ref, total_iva_ref, total_ref,
            conv.title, note,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.border = bdr
            if col in (8, 9, 10, 11, 12, 13):
                cell.number_format = money
            if note and col == 15:
                cell.fill = warn_fill

    total_row = len(requests) + 6
    ws.cell(row=total_row, column=10, value="Total sin IVA").font = hf
    ws.cell(row=total_row, column=11, value=f"=SUM({','.join(total_without_iva_refs)})" if total_without_iva_refs else 0).font = hf
    ws.cell(row=total_row, column=11).number_format = money
    ws.cell(row=total_row + 1, column=10, value="IVA").font = hf
    ws.cell(row=total_row + 1, column=11, value=f"=SUM({','.join(total_iva_refs)})" if total_iva_refs else 0).font = hf
    ws.cell(row=total_row + 1, column=11).number_format = money
    ws.cell(row=total_row + 2, column=10, value="Total + IVA").font = hf
    ws.cell(row=total_row + 2, column=11, value=f"=SUM({','.join(total_refs)})" if total_refs else 0).font = hf
    ws.cell(row=total_row + 2, column=11).number_format = money
    ws.cell(row=total_row + 2, column=10).fill = total_fill
    ws.cell(row=total_row + 2, column=11).fill = total_fill

    for col in range(1, 16):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["N"].width = 18
    ws.column_dimensions["O"].width = 36

    fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="cotizacion_molduras_")
    os.close(fd)
    wb.save(path)
    for image_path in image_paths:
        try:
            os.unlink(image_path)
        except OSError:
            pass
    return {"excel_path": path}


def handle_export_molduras_excel_session(data: dict) -> dict:
    from carpinteria.quotation_session import get_session

    sid = str(data.get("session_id") or "")
    if not sid:
        return {"error": "missing session_id"}
    s = get_session(sid)
    if s is None:
        return {"error": "session not found"}
    items = []
    for q in s.moldura_quotes:
        length_m = (q.breakdown or {}).get("largo_solicitado_m")
        items.append({
            "width_mm": q.width_mm,
            "height_mm": q.height_mm,
            "quantity": q.quantity,
            "material": q.material,
            "family": q.family,
            "unit": q.unit,
            "include_iva": q.iva_included,
            "length_m": length_m,
        })
    if not items:
        return {"error": "No hay cotizaciones de molduras para exportar"}
    general_specs = s.general_specs.model_dump() if s.general_specs else {}
    return handle_export_molduras_excel({
        "items": items,
        "commercial": {
            "payment_days": _effective_payment_days(s.payment_days, str(general_specs.get("payment_terms") or "")),
            "destination": _effective_destination(s.destination or str(general_specs.get("delivery_location") or "")),
        },
        "general_specs": general_specs,
        "additional_services": s.additional_services.model_dump() if s.additional_services else {},
    })


def handle_export_docx(data: dict) -> dict:
    import subprocess
    import tempfile

    fd_in, json_path = tempfile.mkstemp(suffix=".json", prefix="docx_input_")
    fd_out, docx_path = tempfile.mkstemp(suffix=".docx", prefix="cotizacion_")

    with open(json_path, "w") as f:
        json.dump(data, f, ensure_ascii=False)

    script = os.path.join(os.path.dirname(__file__), "generate_docx.js")
    result = subprocess.run(
        ["node", script, json_path, docx_path],
        capture_output=True, text=True, timeout=30,
    )

    os.unlink(json_path)

    if result.returncode != 0:
        return {"error": f"Error generando Word: {result.stderr}"}

    return {"docx_path": docx_path}


def handle_lista_precios_preview(data: dict) -> dict:
    from dataclasses import asdict
    from carpinteria.lista_precios_parser import parse_pdf
    from carpinteria.lista_precios_sheets import read_activa
    from carpinteria.lista_precios_diff import compute_diff

    pdf_path = data["pdf_path"]
    sheet_id = data.get("sheet_id")
    tc, tc_source = _get_tc()

    items = parse_pdf(pdf_path, tc=tc)
    current_rows = read_activa(sheet_id)
    diff = compute_diff(items, current_rows)

    return {
        "lista": items[0].lista if items else "",
        "periodo": items[0].periodo if items else "",
        "current_lista": diff["current_lista"],
        "current_periodo": diff["current_periodo"],
        "tc": tc,
        "tc_source": f"BCU {tc_source}",
        "summary": diff["summary"],
        "nuevos": diff["nuevos"],
        "removidos": diff["removidos"],
        "cambios": diff["cambios"],
        "items": [asdict(it) for it in items],
    }


def handle_lista_precios_confirm(data: dict) -> dict:
    from carpinteria.lista_precios_sheets import items_from_dicts, write_items

    items = items_from_dicts(data.get("items") or [])
    if not items:
        return {"error": "no items in payload"}
    return write_items(items, sheet_id=data.get("sheet_id"))


def handle_hardware_prices_get(data: dict) -> dict:
    from carpinteria.hardware_prices_sheet import read_all
    rows = read_all(sheet_id=data.get("sheet_id"))
    return {"rows": list(rows.values())}


def handle_hardware_prices_set(data: dict) -> dict:
    from carpinteria.hardware_prices_sheet import upsert_price
    code = (data.get("code") or "").strip()
    if not code:
        return {"error": "missing code"}
    try:
        price = float(data.get("price") or 0)
    except (TypeError, ValueError):
        return {"error": "invalid price"}
    if price < 0:
        return {"error": "price must be >= 0"}
    try:
        row = upsert_price(
            code,
            price,
            updated_by=str(data.get("updated_by") or ""),
            sheet_id=data.get("sheet_id"),
        )
    except ValueError as e:
        return {"error": str(e)}
    return {"row": row}


# ---------------------------------------------------------------------------
# Chat / sessions
# ---------------------------------------------------------------------------

def handle_session_create(data: dict) -> dict:
    from carpinteria.quotation_session import create_session, ensure_indexes
    ensure_indexes()
    s = create_session(
        user_id=str(data.get("user_id") or "anonymous"),
        title=str(data.get("title") or ""),
        brand_id=str(data.get("brand_id") or "casa"),
        request_area=str(data.get("area") or "personal"),
    )
    return {"session": s.model_dump(mode="json")}


def handle_session_get(data: dict) -> dict:
    from carpinteria.quotation_session import get_session
    s = get_session(str(data.get("session_id") or ""))
    if s is None:
        return {"error": "session not found"}
    return {"session": s.model_dump(mode="json")}


def handle_session_list(data: dict) -> dict:
    from carpinteria.quotation_session import current_year_month, list_sessions
    year = data.get("year")
    month = data.get("month")
    if data.get("current_month", False) and not year and not month:
        year, month = current_year_month()
    rows = list_sessions(
        user_id=data.get("user_id") if data.get("user_id") else None,
        brand_id=data.get("brand_id") if data.get("brand_id") else None,
        area=data.get("area") if data.get("area") else None,
        limit=int(data.get("limit") or 30),
        year=int(year) if year else None,
        month=int(month) if month else None,
    )
    return {"sessions": rows}


def handle_session_archive(data: dict) -> dict:
    from carpinteria.quotation_session import list_session_archive
    months = list_session_archive(
        user_id=data.get("user_id") if data.get("user_id") else None,
        brand_id=data.get("brand_id") if data.get("brand_id") else None,
        area=data.get("area") if data.get("area") else None,
        year=int(data.get("year") or 2026),
    )
    return {"months": months}


def handle_session_ingest_pliego(data: dict) -> dict:
    """Direct ingest: upload + decompose without going through the chat agent."""
    from carpinteria.agents.cotizador_chat import _ingest_pliego_into_session
    from carpinteria.quotation_session import append_message, get_session

    sid = str(data.get("session_id") or "")
    file_paths = list(data.get("file_paths") or [])
    if not sid or not file_paths:
        return {"error": "missing session_id or file_paths"}

    summary = _ingest_pliego_into_session(sid, file_paths)
    # Direct (non-chat) ingest: surface the summary in the conversation so the
    # UI shows it after reload. The chat path handles its own persistence.
    append_message(sid, "assistant", summary)
    s = get_session(sid)
    return {"summary": summary, "session": s.model_dump(mode="json") if s else None}


def _recalc_all_items(session) -> None:
    """Reuse one catalog/TC/hw_prices fetch across every item recalculation.

    Direct-mutation handlers below all need this same boilerplate, so it lives
    here instead of being duplicated.
    """
    from carpinteria.agents.cotizador_chat import _recalculate_item, _hw_prices_map, _tc
    from carpinteria.catalog import ProductCatalog
    catalog = ProductCatalog.from_activa()
    tc = _tc()
    hw_prices = _hw_prices_map()
    for it in session.items:
        try:
            _recalculate_item(it, session, catalog=catalog, tc=tc, hw_prices=hw_prices)
        except Exception:
            pass


def handle_session_update(data: dict) -> dict:
    """Patch session-level fields and recalculate every item.

    Accepts any subset of: color_default, payment_days, destination, additional_services.
    Empty strings/None clear the field.
    """
    from carpinteria.quotation_session import get_session, save_session
    sid = str(data.get("session_id") or "")
    if not sid:
        return {"error": "missing session_id"}
    s = get_session(sid)
    if s is None:
        return {"error": "session not found"}

    if "color_default" in data:
        s.color_default = str(data.get("color_default") or "")
    if "payment_days" in data:
        v = data.get("payment_days")
        s.payment_days = int(v) if v not in (None, "", 0) else None
    if "destination" in data:
        s.destination = str(data.get("destination") or "")
    if "title" in data:
        s.title = str(data.get("title") or "").strip()
    if "additional_services" in data:
        raw = data.get("additional_services") or {}
        current = s.additional_services.model_dump() if s.additional_services else {}
        current.update({key: bool(raw.get(key)) for key in ADDITIONAL_SERVICE_LABELS if key in raw})
        s.additional_services = type(s.additional_services)(**current)

    _recalc_all_items(s)
    save_session(s)
    return {"session": s.model_dump(mode="json")}


def handle_session_delete(data: dict) -> dict:
    """Hard-delete a session document. Used by the sidebar's «eliminar» menu."""
    from carpinteria.quotation_session import COLLECTION
    from carpinteria.db import collection

    sid = str(data.get("session_id") or "")
    if not sid:
        return {"error": "missing session_id"}
    res = collection(COLLECTION).delete_one({"id": sid})
    return {"deleted": res.deleted_count > 0}


def handle_session_commercial_status(data: dict) -> dict:
    from carpinteria.quotation_session import update_commercial_status
    sid = str(data.get("session_id") or "")
    if not sid:
        return {"error": "missing session_id"}
    fields = {
        key: data[key]
        for key in (
            "approval_status",
            "client_sent",
            "client_accepted",
            "deposit_amount",
            "order_number",
            "ready_to_deliver",
            "delivered",
            "final_payment_amount",
        )
        if key in data
    }
    try:
        s = update_commercial_status(sid, fields)
    except ValueError as e:
        return {"error": str(e)}
    if s is None:
        return {"error": "session not found"}
    return {"session": s.model_dump(mode="json")}


def handle_session_approval(data: dict) -> dict:
    return handle_session_commercial_status(data)


def handle_set_item_placa(data: dict) -> dict:
    """Pin (or unpin) a specific catalog placa SKU on a quotation item."""
    from carpinteria.quotation_session import find_item, get_session, save_session
    sid = str(data.get("session_id") or "")
    code = str(data.get("item_code") or "")
    sku = data.get("placa_sku")  # None to clear
    if not sid or not code:
        return {"error": "missing session_id or item_code"}
    s = get_session(sid)
    if s is None:
        return {"error": "session not found"}
    it = find_item(s, code)
    if it is None:
        return {"error": f"item {code} not found"}
    it.placa_sku = (str(sku) if sku else None)

    from carpinteria.agents.cotizador_chat import _recalculate_item
    _recalculate_item(it, s)
    save_session(s)
    return {"session": s.model_dump(mode="json")}


def handle_item_update(data: dict) -> dict:
    """Patch one or more fields on a single quotation item, then recalc.

    Whitelist of editable fields: color, material, thickness_mm, quantity, name,
    edge_banding. Other fields (pieces, hardware, last_quote) have dedicated
    handlers because they need stricter typing.
    """
    from carpinteria.agents.cotizador_chat import _recalculate_item
    from carpinteria.quotation_session import find_item, get_session, save_session

    sid = str(data.get("session_id") or "")
    code = str(data.get("item_code") or "")
    if not sid or not code:
        return {"error": "missing session_id or item_code"}
    s = get_session(sid)
    if s is None:
        return {"error": "session not found"}
    it = find_item(s, code)
    if it is None:
        return {"error": f"item {code} not found"}

    fields = data.get("fields") or {}
    if "color" in fields:
        it.color = str(fields["color"] or "")
    if "material" in fields:
        it.material = str(fields["material"] or "")
    if "thickness_mm" in fields:
        v = fields["thickness_mm"]
        try:
            it.thickness_mm = float(v) if v not in (None, "") else 18.0
        except (TypeError, ValueError):
            return {"error": "thickness_mm must be a number"}
    if "quantity" in fields:
        v = fields["quantity"]
        try:
            it.quantity = max(1, int(v))
        except (TypeError, ValueError):
            return {"error": "quantity must be an integer"}
    if "name" in fields:
        it.name = str(fields["name"] or "")
    if "edge_banding" in fields:
        it.edge_banding = str(fields["edge_banding"] or "")
    if "notes" in fields:
        it.notes = str(fields["notes"] or "")

    _recalculate_item(it, s)
    save_session(s)
    return {"session": s.model_dump(mode="json")}


def handle_item_delete(data: dict) -> dict:
    from carpinteria.quotation_session import get_session, save_session

    sid = str(data.get("session_id") or "")
    code = str(data.get("item_code") or "")
    if not sid or not code:
        return {"error": "missing session_id or item_code"}
    s = get_session(sid)
    if s is None:
        return {"error": "session not found"}
    before = len(s.items)
    s.items = [it for it in s.items if it.code.lower() != code.lower()]
    if len(s.items) == before:
        return {"error": f"item {code} not found"}
    save_session(s)
    return {"session": s.model_dump(mode="json")}


def handle_piece_set_quantity(data: dict) -> dict:
    """Update one piece's quantity inside an item (matched by label)."""
    from carpinteria.agents.cotizador_chat import _recalculate_item
    from carpinteria.quotation_session import find_item, get_session, save_session

    sid = str(data.get("session_id") or "")
    code = str(data.get("item_code") or "")
    label = str(data.get("piece_label") or "")
    qty_raw = data.get("quantity")
    if not sid or not code or not label or qty_raw is None:
        return {"error": "missing session_id, item_code, piece_label or quantity"}
    s = get_session(sid)
    if s is None:
        return {"error": "session not found"}
    it = find_item(s, code)
    if it is None:
        return {"error": f"item {code} not found"}
    label_l = label.strip().lower()
    target = next((p for p in it.pieces if p.label.lower() == label_l), None)
    if target is None:
        return {"error": f"piece '{label}' not found"}
    try:
        qty = max(0, int(qty_raw))
    except (TypeError, ValueError):
        return {"error": "quantity must be an integer"}
    if qty == 0:
        it.pieces = [p for p in it.pieces if p is not target]
    else:
        target.quantity = qty
    _recalculate_item(it, s)
    save_session(s)
    return {"session": s.model_dump(mode="json")}


def handle_piece_upsert(data: dict) -> dict:
    """Create or replace a cut piece by label, then recalc."""
    from carpinteria.agents.cotizador_chat import _recalculate_item
    from carpinteria.quotation_session import CutPiece, find_item, get_session, save_session

    sid = str(data.get("session_id") or "")
    code = str(data.get("item_code") or "")
    label = str(data.get("piece", {}).get("label") or "")
    if not sid or not code or not label:
        return {"error": "missing session_id, item_code or piece.label"}
    s = get_session(sid)
    if s is None:
        return {"error": "session not found"}
    it = find_item(s, code)
    if it is None:
        return {"error": f"item {code} not found"}
    raw = data.get("piece") or {}
    try:
        piece = CutPiece(
            width_mm=float(raw.get("width_mm") or 0),
            height_mm=float(raw.get("height_mm") or 0),
            quantity=max(1, int(raw.get("quantity") or 1)),
            label=label,
            edge_sides=list(raw.get("edge_sides") or []),
        )
    except (TypeError, ValueError):
        return {"error": "invalid piece dimensions or quantity"}
    label_l = label.strip().lower()
    idx = next((i for i, p in enumerate(it.pieces) if p.label.strip().lower() == label_l), None)
    if idx is None:
        it.pieces.append(piece)
    else:
        it.pieces[idx] = piece
    _recalculate_item(it, s)
    save_session(s)
    return {"session": s.model_dump(mode="json")}


def handle_hardware_set_quantity(data: dict) -> dict:
    """Update one hardware row's quantity inside an item. qty=0 removes it."""
    from carpinteria.agents.cotizador_chat import _recalculate_item
    from carpinteria.hardware_catalog import get_by_code
    from carpinteria.quotation_session import (
        HardwareUsage,
        find_item,
        get_session,
        save_session,
    )

    sid = str(data.get("session_id") or "")
    code = str(data.get("item_code") or "")
    hw_code = str(data.get("hardware_code") or "")
    qty_raw = data.get("quantity")
    if not sid or not code or not hw_code or qty_raw is None:
        return {"error": "missing session_id, item_code, hardware_code or quantity"}
    s = get_session(sid)
    if s is None:
        return {"error": "session not found"}
    it = find_item(s, code)
    if it is None:
        return {"error": f"item {code} not found"}
    spec = get_by_code(hw_code)
    if spec is None:
        return {"error": f"hardware {hw_code} is not in the curated catalog"}
    try:
        qty = max(0, int(qty_raw))
    except (TypeError, ValueError):
        return {"error": "quantity must be an integer"}

    existing = next((h for h in it.hardware if h.code == spec.code), None)
    if qty == 0:
        it.hardware = [h for h in it.hardware if h.code != spec.code]
    elif existing is None:
        it.hardware.append(
            HardwareUsage(
                code=spec.code,
                name=spec.name,
                category=spec.category,
                unit=spec.unit,
                quantity=qty,
            )
        )
    else:
        existing.quantity = qty
    _recalculate_item(it, s)
    save_session(s)
    return {"session": s.model_dump(mode="json")}


def handle_hardware_catalog_list(_data: dict) -> dict:
    """Return the curated hardware catalog (codes + display names)."""
    from carpinteria.hardware_catalog import CURATED_HARDWARE
    rows = [
        {"code": h.code, "name": h.name, "category": h.category, "unit": h.unit}
        for h in CURATED_HARDWARE
    ]
    rows.sort(key=lambda r: (r["category"], r["name"]))
    return {"hardware": rows}


def _session_to_quotes_payload(session) -> list[dict]:
    """Translate a QuotationSession's items into the legacy `quotes` shape that
    `handle_export_excel` / `handle_export_docx` were originally built for.

    Keeps the export handlers untouched — they pre-date the session model and
    expect items to already be denormalised.
    """
    quotes: list[dict] = []
    for it in session.items:
        last = it.last_quote or {}
        total = float(last.get("total", 0) or 0)
        pending = list(last.get("pending_hardware_codes") or [])
        has_error = total == 0
        # When there's a pending hardware price but the quote otherwise computed,
        # we still want the item rendered (don't mark has_error). The exporters
        # handle missing_inputs vs notes vs has_error independently.
        missing = []
        notes = str(last.get("notes") or "")
        if has_error and notes:
            missing = [notes]

        quotes.append({
            "item_code": it.code,
            "item_name": it.name,
            "item_quantity": it.quantity,
            "dimensions": dict(it.dimensions),
            "has_error": has_error,
            "missing_inputs": missing,
            "notes": notes,
            "pending_hardware_codes": pending,
            "decomposition": {
                "pieces": [p.model_dump() for p in it.pieces if not _is_drawer_base_piece(p.model_dump())],
            },
            "lines": list(last.get("lines") or []),
            "hardware_lines": list(last.get("hardware_lines") or []),
            "material": it.material,
            "color": it.color or session.color_default,
            "thickness_mm": it.thickness_mm,
            "edge_banding": it.edge_banding,
            "subtotal": float(last.get("subtotal", 0) or 0),
            "margin_percent": float(last.get("margin_percent", 0) or 0),
            "margin_amount": float(last.get("margin_amount", 0) or 0),
            "total": total,
            "total_with_hardware": float(last.get("total_with_hardware", 0) or 0),
            "_tc": float(last.get("tc", 40) or 40),
            "_payment_days": _effective_payment_days(session.payment_days, session.general_specs.payment_terms),
            "_payment_terms": session.general_specs.payment_terms,
            "_destination": _effective_destination(session.destination or session.general_specs.delivery_location),
            "metadata": dict(last.get("metadata") or {}),
        })
    return quotes


def handle_export_excel_session(data: dict) -> dict:
    from carpinteria.quotation_session import get_session, save_session
    sid = str(data.get("session_id") or "")
    if not sid:
        return {"error": "missing session_id"}
    s = get_session(sid)
    if s is None:
        return {"error": "session not found"}
    try:
        _recalc_all_items(s)
        save_session(s)
    except Exception:
        pass
    return handle_export_excel({
        "quotes": _session_to_quotes_payload(s),
        "additional_services": s.additional_services.model_dump() if s.additional_services else {},
    })


def handle_export_docx_session(data: dict) -> dict:
    from carpinteria.quotation_session import get_session
    sid = str(data.get("session_id") or "")
    if not sid:
        return {"error": "missing session_id"}
    s = get_session(sid)
    if s is None:
        return {"error": "session not found"}
    return handle_export_docx({"quotes": _session_to_quotes_payload(s)})


def handle_catalog_list_boards(_data: dict) -> dict:
    """Return all PLACA rows from the Activa catalog with a stable id (sku)
    and a human-readable label, ready to feed a frontend dropdown."""
    from carpinteria.catalog import ProductCatalog
    cat = ProductCatalog.from_activa()
    rows: list[dict] = []
    for p in cat.filter(tipo_producto="PLACA"):
        thickness = f"{p.espesor_mm:.0f}mm" if p.espesor_mm else "?mm"
        size = ""
        if p.ancho_mm and p.largo_mm:
            size = f" {p.largo_mm/1000:.2f}×{p.ancho_mm/1000:.2f}m"
        label = f"{p.material or p.familia} {thickness}{size} — {p.nombre}"
        rows.append({
            "sku": p.sku,
            "label": label,
            "familia": p.familia,
            "material": p.material,
            "espesor_mm": p.espesor_mm,
            "precio_usd": p.precio_usd_simp,
        })
    rows.sort(key=lambda r: (r["familia"] or "", r["espesor_mm"] or 0, r["label"]))
    return {"boards": rows}


def handle_chat(data: dict) -> dict:
    import asyncio
    from carpinteria.agents.cotizador_chat import run_turn

    sid = str(data.get("session_id") or "")
    message = str(data.get("message") or "")
    if not sid or not message:
        return {"error": "missing session_id or message"}

    return asyncio.run(run_turn(sid, message))


def handle_chat_stream(data: dict) -> None:
    """Stream-emit NDJSON events to stdout instead of returning a dict.

    Each line is a JSON object `{type, ...}`. The Next route reads this
    line-by-line from the subprocess stdout and forwards it as SSE.
    """
    import asyncio
    from carpinteria.agents.cotizador_chat import run_turn_stream

    sid = str(data.get("session_id") or "")
    message = str(data.get("message") or "")
    if not sid or not message:
        sys.stdout.write(json.dumps({"type": "error", "message": "missing session_id or message"}) + "\n")
        sys.stdout.flush()
        return

    async def consume() -> None:
        try:
            async for ev_type, payload in run_turn_stream(sid, message):
                line = json.dumps({"type": ev_type, **payload}, ensure_ascii=False, default=str)
                sys.stdout.write(line + "\n")
                sys.stdout.flush()
        except Exception as e:
            line = json.dumps({"type": "error", "message": str(e)}, ensure_ascii=False)
            sys.stdout.write(line + "\n")
            sys.stdout.flush()

    asyncio.run(consume())


# ---------------------------------------------------------------------------
# Memory (cross-session facts)
# ---------------------------------------------------------------------------

def handle_memory_list(_data: dict) -> dict:
    from carpinteria import memory as agent_memory
    agent_memory.ensure_indexes()
    facts = [f.model_dump(mode="json") for f in agent_memory.list_facts()]
    return {"facts": facts}


def handle_memory_add(data: dict) -> dict:
    from carpinteria import memory as agent_memory
    text = str(data.get("text") or "").strip()
    if not text:
        return {"error": "missing text"}
    tags_raw = data.get("tags") or []
    tags = [str(t) for t in tags_raw] if isinstance(tags_raw, list) else []
    fact = agent_memory.add_fact(text, tags)
    return {"fact": fact.model_dump(mode="json")}


def handle_memory_delete(data: dict) -> dict:
    from carpinteria import memory as agent_memory
    fid = str(data.get("id") or "")
    if not fid:
        return {"error": "missing id"}
    deleted = agent_memory.delete_fact(fid)
    return {"deleted": deleted}


def handle_auth_login(data: dict) -> dict:
    from carpinteria.auth_users import authenticate
    user = authenticate(
        username=str(data.get("user") or ""),
        password=str(data.get("password") or ""),
        brand_id=str(data.get("brand_id") or ""),
        area=str(data.get("area") or ""),
    )
    if user is None:
        return {"error": "Usuario o contraseña incorrectos"}
    return {"user": user}


def handle_auth_users_list(data: dict) -> dict:
    from carpinteria.auth_users import list_users
    return {"users": list_users(brand_id=data.get("brand_id") or None)}


def handle_auth_users_create(data: dict) -> dict:
    from carpinteria.auth_users import create_user
    try:
        user = create_user(
            username=str(data.get("username") or ""),
            password=str(data.get("password") or ""),
            brand_id=str(data.get("brand_id") or ""),
            area=str(data.get("area") or ""),
            must_change_password=bool(data.get("must_change_password", True)),
        )
    except ValueError as e:
        return {"error": str(e)}
    return {"user": user}


def handle_auth_password_reset_request(data: dict) -> dict:
    from carpinteria.auth_users import request_password_reset
    return request_password_reset(
        username=str(data.get("username") or ""),
        brand_id=str(data.get("brand_id") or ""),
        area=str(data.get("area") or ""),
    )


def handle_auth_password_reset_confirm(data: dict) -> dict:
    from carpinteria.auth_users import reset_password_with_code
    user = reset_password_with_code(
        username=str(data.get("username") or ""),
        brand_id=str(data.get("brand_id") or ""),
        area=str(data.get("area") or ""),
        code=str(data.get("code") or ""),
        password=str(data.get("password") or ""),
    )
    if user is None:
        return {"error": "codigo invalido o vencido"}
    return {"user": user}


def handle_auth_users_update(data: dict) -> dict:
    from carpinteria.auth_users import set_user_active, update_password
    user_id = str(data.get("user_id") or "")
    if not user_id:
        return {"error": "missing user_id"}
    try:
        if "password" in data and data.get("password"):
            user = update_password(user_id, str(data.get("password") or ""))
        elif "active" in data:
            user = set_user_active(user_id, bool(data.get("active")))
        else:
            return {"error": "nothing to update"}
    except ValueError as e:
        return {"error": str(e)}
    if user is None:
        return {"error": "user not found"}
    return {"user": user}


def main() -> None:
    raw = sys.stdin.read()
    data = json.loads(raw)
    action = data.get("action", "")

    # Streaming actions write their own NDJSON to stdout — don't wrap in
    # a single JSON response or json.dump at the end.
    if action == "chat_stream":
        handle_chat_stream(data)
        return

    try:
        if action == "prices":
            result = handle_prices()
        elif action == "quote":
            result = handle_quote(data)
        elif action == "analyze":
            result = handle_analyze(data)
        elif action == "analyze_pliego":
            result = handle_analyze_pliego(data)
        elif action == "quote_item":
            result = handle_quote_item(data)
        elif action == "export_excel":
            result = handle_export_excel(data)
        elif action == "export_molduras_excel":
            result = handle_export_molduras_excel(data)
        elif action == "export_molduras_excel_session":
            result = handle_export_molduras_excel_session(data)
        elif action == "export_docx":
            result = handle_export_docx(data)
        elif action == "lista_precios_preview":
            result = handle_lista_precios_preview(data)
        elif action == "lista_precios_confirm":
            result = handle_lista_precios_confirm(data)
        elif action == "hardware_prices_get":
            result = handle_hardware_prices_get(data)
        elif action == "hardware_prices_set":
            result = handle_hardware_prices_set(data)
        elif action == "session_create":
            result = handle_session_create(data)
        elif action == "session_get":
            result = handle_session_get(data)
        elif action == "session_list":
            result = handle_session_list(data)
        elif action == "session_archive":
            result = handle_session_archive(data)
        elif action == "session_ingest_pliego":
            result = handle_session_ingest_pliego(data)
        elif action == "chat":
            result = handle_chat(data)
        elif action == "memory_list":
            result = handle_memory_list(data)
        elif action == "memory_add":
            result = handle_memory_add(data)
        elif action == "memory_delete":
            result = handle_memory_delete(data)
        elif action == "auth_login":
            result = handle_auth_login(data)
        elif action == "auth_users_list":
            result = handle_auth_users_list(data)
        elif action == "auth_users_create":
            result = handle_auth_users_create(data)
        elif action == "auth_password_reset_request":
            result = handle_auth_password_reset_request(data)
        elif action == "auth_password_reset_confirm":
            result = handle_auth_password_reset_confirm(data)
        elif action == "auth_users_update":
            result = handle_auth_users_update(data)
        elif action == "session_update":
            result = handle_session_update(data)
        elif action == "session_delete":
            result = handle_session_delete(data)
        elif action == "session_approval":
            result = handle_session_approval(data)
        elif action == "session_commercial_status":
            result = handle_session_commercial_status(data)
        elif action == "set_item_placa":
            result = handle_set_item_placa(data)
        elif action == "catalog_list_boards":
            result = handle_catalog_list_boards(data)
        elif action == "item_update":
            result = handle_item_update(data)
        elif action == "item_delete":
            result = handle_item_delete(data)
        elif action == "piece_set_quantity":
            result = handle_piece_set_quantity(data)
        elif action == "piece_upsert":
            result = handle_piece_upsert(data)
        elif action == "hardware_set_quantity":
            result = handle_hardware_set_quantity(data)
        elif action == "hardware_catalog_list":
            result = handle_hardware_catalog_list(data)
        elif action == "export_excel_session":
            result = handle_export_excel_session(data)
        elif action == "export_docx_session":
            result = handle_export_docx_session(data)
        else:
            result = {"error": f"Unknown action: {action}"}
    except Exception as e:
        result = {"error": str(e)}

    json.dump(result, sys.stdout, ensure_ascii=False, default=str)


if __name__ == "__main__":
    main()
