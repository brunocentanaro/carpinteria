from __future__ import annotations

import unicodedata
from calendar import monthrange
from datetime import date, datetime, timezone
from pathlib import Path
from tempfile import gettempdir
from uuid import uuid4

from pymongo import ReturnDocument

from .db import collection
from .exchange_rate import fetch_bcu_accounting_usd


MOVEMENT_DIRECTIONS = {"income", "expense", "transfer"}
PAYMENT_METHODS = {
    "efectivo",
    "cheque",
    "credito",
    "deposito",
    "transferencia",
    "tarjeta",
    "visa",
    "master",
    "maestro",
    "mercadolibre",
    "otro",
}
INVOICE_STATUSES = {"pendiente", "parcial", "pagada", "no_aplica"}
CARD_PAYMENT_METHODS = {"tarjeta", "visa", "master", "maestro", "mercadolibre"}
BANK_PAYMENT_METHODS = {"deposito", "transferencia"}
CARD_SETTLEMENT_CATEGORY = "acreditacion_tarjeta"
CARD_PAYMENT_PLANS = {
    "debito": ("debito", 1),
    "credito_1": ("credito", 1),
    "credito_2": ("credito", 2),
    "credito_3": ("credito", 3),
}
# Medios de tarjeta que declara el cajero al entregar la caja. Master prepago se
# pliega sobre master_debito porque el POS lo liquida como debito.
CARD_CONCILIATION_MEDIOS = ("visa_debito", "visa_credito", "master_debito", "master_credito", "maestro")
# Diferencia POS vs caja por debajo de la cual no hay faltante/sobrante real:
# el cajero declara totales redondeados al peso, el POS reporta con centavos.
CARD_CONCILIATION_EPSILON = 0.5

CHART_OF_ACCOUNTS = [
    {"code": "1111", "name": "Caja", "class": "activo", "group": "Disponibilidades", "nature": "debit"},
    {"code": "1112", "name": "Cheques recibidos", "class": "activo", "group": "Disponibilidades", "nature": "debit"},
    {"code": "1113", "name": "Banco BROU", "class": "activo", "group": "Disponibilidades", "nature": "debit"},
    {"code": "1119", "name": "Otros bancos", "class": "activo", "group": "Disponibilidades", "nature": "debit"},
    {"code": "1120", "name": "Creditos por ventas", "class": "activo", "group": "Creditos", "nature": "debit"},
    {"code": "1131", "name": "Visa a cobrar", "class": "activo", "group": "Financieras", "nature": "debit"},
    {"code": "1132", "name": "Mastercard a cobrar", "class": "activo", "group": "Financieras", "nature": "debit"},
    {"code": "1139", "name": "Otras financieras a cobrar", "class": "activo", "group": "Financieras", "nature": "debit"},
    {"code": "1140", "name": "Mercaderias e inventarios", "class": "activo", "group": "Inventarios", "nature": "debit"},
    {"code": "1190", "name": "Otros activos corrientes", "class": "activo", "group": "Otros activos", "nature": "debit"},
    {"code": "1210", "name": "Propiedad, planta y equipo", "class": "activo", "group": "Activo no corriente", "nature": "debit"},
    {"code": "2110", "name": "Proveedores", "class": "pasivo", "group": "Deudas comerciales", "nature": "credit"},
    {"code": "2121", "name": "DGI a pagar", "class": "pasivo", "group": "Deuda fiscal", "nature": "credit"},
    {"code": "2122", "name": "BPS a pagar", "class": "pasivo", "group": "Deuda fiscal", "nature": "credit"},
    {"code": "2123", "name": "IMM a pagar", "class": "pasivo", "group": "Deuda fiscal", "nature": "credit"},
    {"code": "2129", "name": "Otros impuestos a pagar", "class": "pasivo", "group": "Deuda fiscal", "nature": "credit"},
    {"code": "2131", "name": "Sueldos a pagar", "class": "pasivo", "group": "Deudas laborales", "nature": "credit"},
    {"code": "2132", "name": "Provision para aguinaldo", "class": "pasivo", "group": "Provisiones laborales", "nature": "credit"},
    {"code": "2133", "name": "Provision para licencia", "class": "pasivo", "group": "Provisiones laborales", "nature": "credit"},
    {"code": "2134", "name": "Provision para salario vacacional", "class": "pasivo", "group": "Provisiones laborales", "nature": "credit"},
    {"code": "2140", "name": "Tarjetas a pagar", "class": "pasivo", "group": "Deudas financieras", "nature": "credit"},
    {"code": "2190", "name": "Otros pasivos corrientes", "class": "pasivo", "group": "Otros pasivos", "nature": "credit"},
    {"code": "2210", "name": "Pasivos no corrientes", "class": "pasivo", "group": "Pasivo no corriente", "nature": "credit"},
    {"code": "3110", "name": "Capital y aportes", "class": "patrimonio", "group": "Capital", "nature": "credit"},
    {"code": "3210", "name": "Reservas", "class": "patrimonio", "group": "Reservas", "nature": "credit"},
    {"code": "3310", "name": "Resultados acumulados", "class": "patrimonio", "group": "Resultados acumulados", "nature": "credit"},
    {"code": "3510", "name": "Retiros de propietarios", "class": "patrimonio", "group": "Retiros", "nature": "debit"},
    {"code": "4110", "name": "Ventas", "class": "ingreso", "group": "Ingresos operativos", "nature": "credit"},
    {"code": "4210", "name": "Otros ingresos", "class": "ingreso", "group": "Otros ingresos", "nature": "credit"},
    {"code": "4220", "name": "Ganancias por diferencia de cambio", "class": "ingreso", "group": "Resultados financieros", "nature": "credit"},
    {"code": "5110", "name": "Costo de ventas", "class": "gasto", "group": "Costo de ventas", "nature": "debit"},
    {"code": "5210", "name": "Sueldos y jornales", "class": "gasto", "group": "Gastos de personal", "nature": "debit"},
    {"code": "5220", "name": "Gasto de aguinaldo", "class": "gasto", "group": "Gastos de personal", "nature": "debit"},
    {"code": "5230", "name": "Gasto de licencia", "class": "gasto", "group": "Gastos de personal", "nature": "debit"},
    {"code": "5240", "name": "Gasto de salario vacacional", "class": "gasto", "group": "Gastos de personal", "nature": "debit"},
    {"code": "5310", "name": "Impuestos y tasas", "class": "gasto", "group": "Impuestos", "nature": "debit"},
    {"code": "5410", "name": "Servicios y costos fijos", "class": "gasto", "group": "Gastos operativos", "nature": "debit"},
    {"code": "5420", "name": "Marketing y publicidad", "class": "gasto", "group": "Gastos operativos", "nature": "debit"},
    {"code": "5510", "name": "Comisiones de tarjetas", "class": "gasto", "group": "Gastos financieros", "nature": "debit"},
    {"code": "5520", "name": "Perdidas por diferencia de cambio", "class": "gasto", "group": "Gastos financieros", "nature": "debit"},
    {"code": "5530", "name": "Mantenimiento de cuentas y comisiones bancarias", "class": "gasto", "group": "Gastos financieros", "nature": "debit"},
    {"code": "5910", "name": "Otros gastos y perdidas", "class": "gasto", "group": "Otros gastos", "nature": "debit"},
]
ACCOUNT_BY_CODE = {account["code"]: account for account in CHART_OF_ACCOUNTS}
SUPPLIER_CLASSIFICATION_ACCOUNTS = {
    "inventory": "1140",
    "cost_of_sales": "5110",
    "marketing": "5420",
    "services": "5410",
    "taxes": "5310",
    "property_plant_equipment": "1210",
    "other_asset": "1190",
    "other_expense": "5910",
    "bank_fees": "5530",
}


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _clean(value: object) -> str:
    return str(value or "").strip()


def _money(value: object) -> float:
    try:
        amount = round(float(value or 0), 2)
    except (TypeError, ValueError) as exc:
        raise ValueError("El importe debe ser numerico") from exc
    if amount < 0:
        raise ValueError("El importe no puede ser negativo")
    return amount


def _positive_money(value: object) -> float:
    amount = _money(value)
    if amount <= 0:
        raise ValueError("El importe debe ser mayor a cero")
    return amount


def _currency(value: object) -> str:
    currency = _clean(value).upper() or "UYU"
    if currency in {"$", "PESOS"}:
        return "UYU"
    if currency in {"DOLAR", "DOLARES", "U$S"}:
        return "USD"
    if currency not in {"UYU", "USD"}:
        raise ValueError("Moneda invalida")
    return currency


def _required_date(value: object, label: str) -> str:
    cleaned = _clean(value)
    if not cleaned:
        raise ValueError(f"Falta {label}")
    try:
        date.fromisoformat(cleaned)
    except ValueError as exc:
        raise ValueError(f"{label.capitalize()} invalida") from exc
    return cleaned


def _safe_currency(value: object) -> str:
    try:
        return _currency(value)
    except ValueError:
        return "UYU"


def _day_key(value: object) -> str:
    return _clean(value)[:10]


def _normalize_text(value: object) -> str:
    raw = str(value or "").lower()
    raw = unicodedata.normalize("NFKD", raw)
    return "".join(ch for ch in raw if not unicodedata.combining(ch))


def _numeric(value: object) -> int | None:
    digits = "".join(ch for ch in _clean(value) if ch.isdigit())
    return int(digits) if digits else None


def _functional_currency_fields(currency: str, amount: float, transaction_date: str) -> dict:
    if currency == "UYU":
        return {
            "functional_currency": "UYU",
            "presentation_currency": "UYU",
            "functional_amount": amount,
            "exchange_rate": 1.0,
            "exchange_rate_date": transaction_date,
            "exchange_rate_source": "Moneda funcional UYU",
        }
    cached = collection("accounting_exchange_rates").find_one(
        {"currency": "USD", "transaction_date": transaction_date}, {"_id": 0}
    )
    rate_data = cached or fetch_bcu_accounting_usd(transaction_date)
    if not cached:
        collection("accounting_exchange_rates").update_one(
            {"currency": "USD", "transaction_date": transaction_date},
            {"$set": {**rate_data, "created_at": _now()}},
            upsert=True,
        )
    rate = float(rate_data["rate"])
    return {
        "functional_currency": "UYU",
        "presentation_currency": "UYU",
        "foreign_amount": amount,
        "functional_amount": round(amount * rate, 2),
        "exchange_rate": rate,
        "exchange_rate_date": rate_data["rate_date"],
        "exchange_rate_source": rate_data["source"],
        "bcu_currency_code": rate_data.get("bcu_currency_code", "2225"),
    }


def _destination_account(category: object, payment_method: object, direction: object = "income") -> str:
    category_value = _clean(category).lower()
    method = _clean(payment_method).lower()
    direction_value = _clean(direction).lower()
    if category_value == CARD_SETTLEMENT_CATEGORY:
        return "banco"
    if direction_value == "income" and (category_value == "factura_credito" or method == "credito"):
        return "cuentas_por_cobrar"
    if method == "efectivo":
        return "caja"
    if method in BANK_PAYMENT_METHODS:
        return "banco"
    if method in CARD_PAYMENT_METHODS:
        return "tarjetas_por_pagar" if direction_value == "expense" else "financiera"
    if method == "cheque":
        return "banco" if direction_value == "expense" else "caja"
    return "por_clasificar"


def _account_balances(movements: list[dict]) -> dict:
    balances = {
        "cash": 0.0,
        "bank": 0.0,
        "card_receivables": 0.0,
        "accounts_receivable": 0.0,
    }
    balance_key = {
        "caja": "cash",
        "banco": "bank",
        "financiera": "card_receivables",
        "cuentas_por_cobrar": "accounts_receivable",
    }
    for movement in movements:
        amount = float(movement.get("functional_amount") or (movement.get("amount") if _safe_currency(movement.get("currency")) == "UYU" else 0) or 0)
        if amount <= 0:
            continue
        destination = _clean(movement.get("destination_account")) or _destination_account(
            movement.get("category"), movement.get("payment_method"), movement.get("direction")
        )
        # Legacy expense movements could have been stored as accounts receivable
        # solely because their category was "factura_credito".  Their payment
        # method determines the real cash/bank destination, just as it does in
        # the journal entry.
        if movement.get("direction") == "expense" and _clean(movement.get("category")) == "factura_credito":
            destination = _destination_account(
                movement.get("category"), movement.get("payment_method"), movement.get("direction")
            )
        key = balance_key.get(destination)
        if not key:
            continue
        if movement.get("direction") == "transfer" and _clean(movement.get("category")) == CARD_SETTLEMENT_CATEGORY:
            balances["card_receivables"] -= amount
            balances["bank"] += amount
            continue
        balances[key] += -amount if movement.get("direction") == "expense" else amount
    return {key: round(value, 2) for key, value in balances.items()}


def _payment_method_detail(movement: dict) -> str:
    method = _clean(movement.get("payment_method"))
    card_type = _clean(movement.get("card_payment_type"))
    installments = int(movement.get("card_installments") or 0)
    if method not in CARD_PAYMENT_METHODS or not card_type:
        return method
    if card_type == "debito":
        return f"{method} - debito"
    installment_label = "cuota" if installments == 1 else "cuotas"
    return f"{method} - credito {installments} {installment_label}"


def _ensure_storage() -> None:
    collection("accounting_movements").create_index("id", unique=True)
    collection("accounting_movements").create_index([("brand_id", 1), ("year", 1), ("month", 1), ("workday_number", 1)])
    collection("accounting_movements").create_index("source_key", unique=True, sparse=True)
    collection("supplier_invoices").create_index("id", unique=True)
    collection("supplier_invoices").create_index("source_key", unique=True, sparse=True)
    collection("supplier_invoices").create_index([("brand_id", 1), ("supplier", 1), ("status", 1)])
    collection("supplier_payments").create_index("id", unique=True)
    collection("supplier_payments").create_index("supplier_invoice_id")
    collection("accounting_adjustments").create_index("id", unique=True)
    collection("accounting_adjustments").create_index([("brand_id", 1), ("date", 1)])
    collection("accounting_sale_costs").create_index([("brand_id", 1), ("date", 1)], unique=True)
    collection("accounting_exchange_rates").create_index([("currency", 1), ("transaction_date", 1)], unique=True)
    collection("accounting_day_closures").create_index([("brand_id", 1), ("date", 1)], unique=True)
    collection("accounting_till_handovers").create_index([("brand_id", 1), ("date", 1)], unique=True)


def _last_closed_date() -> str:
    closure = collection("accounting_day_closures").find_one(
        {"brand_id": "casa"}, {"_id": 0, "date": 1}, sort=[("date", -1)]
    )
    return _day_key(closure.get("date")) if closure else ""


def _assert_date_open(transaction_date: str) -> None:
    last_closed = _last_closed_date()
    if last_closed and transaction_date <= last_closed:
        raise ValueError(f"El dia contable {transaction_date} esta cerrado. Ultimo cierre: {last_closed}")


def _daily_control(invoices: list[dict], movements: list[dict], adjustments: list[dict], sale_costs: list[dict], cutoff: str) -> dict:
    last_closed = _last_closed_date()
    activity_dates = {
        _day_key(movement.get("date")) for movement in movements if _day_key(movement.get("date")) <= cutoff
    }
    activity_dates.update(
        _day_key(invoice.get("purchase_date")) for invoice in invoices
        if _clean(invoice.get("status")) != "no_aplica" and _day_key(invoice.get("purchase_date")) <= cutoff
    )
    activity_dates.update(
        _day_key(adjustment.get("date")) for adjustment in adjustments if _day_key(adjustment.get("date")) <= cutoff
    )
    activity_dates.discard("")
    next_open_date = next((item for item in sorted(activity_dates) if not last_closed or item > last_closed), "")
    blockers: list[dict] = []
    if next_open_date:
        day_invoices = [invoice for invoice in invoices if _day_key(invoice.get("purchase_date")) == next_open_date and _clean(invoice.get("status")) != "no_aplica"]
        for invoice in day_invoices:
            classification = _clean(invoice.get("accounting_classification")).lower()
            if classification not in SUPPLIER_CLASSIFICATION_ACCOUNTS:
                blockers.append({"type": "supplier_classification", "id": invoice.get("id"), "label": f"Factura {invoice.get('invoice_number', '')} de {invoice.get('supplier', '')} sin clasificar"})
            if _safe_currency(invoice.get("currency")) == "USD" and not invoice.get("functional_amount"):
                blockers.append({"type": "exchange_rate", "id": invoice.get("id"), "label": f"Factura {invoice.get('invoice_number', '')} sin conversion BCU"})
        has_sales = any(
            _day_key(movement.get("date")) == next_open_date
            and _clean(movement.get("direction")) == "income"
            and _clean(movement.get("category")) in {"facturas", "factura_credito"}
            for movement in movements
        )
        has_sale_cost_resolution = any(_day_key(record.get("date")) == next_open_date for record in sale_costs)
        if has_sales and not has_sale_cost_resolution:
            blockers.append({"type": "sale_cost", "id": next_open_date, "label": "Costo de ventas o confirmacion sin inventario pendiente"})
        conciliation = conciliate_cards(next_open_date)
        if conciliation.get("has_faltante"):
            faltantes = ", ".join(item["medio"] for item in conciliation["per_medio"] if item["flag"] == "faltante")
            blockers.append({"type": "card_faltante", "id": next_open_date, "label": f"Faltante de tarjetas sin registrar en caja: {faltantes}"})
    return {
        "last_closed_date": last_closed,
        "next_open_date": next_open_date,
        "blockers": blockers,
        "can_close": bool(next_open_date) and not blockers,
        "remaining_activity_days": sum(1 for item in activity_dates if not last_closed or item > last_closed),
    }


def close_accounting_day(close_date: str, user: str) -> dict:
    _ensure_storage()
    close_date = _required_date(close_date, "fecha de cierre")
    invoices = list(collection("supplier_invoices").find({"brand_id": "casa"}, {"_id": 0}))
    movements = list(collection("accounting_movements").find({"brand_id": "casa"}, {"_id": 0}))
    adjustments = list(collection("accounting_adjustments").find({"brand_id": "casa"}, {"_id": 0}))
    sale_costs = list(collection("accounting_sale_costs").find({"brand_id": "casa"}, {"_id": 0}))
    control = _daily_control(invoices, movements, adjustments, sale_costs, date.today().isoformat())
    if close_date != control["next_open_date"]:
        raise ValueError(f"El siguiente dia contable a cerrar es {control['next_open_date'] or 'ninguno'}")
    if control["blockers"]:
        raise ValueError("El dia tiene pendientes y no puede cerrarse")
    closure = {
        "id": str(uuid4()), "brand_id": "casa", "date": close_date,
        "closed_at": _now(), "closed_by": user,
    }
    collection("accounting_day_closures").insert_one(closure)
    closure.pop("_id", None)
    return {"closure": closure}


def sync_ucfe_supplier_invoices(user: str = "ucfe") -> dict:
    _ensure_storage()
    created = 0
    updated = 0
    skipped_closed = 0
    last_closed = _last_closed_date()
    for cfe in collection("ucfe_received_cfe").find({}, {"_id": 0, "xml": 0}):
        ucfe_id = _clean(cfe.get("ucfe_id"))
        if not ucfe_id:
            continue
        amount = _money(cfe.get("amount_payable") if cfe.get("amount_payable") is not None else cfe.get("total_amount"))
        if amount <= 0:
            continue
        series_number = _clean(cfe.get("series_number"))
        supplier = _clean(cfe.get("supplier_name")) or _clean(cfe.get("supplier_rut")) or "Proveedor UCFE"
        source_key = f"UCFE_CFE:{ucfe_id}"
        existing = collection("supplier_invoices").find_one({"source_key": source_key}, {"_id": 0})
        document_date = _clean(cfe.get("document_date"))[:10]
        if existing and last_closed and document_date <= last_closed:
            skipped_closed += 1
            continue
        payload = {
            "supplier": supplier,
            "rut": _clean(cfe.get("supplier_rut")),
            "invoice_number": series_number or ucfe_id,
            "currency": _safe_currency(cfe.get("currency") or "UYU"),
            "amount": amount,
            "paid_amount": float(existing.get("paid_amount") or 0) if existing else 0,
            "purchase_date": document_date,
            "due_date": (_clean(existing.get("due_date"))[:10] or document_date) if existing else document_date,
            "status": existing.get("status") if existing and existing.get("paid_amount") else "pendiente",
            "ucfe_cfe_id": ucfe_id,
            "source_key": source_key,
            "notes": "Factura recibida desde UCFE",
            "accounting_classification": _clean(existing.get("accounting_classification")) if existing else "",
        }
        upsert_supplier_invoice(payload, user, ensure_storage=False)
        if existing:
            updated += 1
        else:
            created += 1
    return {"created": created, "updated": updated, "skipped_closed": skipped_closed}


def register_movement(data: dict, user: str) -> dict:
    _ensure_storage()
    direction = _clean(data.get("direction")).lower()
    if direction not in MOVEMENT_DIRECTIONS:
        raise ValueError("Direccion de movimiento invalida")
    category = _clean(data.get("category")).lower() or "general"
    supplied_payment_method = _clean(data.get("payment_method")).lower()
    if direction == "income" and category in {"facturas", "factura_credito"} and not supplied_payment_method:
        raise ValueError("Falta medio de cobro de la factura de venta")
    payment_method = supplied_payment_method or "efectivo"
    if payment_method not in PAYMENT_METHODS:
        raise ValueError("Medio de pago invalido")
    if direction == "income" and category == "factura_credito" and payment_method != "credito":
        raise ValueError("Las facturas a credito deben quedar en cuentas por cobrar")
    if direction == "income" and category == "facturas" and payment_method == "credito":
        raise ValueError("Usa la categoria factura_credito para una venta a credito")
    card_plan = _clean(data.get("card_plan")).lower()
    card_payment_type = ""
    card_installments = 0
    if direction == "income" and category == "facturas" and payment_method in CARD_PAYMENT_METHODS:
        if card_plan not in CARD_PAYMENT_PLANS:
            raise ValueError("Selecciona debito o credito en 1, 2 o 3 cuotas")
        card_payment_type, card_installments = CARD_PAYMENT_PLANS[card_plan]
    if direction == "transfer" and category == CARD_SETTLEMENT_CATEGORY and payment_method not in CARD_PAYMENT_METHODS:
        raise ValueError("Selecciona la tarjeta o financiera que realizo la acreditacion")
    amount = _positive_money(data.get("amount"))
    currency = _currency(data.get("currency"))
    if direction == "transfer" and category == CARD_SETTLEMENT_CATEGORY:
        if currency != "UYU":
            raise ValueError("Las acreditaciones de tarjeta se registran en UYU")
        existing_movements = list(collection("accounting_movements").find(
            {"brand_id": "casa"},
            {"_id": 0, "direction": 1, "category": 1, "payment_method": 1, "destination_account": 1, "amount": 1, "currency": 1},
        ))
        if amount > _account_balances(existing_movements)["card_receivables"]:
            raise ValueError("La acreditacion no puede superar el saldo pendiente en financieras")
    year = int(data.get("year") or datetime.now().year)
    month = int(data.get("month") or datetime.now().month)
    if not 1 <= month <= 12:
        raise ValueError("Mes invalido")
    workday_number = int(data.get("workday_number") or 1)
    if workday_number <= 0:
        raise ValueError("Dia trabajado invalido")
    invoice_number = _clean(data.get("invoice_number"))
    source_key = _clean(data.get("source_key"))
    issue_date = _clean(data.get("issue_date"))
    due_date = _clean(data.get("due_date"))
    if direction == "income" and category in {"facturas", "factura_credito"}:
        if not invoice_number:
            raise ValueError("Falta numero de factura de venta")
        issue_date = _required_date(issue_date, "fecha de emision")
        due_date = _required_date(due_date, "fecha de vencimiento")
        if due_date < issue_date:
            raise ValueError("El vencimiento no puede ser anterior a la emision")
    movement_date = _required_date(data.get("date"), "fecha del movimiento")
    _assert_date_open(movement_date)
    functional_fields = _functional_currency_fields(currency, amount, movement_date)
    movement = {
        "id": str(uuid4()),
        "brand_id": "casa",
        "year": year,
        "month": month,
        "workday_number": workday_number,
        "date": movement_date,
        "direction": direction,
        "category": category,
        "subcategory": _clean(data.get("subcategory")),
        "payment_method": payment_method,
        "card_payment_type": card_payment_type,
        "card_installments": card_installments,
        "origin_account": "financiera" if category == CARD_SETTLEMENT_CATEGORY else "",
        "destination_account": _destination_account(category, payment_method, direction),
        "amount": amount,
        "currency": currency,
        **functional_fields,
        "description": _clean(data.get("description")),
        "reference": _clean(data.get("reference")) or invoice_number,
        "invoice_number": invoice_number,
        "issue_date": issue_date,
        "due_date": due_date,
        "source": _clean(data.get("source")) or "manual",
        "supplier_invoice_id": _clean(data.get("supplier_invoice_id")),
        "reconciled": bool(data.get("reconciled", False)),
        "created_at": _now(),
        "created_by": user,
        "updated_at": _now(),
        "updated_by": user,
    }
    if source_key:
        movement["source_key"] = source_key
        existing = collection("accounting_movements").find_one({"source_key": source_key}, {"_id": 0})
        if existing:
            return {"movement": existing, "already_registered": True}
    collection("accounting_movements").insert_one(movement)
    movement.pop("_id", None)
    return {"movement": movement, "already_registered": False}


def replace_movement(movement_id: str, replacements: list[dict], user: str) -> dict:
    _ensure_storage()
    movement_id = _clean(movement_id)
    original = collection("accounting_movements").find_one(
        {"id": movement_id, "brand_id": "casa"}, {"_id": 0}
    )
    if not original:
        raise ValueError("El movimiento a reemplazar no existe")
    if original.get("direction") != "income" or original.get("category") != "facturas":
        raise ValueError("Solo se pueden desglosar movimientos de ventas")
    if not replacements:
        raise ValueError("Falta el detalle que reemplaza al movimiento")
    replacement_total = round(sum(_positive_money(item.get("amount")) for item in replacements), 2)
    if replacement_total != round(float(original.get("amount") or 0), 2):
        raise ValueError("El detalle debe sumar exactamente el importe original")

    created_ids: list[str] = []
    try:
        created = []
        for index, item in enumerate(replacements, start=1):
            payload = {
                **item,
                "year": original.get("year"),
                "month": original.get("month"),
                "workday_number": original.get("workday_number"),
                "date": original.get("date"),
                "direction": "income",
                "category": "facturas",
                "currency": original.get("currency"),
                "issue_date": original.get("issue_date") or original.get("date"),
                "due_date": original.get("due_date") or original.get("date"),
                "source": "sales_breakdown",
                "source_key": f"SALES_BREAKDOWN:{movement_id}:{index}",
            }
            result = register_movement(payload, user)
            created.append(result["movement"])
            if not result.get("already_registered"):
                created_ids.append(result["movement"]["id"])
        collection("accounting_movements").delete_one({"id": movement_id, "brand_id": "casa"})
        return {"replaced_movement_id": movement_id, "movements": created, "total": replacement_total}
    except Exception:
        if created_ids:
            collection("accounting_movements").delete_many({"id": {"$in": created_ids}})
        raise


def correct_movement_direction(movement_id: str, direction: str, user: str) -> dict:
    _ensure_storage()
    direction = _clean(direction).lower()
    if direction not in {"income", "expense"}:
        raise ValueError("La correccion debe ser entrada o salida")
    movement = collection("accounting_movements").find_one({"id": _clean(movement_id), "brand_id": "casa"}, {"_id": 0})
    if not movement:
        raise ValueError("El movimiento no existe")
    if _clean(movement.get("category")) in {"facturas", "factura_credito", "proveedores", CARD_SETTLEMENT_CATEGORY}:
        raise ValueError("Este tipo de movimiento requiere una correccion contable especifica")
    _assert_date_open(_day_key(movement.get("date")))
    if _clean(movement.get("direction")) == direction:
        return {"movement": movement, "already_corrected": True}
    correction = {
        "id": str(uuid4()), "brand_id": "casa", "movement_id": movement["id"],
        "date": _day_key(movement.get("date")), "previous_direction": movement.get("direction"),
        "new_direction": direction, "created_at": _now(), "created_by": user,
    }
    collection("accounting_movement_corrections").insert_one(correction)
    updated = collection("accounting_movements").find_one_and_update(
        {"id": movement["id"], "brand_id": "casa"},
        {"$set": {
            "direction": direction,
            "destination_account": _destination_account(movement.get("category"), movement.get("payment_method"), direction),
            "corrected_at": _now(), "corrected_by": user, "updated_at": _now(), "updated_by": user,
        }},
        return_document=ReturnDocument.AFTER,
        projection={"_id": 0},
    )
    correction.pop("_id", None)
    return {"movement": updated, "correction": correction}


def correct_movement_amount(movement_id: str, amount: object, user: str) -> dict:
    _ensure_storage()
    new_amount = _positive_money(amount)
    movement = collection("accounting_movements").find_one(
        {"id": _clean(movement_id), "brand_id": "casa"}, {"_id": 0}
    )
    if not movement:
        raise ValueError("El movimiento no existe")
    _assert_date_open(_day_key(movement.get("date")))
    if _safe_currency(movement.get("currency")) != "UYU":
        raise ValueError("La correccion directa de importe solo admite movimientos en UYU")
    previous_amount = round(float(movement.get("amount") or 0), 2)
    if previous_amount == new_amount:
        return {"movement": movement, "already_corrected": True}
    now = _now()
    correction = {
        "id": str(uuid4()), "brand_id": "casa", "movement_id": movement["id"],
        "date": _day_key(movement.get("date")), "field": "amount",
        "previous_amount": previous_amount, "new_amount": new_amount,
        "created_at": now, "created_by": user,
    }
    collection("accounting_movement_corrections").insert_one(correction)
    updated = collection("accounting_movements").find_one_and_update(
        {"id": movement["id"], "brand_id": "casa"},
        {"$set": {
            "amount": new_amount, "functional_amount": new_amount,
            "corrected_at": now, "corrected_by": user,
            "updated_at": now, "updated_by": user,
        }},
        return_document=ReturnDocument.AFTER,
        projection={"_id": 0},
    )
    correction.pop("_id", None)
    return {"movement": updated, "correction": correction}


def correct_movement_date(movement_id: str, new_date: str, user: str) -> dict:
    _ensure_storage()
    corrected_date = _required_date(new_date, "fecha corregida")
    movement = collection("accounting_movements").find_one(
        {"id": _clean(movement_id), "brand_id": "casa"}, {"_id": 0}
    )
    if not movement:
        raise ValueError("El movimiento no existe")
    previous_date = _day_key(movement.get("date"))
    _assert_date_open(previous_date)
    _assert_date_open(corrected_date)
    if previous_date == corrected_date:
        return {"movement": movement, "already_corrected": True}
    now = _now()
    correction = {
        "id": str(uuid4()), "brand_id": "casa", "movement_id": movement["id"],
        "date": previous_date, "field": "date", "previous_date": previous_date,
        "new_date": corrected_date, "created_at": now, "created_by": user,
    }
    collection("accounting_movement_corrections").insert_one(correction)
    updated = collection("accounting_movements").find_one_and_update(
        {"id": movement["id"], "brand_id": "casa"},
        {"$set": {
            "date": corrected_date, "year": int(corrected_date[:4]),
            "month": int(corrected_date[5:7]), "corrected_at": now,
            "corrected_by": user, "updated_at": now, "updated_by": user,
        }},
        return_document=ReturnDocument.AFTER,
        projection={"_id": 0},
    )
    correction.pop("_id", None)
    return {"movement": updated, "correction": correction}


def delete_duplicate_movement(movement_id: str, user: str) -> dict:
    _ensure_storage()
    movement = collection("accounting_movements").find_one(
        {"id": _clean(movement_id), "brand_id": "casa"}, {"_id": 0}
    )
    if not movement:
        raise ValueError("El movimiento no existe")
    _assert_date_open(_day_key(movement.get("date")))
    audit = {
        "id": str(uuid4()), "brand_id": "casa", "movement_id": movement["id"],
        "date": _day_key(movement.get("date")), "field": "deleted_duplicate",
        "previous_movement": movement, "created_at": _now(), "created_by": user,
    }
    collection("accounting_movement_corrections").insert_one(audit)
    deleted = collection("accounting_movements").delete_one(
        {"id": movement["id"], "brand_id": "casa"}
    )
    if deleted.deleted_count != 1:
        collection("accounting_movement_corrections").delete_one({"id": audit["id"]})
        raise ValueError("No se pudo eliminar el movimiento duplicado")
    audit.pop("_id", None)
    return {"deleted_movement": movement, "correction": audit}


def upsert_supplier_invoice(data: dict, user: str, *, ensure_storage: bool = True) -> dict:
    if ensure_storage:
        _ensure_storage()
    invoice_id = _clean(data.get("id"))
    source_key = _clean(data.get("source_key"))
    supplier = _clean(data.get("supplier"))
    invoice_number = _clean(data.get("invoice_number"))
    if not supplier:
        raise ValueError("Falta proveedor")
    if not invoice_number:
        raise ValueError("Falta numero de factura de compra")
    purchase_date = _required_date(data.get("purchase_date"), "fecha de emision")
    _assert_date_open(purchase_date)
    due_date = _required_date(data.get("due_date"), "fecha de vencimiento")
    if due_date < purchase_date:
        raise ValueError("El vencimiento no puede ser anterior a la emision")
    amount = _positive_money(data.get("amount"))
    paid_amount = _money(data.get("paid_amount"))
    if paid_amount > amount:
        raise ValueError("El pago no puede superar el monto de la factura")
    status = _clean(data.get("status")).lower()
    if not status:
        status = "pagada" if paid_amount >= amount else "parcial" if paid_amount > 0 else "pendiente"
    if status not in INVOICE_STATUSES:
        raise ValueError("Estado de factura invalido")
    now = _now()
    currency = _currency(data.get("currency"))
    functional_fields = _functional_currency_fields(currency, amount, purchase_date)
    invoice = {
        "brand_id": "casa",
        "supplier": supplier,
        "rut": _clean(data.get("rut")),
        "invoice_number": invoice_number,
        "currency": currency,
        "amount": amount,
        **functional_fields,
        "paid_amount": paid_amount,
        "balance": round(max(0, amount - paid_amount), 2),
        "purchase_date": purchase_date,
        "due_date": due_date,
        "status": status,
        "ucfe_cfe_id": _clean(data.get("ucfe_cfe_id")),
        "notes": _clean(data.get("notes")),
        "accounting_classification": _clean(data.get("accounting_classification")),
        "updated_at": now,
        "updated_by": user,
    }
    if source_key:
        invoice["source_key"] = source_key
    query = {"id": invoice_id} if invoice_id else None
    if query is None and source_key:
        query = {"source_key": source_key}
    if query is None and supplier and invoice_number:
        query = {"brand_id": "casa", "supplier": supplier, "invoice_number": invoice_number}
    existing = collection("supplier_invoices").find_one(query or {"id": "__none__"}, {"_id": 0})
    if existing:
        invoice_id = existing["id"]
    else:
        invoice_id = str(uuid4())
    invoice["id"] = invoice_id
    collection("supplier_invoices").update_one(
        {"id": invoice_id},
        {"$set": invoice, "$setOnInsert": {"created_at": now, "created_by": user}},
        upsert=True,
    )
    saved = collection("supplier_invoices").find_one({"id": invoice_id}, {"_id": 0}) or invoice
    return {"invoice": saved}


def register_supplier_payment(data: dict, user: str) -> dict:
    _ensure_storage()
    invoice_id = _clean(data.get("supplier_invoice_id"))
    invoice = None
    if invoice_id:
        invoice = collection("supplier_invoices").find_one({"id": invoice_id}, {"_id": 0})
        if not invoice:
            raise ValueError("La factura no existe")
    amount = _positive_money(data.get("amount"))
    payment_date = _required_date(data.get("payment_date"), "fecha de pago")
    _assert_date_open(payment_date)
    payment_currency = _currency(data.get("currency") or (invoice or {}).get("currency"))
    functional_fields = _functional_currency_fields(payment_currency, amount, payment_date)
    invoice_currency = ""
    invoice_currency_amount = None
    settlement_exchange_rate = None
    if invoice:
        invoice_currency = _currency(invoice.get("currency"))
        settlement_exchange_rate = 1.0
        if payment_currency == invoice_currency:
            invoice_currency_amount = amount
            if invoice_currency == "USD":
                settlement_exchange_rate = float(functional_fields.get("exchange_rate") or 0)
        elif {payment_currency, invoice_currency} == {"UYU", "USD"}:
            invoice_currency_amount = _positive_money(data.get("invoice_currency_amount"))
            if payment_currency == "UYU":
                settlement_exchange_rate = round(amount / invoice_currency_amount, 6)
            else:
                settlement_exchange_rate = round(invoice_currency_amount / amount, 6)
        else:
            raise ValueError("No se puede convertir la moneda del pago a la moneda de la factura")
        invoice_balance = round(float(invoice.get("balance") or 0), 2)
        if invoice_currency_amount > invoice_balance + 0.01:
            raise ValueError("El pago supera el saldo pendiente convertido de la factura")
    supplier = _clean((invoice or {}).get("supplier") or data.get("supplier"))
    if not supplier:
        raise ValueError("Indica el proveedor del pago")
    payment = {
        "id": str(uuid4()),
        "brand_id": "casa",
        "supplier_invoice_id": invoice_id,
        "supplier": supplier,
        "payment_date": payment_date,
        "amount": amount,
        "currency": payment_currency,
        "invoice_currency": invoice_currency,
        "invoice_currency_amount": invoice_currency_amount,
        "settlement_exchange_rate": settlement_exchange_rate,
        **functional_fields,
        "receipt_number": _clean(data.get("receipt_number")),
        "payment_method": _clean(data.get("payment_method")).lower() or "transferencia",
        "bank_reference": _clean(data.get("bank_reference")),
        "notes": _clean(data.get("notes")),
        "created_at": _now(),
        "created_by": user,
    }
    if payment["payment_method"] not in PAYMENT_METHODS - {"credito"}:
        raise ValueError("Medio de pago invalido")
    collection("supplier_payments").insert_one(payment)
    if not invoice:
        payment.pop("_id", None)
        return {"payment": payment, "invoice": None}
    paid = round(float(invoice.get("paid_amount") or 0) + invoice_currency_amount, 2)
    balance = round(max(0, float(invoice.get("amount") or 0) - paid), 2)
    status = "pagada" if balance == 0 else "parcial"
    updated = collection("supplier_invoices").find_one_and_update(
        {"id": invoice_id},
        {"$set": {"paid_amount": paid, "balance": balance, "status": status, "updated_at": _now(), "updated_by": user}},
        return_document=ReturnDocument.AFTER,
        projection={"_id": 0},
    )
    payment.pop("_id", None)
    return {"payment": payment, "invoice": updated}


def register_daily_supplier_payment(data: dict, user: str) -> dict:
    result = register_supplier_payment(data, user)
    payment = result["payment"]
    invoice = result["invoice"] or {}
    invoice_number = _clean(invoice.get("invoice_number"))
    description = _clean(data.get("description"))
    if not description:
        description = f"Pago proveedor {payment.get('supplier', '')}".strip()
        if invoice_number:
            description = f"{description} factura {invoice_number}"
    reference = _clean(data.get("reference")) or _clean(data.get("receipt_number")) or _clean(data.get("bank_reference")) or invoice_number
    movement_result = register_movement({
        "year": data.get("year"),
        "month": data.get("month"),
        "workday_number": data.get("workday_number"),
        "date": data.get("payment_date") or data.get("date"),
        "direction": "expense",
        "category": "proveedores",
        "subcategory": payment.get("supplier", ""),
        "payment_method": data.get("payment_method"),
        "amount": payment.get("amount"),
        "currency": payment.get("currency"),
        "description": description,
        "reference": reference,
        "invoice_number": invoice_number,
        "issue_date": invoice.get("purchase_date", ""),
        "due_date": invoice.get("due_date", ""),
        "source": "supplier_payment",
        "source_key": f"SUPPLIER_PAYMENT:{payment['id']}",
        "supplier_invoice_id": payment.get("supplier_invoice_id"),
    }, user)
    collection("supplier_payments").update_one(
        {"id": payment["id"]},
        {"$set": {"accounting_movement_id": movement_result["movement"]["id"]}},
    )
    payment["accounting_movement_id"] = movement_result["movement"]["id"]
    return {**result, "movement": movement_result["movement"]}


def register_labor_provision(data: dict, user: str) -> dict:
    _ensure_storage()
    provision_type = _clean(data.get("provision_type")).lower()
    if provision_type not in {"aguinaldo", "licencia", "salario_vacacional"}:
        raise ValueError("Tipo de provision invalido")
    provision_date = _required_date(data.get("date"), "fecha de provision")
    _assert_date_open(provision_date)
    amount = _positive_money(data.get("amount"))
    adjustment = {
        "id": str(uuid4()),
        "brand_id": "casa",
        "type": "labor_provision",
        "provision_type": provision_type,
        "date": provision_date,
        "amount": amount,
        "currency": "UYU",
        "description": _clean(data.get("description")) or f"Provision de {provision_type.replace('_', ' ')}",
        "created_at": _now(),
        "created_by": user,
    }
    collection("accounting_adjustments").insert_one(adjustment)
    adjustment.pop("_id", None)
    return {"adjustment": adjustment}


def register_sale_cost(data: dict, user: str) -> dict:
    _ensure_storage()
    cost_date = _required_date(data.get("date"), "fecha del costo de ventas")
    _assert_date_open(cost_date)
    treatment = _clean(data.get("treatment")).lower()
    if treatment not in {"inventory", "not_applicable"}:
        raise ValueError("Tratamiento de costo de venta invalido")
    amount = _positive_money(data.get("amount")) if treatment == "inventory" else 0.0
    record = {
        "id": _clean(data.get("id")) or str(uuid4()),
        "brand_id": "casa",
        "date": cost_date,
        "treatment": treatment,
        "amount": amount,
        "currency": "UYU",
        "functional_currency": "UYU",
        "presentation_currency": "UYU",
        "description": _clean(data.get("description")) or ("Costo real de mercaderias vendidas" if treatment == "inventory" else "Venta sin efecto en inventarios"),
        "source": _clean(data.get("source")) or "determinacion_administrativa",
        "updated_at": _now(),
        "updated_by": user,
    }
    existing = collection("accounting_sale_costs").find_one({"brand_id": "casa", "date": cost_date}, {"_id": 0})
    if existing:
        record["id"] = existing["id"]
    collection("accounting_sale_costs").update_one(
        {"brand_id": "casa", "date": cost_date},
        {"$set": record, "$setOnInsert": {"created_at": _now(), "created_by": user}},
        upsert=True,
    )
    return {"sale_cost": collection("accounting_sale_costs").find_one({"brand_id": "casa", "date": cost_date}, {"_id": 0})}


def classify_supplier_invoice(invoice_id: str, classification: str, user: str) -> dict:
    classification = _clean(classification).lower()
    if classification not in SUPPLIER_CLASSIFICATION_ACCOUNTS:
        raise ValueError("Clasificacion contable invalida")
    existing = collection("supplier_invoices").find_one({"id": _clean(invoice_id), "brand_id": "casa"}, {"_id": 0})
    if not existing:
        raise ValueError("La factura no existe")
    _assert_date_open(_day_key(existing.get("purchase_date")))
    invoice = collection("supplier_invoices").find_one_and_update(
        {"id": _clean(invoice_id), "brand_id": "casa"},
        {"$set": {"accounting_classification": classification, "classified_at": _now(), "classified_by": user, "updated_at": _now(), "updated_by": user}},
        return_document=ReturnDocument.AFTER,
        projection={"_id": 0},
    )
    return {"invoice": invoice}


def _financial_account(payment_method: object) -> str:
    method = _clean(payment_method).lower()
    if method == "visa":
        return "1131"
    if method in {"master", "maestro"}:
        return "1132"
    return "1139"


def _cash_account(payment_method: object, *, incoming: bool) -> str:
    method = _clean(payment_method).lower()
    if method == "efectivo":
        return "1111"
    if method == "cheque":
        return "1112" if incoming else "1113"
    if method in BANK_PAYMENT_METHODS:
        return "1113"
    if method in CARD_PAYMENT_METHODS:
        return _financial_account(method) if incoming else "2140"
    if method == "credito":
        return "1120"
    return "1190" if incoming else "2190"


def _expense_account(movement: dict) -> str:
    category = _clean(movement.get("category")).lower()
    if category == "proveedores" and _clean(movement.get("supplier_invoice_id")):
        return "2110"
    if category in {"costo_venta", "proveedores"}:
        return "5110"
    if category == "impuestos":
        return "5310"
    if category == "sueldos":
        return "5210"
    if category in {"servicios", "costos_fijos"}:
        return "5410"
    if category == "retiros":
        return "3510"
    return "5910"


def _entry(entry_id: str, entry_date: str, description: str, reference: str, source: str, lines: list[dict]) -> dict:
    debit = round(sum(float(line.get("debit") or 0) for line in lines), 2)
    credit = round(sum(float(line.get("credit") or 0) for line in lines), 2)
    if debit != credit:
        raise ValueError(f"El asiento {entry_id} no balancea")
    enriched = []
    for line in lines:
        account = ACCOUNT_BY_CODE[line["account_code"]]
        enriched.append({**line, "account_name": account["name"], "account_class": account["class"]})
    return {
        "id": entry_id,
        "date": entry_date,
        "description": description,
        "reference": reference,
        "source": source,
        "currency": "UYU",
        "debit": debit,
        "credit": credit,
        "balanced": True,
        "lines": enriched,
    }


def _movement_entry(movement: dict, classified_invoice_ids: set[str], invoices_by_id: dict[str, dict] | None = None) -> dict | None:
    currency = _safe_currency(movement.get("currency"))
    amount = round(float(movement.get("functional_amount") or (movement.get("amount") if currency == "UYU" else 0) or 0), 2)
    if amount <= 0:
        return None
    direction = _clean(movement.get("direction")).lower()
    category = _clean(movement.get("category")).lower()
    if category == "proveedores" and _clean(movement.get("supplier_invoice_id")) not in classified_invoice_ids:
        return None
    method = _clean(movement.get("payment_method")).lower()
    entry_date = _clean(movement.get("date")) or f"{movement.get('year')}-{int(movement.get('month') or 1):02d}-01"
    description = _clean(movement.get("description")) or category or "Movimiento contable"
    reference = _clean(movement.get("reference")) or _clean(movement.get("invoice_number"))

    if direction == "expense" and category == "proveedores" and _clean(movement.get("supplier_invoice_id")):
        invoice = (invoices_by_id or {}).get(_clean(movement.get("supplier_invoice_id")), {})
        if currency == "USD" and invoice:
            historical_rate = float(invoice.get("exchange_rate") or 0)
            if historical_rate <= 0:
                return None
            carrying_amount = round(float(movement.get("amount") or 0) * historical_rate, 2)
            lines = [{"account_code": "2110", "debit": carrying_amount, "credit": 0.0}]
            if amount > carrying_amount:
                lines.append({"account_code": "5520", "debit": round(amount - carrying_amount, 2), "credit": 0.0})
            lines.append({"account_code": _cash_account(method, incoming=False), "debit": 0.0, "credit": amount})
            if amount < carrying_amount:
                lines.append({"account_code": "4220", "debit": 0.0, "credit": round(carrying_amount - amount, 2)})
            return _entry(f"MOV:{movement.get('id')}", entry_date, description, reference, "pago_proveedor_moneda_extranjera", lines)

    if direction == "transfer":
        if category == CARD_SETTLEMENT_CATEGORY:
            debit_code, credit_code = "1113", _financial_account(method)
        elif _clean(movement.get("source")) == "opening_balance" or category == "saldo_inicial":
            debit_code = _cash_account(method, incoming=True)
            credit_code = "3110"
        elif category == "depositos":
            debit_code, credit_code = "1113", "1111"
        else:
            debit_code = _cash_account(method, incoming=True)
            credit_code = "1190"
    elif direction == "income":
        debit_code = _cash_account(method, incoming=True)
        credit_code = "3110" if category == "aportes" else "4110" if category in {"facturas", "factura_credito", "tarjetas"} else "4210"
    elif direction == "expense":
        debit_code = _expense_account(movement)
        credit_code = _cash_account(method, incoming=False)
    else:
        return None
    return _entry(
        f"MOV:{movement.get('id')}", entry_date, description, reference, "movimiento_diario",
        [
            {"account_code": debit_code, "debit": amount, "credit": 0.0},
            {"account_code": credit_code, "debit": 0.0, "credit": amount},
        ],
    )


def _supplier_invoice_entry(invoice: dict) -> dict | None:
    if _clean(invoice.get("status")) == "no_aplica":
        return None
    currency = _safe_currency(invoice.get("currency"))
    amount = round(float(invoice.get("functional_amount") or (invoice.get("amount") if currency == "UYU" else 0) or 0), 2)
    if amount <= 0:
        return None
    debit_account = SUPPLIER_CLASSIFICATION_ACCOUNTS.get(_clean(invoice.get("accounting_classification")).lower())
    if not debit_account:
        return None
    invoice_date = _clean(invoice.get("purchase_date"))[:10]
    return _entry(
        f"INV:{invoice.get('id')}", invoice_date,
        f"Factura de {invoice.get('supplier', 'proveedor')}", _clean(invoice.get("invoice_number")), "factura_proveedor",
        [
            {"account_code": debit_account, "debit": amount, "credit": 0.0},
            {"account_code": "2110", "debit": 0.0, "credit": amount},
        ],
    )


def _sale_cost_entry(record: dict) -> dict | None:
    if _clean(record.get("treatment")) != "inventory":
        return None
    amount = round(float(record.get("amount") or 0), 2)
    if amount <= 0:
        return None
    return _entry(
        f"COGS:{record.get('id')}", _clean(record.get("date")), _clean(record.get("description")), "", "costo_de_ventas",
        [
            {"account_code": "5110", "debit": amount, "credit": 0.0},
            {"account_code": "1140", "debit": 0.0, "credit": amount},
        ],
    )


def _adjustment_entry(adjustment: dict) -> dict | None:
    if _safe_currency(adjustment.get("currency")) != "UYU" or adjustment.get("type") != "labor_provision":
        return None
    provision_accounts = {
        "aguinaldo": ("5220", "2132"),
        "licencia": ("5230", "2133"),
        "salario_vacacional": ("5240", "2134"),
    }
    accounts = provision_accounts.get(_clean(adjustment.get("provision_type")))
    if not accounts:
        return None
    amount = round(float(adjustment.get("amount") or 0), 2)
    return _entry(
        f"ADJ:{adjustment.get('id')}", _clean(adjustment.get("date")),
        _clean(adjustment.get("description")), "", "ajuste_contable",
        [
            {"account_code": accounts[0], "debit": amount, "credit": 0.0},
            {"account_code": accounts[1], "debit": 0.0, "credit": amount},
        ],
    )


def _journal_and_statements(movements: list[dict], invoices: list[dict], adjustments: list[dict], sale_costs: list[dict], year: int, month: int) -> dict:
    cutoff = date(year, month, monthrange(year, month)[1]).isoformat()
    period_start = f"{year}-01-01"
    entries: list[dict] = []
    pending_currency = 0
    pending_classification = 0
    invoices_by_id = {_clean(invoice.get("id")): invoice for invoice in invoices}
    classified_invoice_ids = {
        _clean(invoice.get("id"))
        for invoice in invoices
        if _clean(invoice.get("accounting_classification")).lower() in SUPPLIER_CLASSIFICATION_ACCOUNTS
    }
    for movement in movements:
        movement_date = _clean(movement.get("date"))
        if movement_date and movement_date <= cutoff:
            entry = _movement_entry(movement, classified_invoice_ids, invoices_by_id)
            if entry:
                entries.append(entry)
            elif _safe_currency(movement.get("currency")) != "UYU" and not movement.get("functional_amount"):
                pending_currency += 1
    for invoice in invoices:
        invoice_date = _clean(invoice.get("purchase_date"))[:10]
        if invoice_date and invoice_date <= cutoff:
            classification_is_valid = _clean(invoice.get("accounting_classification")).lower() in SUPPLIER_CLASSIFICATION_ACCOUNTS
            if not classification_is_valid and _clean(invoice.get("status")) != "no_aplica":
                pending_classification += 1
            entry = _supplier_invoice_entry(invoice)
            if entry:
                entries.append(entry)
            elif _safe_currency(invoice.get("currency")) != "UYU" and not invoice.get("functional_amount"):
                pending_currency += 1
    classified_usd_invoices = [
        invoice for invoice in invoices
        if _safe_currency(invoice.get("currency")) == "USD"
        and _clean(invoice.get("id")) in classified_invoice_ids
        and _clean(invoice.get("purchase_date"))[:10] <= cutoff
        and invoice.get("exchange_rate")
    ]
    if classified_usd_invoices:
        try:
            closing_rate = float(_functional_currency_fields("USD", 1.0, cutoff)["exchange_rate"])
            for invoice in classified_usd_invoices:
                invoice_id = _clean(invoice.get("id"))
                paid_to_cutoff = sum(
                    float(movement.get("amount") or 0)
                    for movement in movements
                    if _clean(movement.get("supplier_invoice_id")) == invoice_id
                    and _clean(movement.get("direction")) == "expense"
                    and _clean(movement.get("category")) == "proveedores"
                    and _clean(movement.get("date")) <= cutoff
                )
                remaining_usd = round(max(0.0, float(invoice.get("amount") or 0) - paid_to_cutoff), 2)
                historical_value = round(remaining_usd * float(invoice.get("exchange_rate") or 0), 2)
                closing_value = round(remaining_usd * closing_rate, 2)
                difference = round(closing_value - historical_value, 2)
                if not remaining_usd or not difference:
                    continue
                lines = (
                    [{"account_code": "5520", "debit": difference, "credit": 0.0}, {"account_code": "2110", "debit": 0.0, "credit": difference}]
                    if difference > 0 else
                    [{"account_code": "2110", "debit": -difference, "credit": 0.0}, {"account_code": "4220", "debit": 0.0, "credit": -difference}]
                )
                entries.append(_entry(
                    f"FXR:{invoice_id}:{cutoff}", cutoff,
                    f"Reexpresion de saldo USD de {invoice.get('supplier', 'proveedor')}",
                    _clean(invoice.get("invoice_number")), "remeasurement_moneda_extranjera", lines,
                ))
        except Exception:
            pending_currency += len(classified_usd_invoices)
    for adjustment in adjustments:
        if _clean(adjustment.get("date")) <= cutoff:
            entry = _adjustment_entry(adjustment)
            if entry:
                entries.append(entry)
    for record in sale_costs:
        if _clean(record.get("date")) <= cutoff:
            entry = _sale_cost_entry(record)
            if entry:
                entries.append(entry)

    sales_dates = {
        _clean(movement.get("date"))
        for movement in movements
        if _clean(movement.get("date")) <= cutoff
        and _clean(movement.get("direction")) == "income"
        and _clean(movement.get("category")) in {"facturas", "factura_credito"}
    }
    resolved_sale_cost_dates = {_clean(record.get("date")) for record in sale_costs}
    pending_sale_cost_dates = []
    for sales_date in sorted(sales_dates - resolved_sale_cost_dates):
        daily_sales = sum(
            float(movement.get("functional_amount") or (movement.get("amount") if _safe_currency(movement.get("currency")) == "UYU" else 0) or 0)
            for movement in movements if _clean(movement.get("date")) == sales_date and _clean(movement.get("direction")) == "income" and _clean(movement.get("category")) in {"facturas", "factura_credito"}
        )
        pending_sale_cost_dates.append({"date": sales_date, "sales_amount_uyu": round(daily_sales, 2)})
    entries.sort(key=lambda item: (item["date"], item["id"]))

    balances = {code: 0.0 for code in ACCOUNT_BY_CODE}
    for entry in entries:
        for line in entry["lines"]:
            balances[line["account_code"]] += float(line.get("debit") or 0) - float(line.get("credit") or 0)
    account_balances = []
    for account in CHART_OF_ACCOUNTS:
        raw = round(balances[account["code"]], 2)
        display = raw if account["nature"] == "debit" else -raw
        account_balances.append({**account, "balance": round(display, 2)})

    def class_total(account_class: str) -> float:
        return round(sum(item["balance"] for item in account_balances if item["class"] == account_class), 2)

    assets = class_total("activo")
    liabilities = class_total("pasivo")
    equity_accounts = class_total("patrimonio")
    cumulative_revenue = class_total("ingreso")
    cumulative_expenses = class_total("gasto")
    cumulative_result = round(cumulative_revenue - cumulative_expenses, 2)
    equity = round(equity_accounts + cumulative_result, 2)

    current_year_entries = [entry for entry in entries if period_start <= entry["date"] <= cutoff]
    current_revenue = 0.0
    current_expenses = 0.0
    contributions = 0.0
    withdrawals = 0.0
    for entry in current_year_entries:
        for line in entry["lines"]:
            account_class = ACCOUNT_BY_CODE[line["account_code"]]["class"]
            if account_class == "ingreso":
                current_revenue += float(line.get("credit") or 0) - float(line.get("debit") or 0)
            elif account_class == "gasto":
                current_expenses += float(line.get("debit") or 0) - float(line.get("credit") or 0)
            if line["account_code"] == "3110":
                contributions += float(line.get("credit") or 0) - float(line.get("debit") or 0)
            elif line["account_code"] == "3510":
                withdrawals += float(line.get("debit") or 0) - float(line.get("credit") or 0)
    current_revenue = round(current_revenue, 2)
    current_expenses = round(current_expenses, 2)
    current_result = round(current_revenue - current_expenses, 2)

    cash_flow = {"operating": 0.0, "investing": 0.0, "financing": 0.0}
    opening_cash = 0.0
    for entry in entries:
        if entry["source"] != "movimiento_diario":
            continue
        cash_change = 0.0
        for line in entry["lines"]:
            if line["account_code"] in {"1111", "1112", "1113", "1119"}:
                cash_change += float(line.get("debit") or 0) - float(line.get("credit") or 0)
        if entry["date"] < period_start:
            opening_cash += cash_change
            continue
        if not period_start <= entry["date"] <= cutoff or cash_change == 0:
            continue
        codes = {line["account_code"] for line in entry["lines"]}
        if codes & {"3110", "3510", "2210"}:
            cash_flow["financing"] += cash_change
        elif codes & {"1210"}:
            cash_flow["investing"] += cash_change
        else:
            cash_flow["operating"] += cash_change
    for key in cash_flow:
        cash_flow[key] = round(cash_flow[key], 2)
    net_cash = round(sum(cash_flow.values()), 2)

    period_entries = [entry for entry in entries if entry["date"][:7] == f"{year}-{month:02d}"]
    return {
        "chart_of_accounts": CHART_OF_ACCOUNTS,
        "journal_entries": period_entries,
        "ledger_balances": account_balances,
        "statement_of_financial_position": {
            "assets": assets,
            "liabilities": liabilities,
            "equity": equity,
            "liabilities_and_equity": round(liabilities + equity, 2),
            "balanced": abs(assets - liabilities - equity) < 0.01,
        },
        "changes_in_equity": {
            "opening_equity": round(equity - contributions + withdrawals - current_result, 2),
            "contributions": round(contributions, 2),
            "withdrawals": round(withdrawals, 2),
            "result": current_result,
            "closing_equity": equity,
        },
        "cash_flow": {
            **cash_flow,
            "opening_cash": round(opening_cash, 2),
            "net_change": net_cash,
            "closing_cash": round(opening_cash + net_cash, 2),
        },
        "result_summary": {"revenue": current_revenue, "expenses": current_expenses, "result": current_result},
        "pending_currency_conversion": pending_currency,
        "pending_classification": pending_classification,
        "pending_sale_cost_dates": pending_sale_cost_dates,
        "sale_cost_records": sorted(sale_costs, key=lambda item: _clean(item.get("date")), reverse=True),
        "cutoff_date": cutoff,
    }


def list_accounting(year: int | None = None, month: int | None = None) -> dict:
    _ensure_storage()
    now = datetime.now()
    year = int(year or now.year)
    month = int(month or now.month)
    movement_query = {"brand_id": "casa", "year": year}
    if month:
        movement_query["month"] = month
    movements = list(collection("accounting_movements").find(movement_query, {"_id": 0}).sort([("year", -1), ("month", -1), ("workday_number", -1), ("created_at", -1)]).limit(500))
    for movement in movements:
        movement["destination_account"] = _clean(movement.get("destination_account")) or _destination_account(
            movement.get("category"), movement.get("payment_method"), movement.get("direction")
        )
    statement_movements = list(collection("accounting_movements").find({
        "brand_id": "casa",
        "$or": [
            {"year": {"$lt": year}},
            {"year": year, "month": {"$lte": month}},
        ],
    }, {"_id": 0}))
    invoices = list(collection("supplier_invoices").find({"brand_id": "casa"}, {"_id": 0}).sort([("status", 1), ("supplier", 1), ("purchase_date", -1)]).limit(500))
    payments = list(collection("supplier_payments").find({"brand_id": "casa"}, {"_id": 0}).sort("created_at", -1).limit(250))
    adjustments = list(collection("accounting_adjustments").find({"brand_id": "casa"}, {"_id": 0}).sort("date", -1).limit(500))
    sale_costs = list(collection("accounting_sale_costs").find({"brand_id": "casa"}, {"_id": 0}).sort("date", -1).limit(500))
    reporting = _journal_and_statements(statement_movements, invoices, adjustments, sale_costs, year, month)
    control_cutoff = min(reporting["cutoff_date"], date.today().isoformat())
    control = _daily_control(invoices, statement_movements, adjustments, sale_costs, control_cutoff)
    conciliation_date = control["next_open_date"] or date.today().isoformat()
    return {
        "year": year,
        "month": month,
        "movements": movements,
        "account_balances": _account_balances(statement_movements),
        "supplier_invoices": invoices,
        "supplier_payments": payments,
        "adjustments": adjustments,
        "monthly_results": monthly_results(year),
        "annual_result": annual_result(year),
        "daily_control": control,
        "card_conciliation": conciliate_cards(conciliation_date),
        "proposed_card_settlements": list_proposed_card_settlements()["proposed"],
        "till_handovers": list_till_handovers(year, month),
        **reporting,
    }


def _daily_cash_summary(all_movements: list[dict], report_date: str) -> dict:
    """Efectivo del dia: saldo inicial + entradas − salidas, contando solo
    movimientos en UYU con medio efectivo. Fuente unica de verdad para el
    reporte diario y para el arqueo de la entrega de caja."""
    def cash_effect(movement: dict) -> float:
        if movement.get("currency") != "UYU" or movement.get("payment_method") != "efectivo":
            return 0.0
        amount = float(movement.get("amount") or 0)
        return -amount if movement.get("direction") == "expense" else amount

    day_movements = [movement for movement in all_movements if movement.get("date") == report_date]
    opening_balance = sum(cash_effect(movement) for movement in all_movements if movement.get("date", "") < report_date)
    opening_movements = [movement for movement in day_movements if movement.get("source") == "opening_balance"]
    opening_balance += sum(cash_effect(movement) for movement in opening_movements)
    report_movements = [movement for movement in day_movements if movement.get("source") != "opening_balance"]
    cash_income = sum(cash_effect(movement) for movement in report_movements if cash_effect(movement) > 0)
    cash_expenses = -sum(cash_effect(movement) for movement in report_movements if cash_effect(movement) < 0)
    return {
        "opening_balance": opening_balance,
        "cash_income": cash_income,
        "cash_expenses": cash_expenses,
        "closing_balance": opening_balance + cash_income - cash_expenses,
        "report_movements": report_movements,
    }


def _theoretical_cash_close(report_date: str) -> float:
    all_movements = list(collection("accounting_movements").find(
        {"brand_id": "casa", "date": {"$lte": report_date}},
        {"_id": 0},
    ).sort([("date", 1), ("created_at", 1)]))
    return round(_daily_cash_summary(all_movements, report_date)["closing_balance"], 2)


def _fiserv_coupon_medio(product_name: object) -> str | None:
    name = _normalize_text(product_name)
    if not name:
        return None
    if "maestro" in name:
        return "maestro"
    is_credito = "credit" in name  # credito / credit
    if "visa" in name:
        return "visa_credito" if is_credito else "visa_debito"
    if "master" in name:
        # Debit Mastercard, Mastercard prepago y Mastercard debito -> master_debito
        return "master_credito" if is_credito else "master_debito"
    return None


def _movement_card_medio(movement: dict) -> str | None:
    method = _clean(movement.get("payment_method")).lower()
    card_type = _clean(movement.get("card_payment_type")).lower()
    if method == "maestro":
        return "maestro"
    if method == "visa":
        return "visa_credito" if card_type == "credito" else "visa_debito"
    if method == "master":
        return "master_credito" if card_type == "credito" else "master_debito"
    return None


def _card_income_by_medio(report_date: str) -> dict[str, float]:
    totals = {medio: 0.0 for medio in CARD_CONCILIATION_MEDIOS}
    for movement in collection("accounting_movements").find(
        {"brand_id": "casa", "date": report_date, "direction": "income", "payment_method": {"$in": list(CARD_PAYMENT_METHODS)}},
        {"_id": 0},
    ):
        medio = _movement_card_medio(movement)
        if medio:
            totals[medio] += float(movement.get("amount") or 0)
    return totals


def _coupon_integrity(report_date: str, coupons: list[dict]) -> dict:
    counts: dict[str, int] = {}
    for coupon in coupons:
        bill = _clean(coupon.get("bill_number"))
        counts[bill] = counts.get(bill, 0) + 1
    sales_numbers = [
        number
        for movement in collection("accounting_movements").find(
            {"brand_id": "casa", "date": report_date, "direction": "income", "category": {"$in": ["facturas", "factura_credito"]}},
            {"_id": 0, "invoice_number": 1},
        )
        if (number := _numeric(movement.get("invoice_number"))) is not None
    ]
    low = min(sales_numbers) if sales_numbers else None
    high = max(sales_numbers) if sales_numbers else None
    rows: list[dict] = []
    for coupon in coupons:
        bill = _clean(coupon.get("bill_number"))
        issues: list[str] = []
        if not bill:
            issues.append("empty")
        else:
            if counts.get(bill, 0) > 1:
                issues.append("duplicated")
            number = _numeric(bill)
            if low is not None and number is not None and not low <= number <= high:
                issues.append("out_of_range")
        rows.append({
            "fiserv_id": coupon.get("fiserv_id"),
            "bill_number": bill,
            "ticket": coupon.get("ticket"),
            "batch": coupon.get("batch"),
            "product_name": coupon.get("product_name"),
            "amount": round(float(coupon.get("total_amount") or 0), 2),
            "issues": issues,
        })
    return {
        "coupons": rows,
        "flagged": [row for row in rows if row["issues"]],
        "registered_bill_range": {"min": low, "max": high} if low is not None else None,
    }


def conciliate_cards(report_date: str) -> dict:
    """Concilia los cupones del POS Fiserv contra los totales declarados por el
    cajero (o, si aun no entrego caja, contra los cobros con tarjeta ya
    registrados). Por medio, no por fecha de lote."""
    _ensure_storage()
    report_date = _required_date(report_date, "fecha de conciliacion")
    coupons = list(collection("fiserv_transactions").find(
        {"sale_date": report_date, "transaction_type": "C", "state": "2"},
        {"_id": 0, "card_number": 0},
    ))
    pos_totals = {medio: 0.0 for medio in CARD_CONCILIATION_MEDIOS}
    unmapped_pos_total = 0.0
    for coupon in coupons:
        medio = _fiserv_coupon_medio(coupon.get("product_name"))
        gross = float(coupon.get("total_amount") or 0)
        if medio:
            pos_totals[medio] += gross
        else:
            unmapped_pos_total += gross
    handover = collection("accounting_till_handovers").find_one({"brand_id": "casa", "date": report_date}, {"_id": 0})
    if handover and handover.get("card_totals"):
        caja_totals = {medio: float((handover.get("card_totals") or {}).get(medio) or 0) for medio in CARD_CONCILIATION_MEDIOS}
        caja_source = "handover"
    else:
        caja_totals = _card_income_by_medio(report_date)
        caja_source = "movements"
    per_medio = []
    for medio in CARD_CONCILIATION_MEDIOS:
        pos = round(pos_totals[medio], 2)
        caja = round(caja_totals[medio], 2)
        difference = round(pos - caja, 2)
        flag = ""
        if difference > CARD_CONCILIATION_EPSILON:
            flag = "faltante"
        elif difference < -CARD_CONCILIATION_EPSILON:
            flag = "sobrante"
        per_medio.append({"medio": medio, "pos_total": pos, "caja_total": caja, "difference": difference, "flag": flag})
    return {
        "date": report_date,
        "has_coupons": bool(coupons),
        "pending_sync": not coupons,
        "caja_source": caja_source,
        "coupon_count": len(coupons),
        "per_medio": per_medio,
        "unmapped_pos_total": round(unmapped_pos_total, 2),
        "has_faltante": any(item["flag"] == "faltante" for item in per_medio),
        "has_sobrante": any(item["flag"] == "sobrante" for item in per_medio),
        "coupon_integrity": _coupon_integrity(report_date, coupons),
    }


def register_till_handover(data: dict, user: str) -> dict:
    """Entrega de caja + arqueo: un acto del cajero, distinto del cierre
    contable administrativo. Nunca autocorrige: registra lo contado y la
    diferencia, y devuelve la conciliacion de tarjetas en vivo."""
    _ensure_storage()
    handover_date = _required_date(data.get("date"), "fecha de entrega de caja")
    raw_counted = data.get("counted_cash")
    if raw_counted is None or _clean(raw_counted) == "":
        raise ValueError("Falta el efectivo contado")
    counted_cash = _money(raw_counted)
    override = bool(data.get("override"))
    existing = collection("accounting_till_handovers").find_one({"brand_id": "casa", "date": handover_date}, {"_id": 0})
    if existing and not override:
        raise ValueError(f"La caja del {handover_date} ya fue entregada. Usa override para volver a registrarla.")
    theoretical_cash = _theoretical_cash_close(handover_date)
    card_totals_in = data.get("card_totals") or {}
    card_totals = {medio: _money(card_totals_in.get(medio)) for medio in CARD_CONCILIATION_MEDIOS}
    raw_ticket_total = data.get("ticket_close_total")
    ticket_close_total = _money(raw_ticket_total) if raw_ticket_total not in (None, "") else None
    now = _now()
    handover = {
        "id": existing["id"] if existing else str(uuid4()),
        "brand_id": "casa",
        "date": handover_date,
        "counted_cash": counted_cash,
        "theoretical_cash": theoretical_cash,
        "difference": round(counted_cash - theoretical_cash, 2),
        "card_totals": card_totals,
        "pos_batches": [_clean(batch) for batch in (data.get("pos_batches") or []) if _clean(batch)],
        "ticket_close_total": ticket_close_total,
        "cashier": user,
        "overridden": bool(existing),
        "override_count": (int(existing.get("override_count") or 0) + 1) if existing else 0,
        "updated_at": now,
        "updated_by": user,
    }
    collection("accounting_till_handovers").update_one(
        {"brand_id": "casa", "date": handover_date},
        {"$set": handover, "$setOnInsert": {"created_at": now, "created_by": user}},
        upsert=True,
    )
    saved = collection("accounting_till_handovers").find_one({"brand_id": "casa", "date": handover_date}, {"_id": 0}) or handover
    return {"handover": saved, "conciliation": conciliate_cards(handover_date)}


def list_till_handovers(year: int, month: int) -> list[dict]:
    _ensure_storage()
    start = date(year, month, 1).isoformat()
    end = date(year, month, monthrange(year, month)[1]).isoformat()
    return list(collection("accounting_till_handovers").find(
        {"brand_id": "casa", "date": {"$gte": start, "$lte": end}},
        {"_id": 0},
    ).sort("date", -1))


def _settlement_payment_method(row: dict) -> str:
    text = _normalize_text(f"{row.get('product_desc')} {row.get('product_code')}")
    if "maestro" in text:
        return "maestro"
    if "master" in text:
        return "master"
    if "visa" in text:
        return "visa"
    raise ValueError(f"No se pudo determinar la tarjeta de la liquidacion {row.get('settlement_number')}")


def list_proposed_card_settlements() -> dict:
    """Cada liquidacion Fiserv con neto>0 que aun no tiene su acreditacion
    (financiera -> banco) registrada. Match por settlement_number."""
    _ensure_storage()
    registered: set[str] = set()
    for movement in collection("accounting_movements").find(
        {"brand_id": "casa", "category": CARD_SETTLEMENT_CATEGORY},
        {"_id": 0, "source_key": 1},
    ):
        source_key = _clean(movement.get("source_key"))
        if source_key.startswith("fiserv-settlement:"):
            registered.add(source_key.split(":", 1)[1])
    by_number: dict[str, dict] = {}
    for row in collection("fiserv_settlements").find({"acquirer": "fiserv", "net_amount": {"$gt": 0}}, {"_id": 0}):
        number = _clean(row.get("settlement_number"))
        if not number or number in registered:
            continue
        agg = by_number.setdefault(number, {
            "settlement_number": number,
            "payment_date": _clean(row.get("payment_date")),
            "product_desc": _clean(row.get("product_desc")),
            "net_amount": 0.0,
            "already_registered": False,
        })
        agg["net_amount"] = round(agg["net_amount"] + float(row.get("net_amount") or 0), 2)
    proposed = sorted(by_number.values(), key=lambda item: (item["payment_date"], item["settlement_number"]))
    return {"proposed": proposed}


def confirm_card_settlement(settlement_number: str, user: str) -> dict:
    _ensure_storage()
    number = _clean(settlement_number)
    if not number:
        raise ValueError("Falta el numero de liquidacion")
    rows = list(collection("fiserv_settlements").find({"acquirer": "fiserv", "settlement_number": number}, {"_id": 0}))
    if not rows:
        raise ValueError(f"No existe la liquidacion {number}")
    net_amount = round(sum(float(row.get("net_amount") or 0) for row in rows), 2)
    if net_amount <= 0:
        raise ValueError("La liquidacion no tiene neto a acreditar")
    payment_date = next((_clean(row.get("payment_date")) for row in rows if _clean(row.get("payment_date"))), "")
    if not payment_date:
        raise ValueError("La liquidacion no tiene fecha de pago")
    payment_method = _settlement_payment_method(rows[0])
    product_desc = _clean(rows[0].get("product_desc")) or number
    source_key = f"fiserv-settlement:{number}"
    # register_movement validates card_receivables before deduping by source_key,
    # so a re-confirm would raise "no puede superar" once the balance dropped.
    # Guard idempotency here.
    already = collection("accounting_movements").find_one({"source_key": source_key}, {"_id": 0})
    if already:
        return {"settlement_number": number, "net_amount": net_amount, "movement": already, "already_registered": True}
    result = register_movement({
        "direction": "transfer",
        "category": CARD_SETTLEMENT_CATEGORY,
        "payment_method": payment_method,
        "amount": net_amount,
        "currency": "UYU",
        "date": payment_date,
        "year": int(payment_date[:4]),
        "month": int(payment_date[5:7]),
        "description": f"Acreditacion Fiserv liquidacion {number} ({product_desc})",
        "reference": number,
        "source": "fiserv_settlement",
        "source_key": source_key,
    }, user)
    return {"settlement_number": number, "net_amount": net_amount, **result}


def export_daily_report(report_date: str, cashier: str) -> dict:
    _ensure_storage()
    try:
        selected_date = date.fromisoformat(_clean(report_date))
    except ValueError as exc:
        raise ValueError("Fecha de reporte invalida") from exc

    all_movements = list(collection("accounting_movements").find(
        {"brand_id": "casa", "date": {"$lte": selected_date.isoformat()}},
        {"_id": 0},
    ).sort([("date", 1), ("created_at", 1)]))
    cash = _daily_cash_summary(all_movements, selected_date.isoformat())
    opening_balance = cash["opening_balance"]
    cash_income = cash["cash_income"]
    cash_expenses = cash["cash_expenses"]
    closing_balance = cash["closing_balance"]
    report_movements = cash["report_movements"]

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte diario"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A10"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = "1:9"
    sheet.oddFooter.center.text = "Pagina &P de &N"
    sheet.oddFooter.right.text = "Firma cajero: ____________________"

    dark_green = "174C45"
    pale_green = "E8F3EF"
    light_gray = "F3F4F6"
    border_color = "D1D5DB"
    thin = Side(style="thin", color=border_color)

    sheet.merge_cells("A1:N1")
    sheet["A1"] = "LA CASA DEL CARPINTERO - REPORTE DIARIO DE CAJA"
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=dark_green)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30

    sheet["A3"] = "Fecha"
    sheet["B3"] = selected_date
    sheet["B3"].number_format = "dd/mm/yyyy"
    sheet["D3"] = "Cajero"
    sheet.merge_cells("E3:F3")
    sheet["E3"] = _clean(cashier) or "Sin identificar"
    sheet["H3"] = "Emitido"
    sheet.merge_cells("I3:J3")
    sheet["I3"] = _now().astimezone().replace(tzinfo=None)
    sheet["I3"].number_format = "dd/mm/yyyy hh:mm"
    for cell in ("A3", "D3", "H3"):
        sheet[cell].font = Font(bold=True, color=dark_green)

    summary = [
        ("Saldo inicial efectivo", opening_balance),
        ("Entradas efectivo", cash_income),
        ("Salidas efectivo", cash_expenses),
        ("Saldo final efectivo", closing_balance),
    ]
    summary_spans = [(1, 3), (4, 6), (7, 9), (10, 14)]
    for (label, amount), (start_col, end_col) in zip(summary, summary_spans, strict=True):
        label_cell = sheet.cell(row=5, column=start_col, value=label)
        value_cell = sheet.cell(row=6, column=start_col, value=amount)
        sheet.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col)
        sheet.merge_cells(start_row=6, start_column=start_col, end_row=6, end_column=end_col)
        label_cell.fill = PatternFill("solid", fgColor=pale_green)
        label_cell.font = Font(size=10, bold=True, color=dark_green)
        label_cell.alignment = Alignment(horizontal="center")
        value_cell.font = Font(size=15, bold=True, color=dark_green)
        value_cell.alignment = Alignment(horizontal="center")
        value_cell.number_format = '[$$-es-UY] #,##0.00'
        for row in (5, 6):
            for column in range(start_col, end_col + 1):
                sheet.cell(row=row, column=column).border = Border(top=thin, bottom=thin, left=thin, right=thin)

    headers = ["Hora", "Tipo", "Categoria", "Subcategoria", "Medio", "Destino", "Factura", "Emision", "Vencimiento", "Descripcion", "Referencia", "Moneda", "Entrada", "Salida"]
    sheet.append([])
    sheet.append(headers)
    header_row = sheet.max_row
    for cell in sheet[header_row]:
        cell.fill = PatternFill("solid", fgColor=dark_green)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    direction_labels = {"income": "Entrada", "expense": "Salida", "transfer": "Transferencia"}
    for movement in report_movements:
        created_at = movement.get("created_at")
        time_value = created_at.astimezone().replace(tzinfo=None) if isinstance(created_at, datetime) else None
        amount = float(movement.get("amount") or 0)
        is_expense = movement.get("direction") == "expense"
        is_transfer = movement.get("direction") == "transfer"
        sheet.append([
            time_value,
            direction_labels.get(movement.get("direction"), movement.get("direction", "")),
            movement.get("category", ""),
            movement.get("subcategory", ""),
            _payment_method_detail(movement),
            "financiera -> banco" if movement.get("category") == CARD_SETTLEMENT_CATEGORY else _clean(movement.get("destination_account")) or _destination_account(movement.get("category"), movement.get("payment_method"), movement.get("direction")),
            movement.get("invoice_number", ""),
            movement.get("issue_date", ""),
            movement.get("due_date", ""),
            movement.get("description", ""),
            movement.get("reference", ""),
            movement.get("currency", "UYU"),
            amount if not is_expense else None,
            amount if is_expense or is_transfer else None,
        ])
        row = sheet.max_row
        sheet.cell(row=row, column=1).number_format = "hh:mm"
        sheet.cell(row=row, column=13).number_format = '#,##0.00'
        sheet.cell(row=row, column=14).number_format = '#,##0.00'
        fill = PatternFill("solid", fgColor="FFFFFF" if row % 2 else light_gray)
        for cell in sheet[row]:
            cell.fill = fill
            cell.border = Border(bottom=Side(style="hair", color=border_color))
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in {4, 5, 6, 7, 10, 11})

    if not report_movements:
        sheet.merge_cells(start_row=header_row + 1, start_column=1, end_row=header_row + 2, end_column=14)
        empty_cell = sheet.cell(row=header_row + 1, column=1, value="Sin movimientos operativos registrados para esta fecha.")
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")
        empty_cell.font = Font(italic=True, color="6B7280")

    totals_row = max(sheet.max_row + 2, header_row + 4)
    sheet.merge_cells(start_row=totals_row, start_column=1, end_row=totals_row, end_column=12)
    sheet.cell(row=totals_row, column=1, value="TOTALES DEL DIA").font = Font(bold=True, color=dark_green)
    income_rows = [float(m.get("amount") or 0) for m in report_movements if m.get("direction") in {"income", "transfer"} and m.get("currency") == "UYU"]
    expense_rows = [float(m.get("amount") or 0) for m in report_movements if m.get("direction") in {"expense", "transfer"} and m.get("currency") == "UYU"]
    sheet.cell(row=totals_row, column=13, value=sum(income_rows)).number_format = '#,##0.00'
    sheet.cell(row=totals_row, column=14, value=sum(expense_rows)).number_format = '#,##0.00'
    for cell in sheet[totals_row]:
        cell.fill = PatternFill("solid", fgColor=pale_green)
        cell.font = Font(bold=True, color=dark_green)
        cell.border = Border(top=Side(style="medium", color=dark_green))

    signature_row = totals_row + 4
    sheet.merge_cells(start_row=signature_row, start_column=1, end_row=signature_row, end_column=5)
    sheet.merge_cells(start_row=signature_row, start_column=9, end_row=signature_row, end_column=14)
    sheet.cell(row=signature_row, column=1, value="Firma del cajero: __________________________________")
    sheet.cell(row=signature_row, column=9, value="Firma del responsable: ____________________________")
    sheet.cell(row=signature_row + 2, column=1, value="Aclaracion: ______________________________________")
    sheet.cell(row=signature_row + 2, column=9, value="Aclaracion: ______________________________________")

    widths = [9, 13, 16, 16, 26, 20, 16, 13, 13, 28, 18, 9, 14, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.auto_filter.ref = f"A{header_row}:N{max(header_row, totals_row - 2)}"
    sheet.print_area = f"A1:N{signature_row + 3}"

    path = Path(gettempdir()) / f"reporte-caja-{selected_date.isoformat()}-{uuid4().hex[:8]}.xlsx"
    workbook.save(path)
    return {"excel_path": str(path), "filename": f"reporte-caja-{selected_date.isoformat()}.xlsx"}


def monthly_results(year: int) -> list[dict]:
    _ensure_storage()
    invoices = list(collection("supplier_invoices").find({"brand_id": "casa"}, {"_id": 0}))
    adjustments = list(collection("accounting_adjustments").find({"brand_id": "casa"}, {"_id": 0}))
    sale_costs = list(collection("accounting_sale_costs").find({"brand_id": "casa"}, {"_id": 0}))
    invoices_by_id = {_clean(invoice.get("id")): invoice for invoice in invoices}
    classified_invoice_ids = {
        _clean(invoice.get("id"))
        for invoice in invoices
        if _clean(invoice.get("accounting_classification")).lower() in SUPPLIER_CLASSIFICATION_ACCOUNTS
    }
    rows: list[dict] = []
    for month in range(1, 13):
        movements = list(collection("accounting_movements").find({"brand_id": "casa", "year": year, "month": month}, {"_id": 0}))
        period = f"{year}-{month:02d}"
        period_invoices = [invoice for invoice in invoices if _clean(invoice.get("purchase_date"))[:7] == period]
        period_adjustments = [adjustment for adjustment in adjustments if _clean(adjustment.get("date"))[:7] == period]
        period_sale_costs = [record for record in sale_costs if _clean(record.get("date"))[:7] == period]
        entries = [entry for movement in movements if (entry := _movement_entry(movement, classified_invoice_ids, invoices_by_id))]
        entries.extend(entry for invoice in period_invoices if (entry := _supplier_invoice_entry(invoice)))
        entries.extend(entry for adjustment in period_adjustments if (entry := _adjustment_entry(adjustment)))
        entries.extend(entry for record in period_sale_costs if (entry := _sale_cost_entry(record)))

        sales = [m for m in movements if m.get("direction") == "income" and m.get("category") in {"facturas", "factura_credito"}]
        functional_value = lambda movement: float(movement.get("functional_amount") or (movement.get("amount") if _safe_currency(movement.get("currency")) == "UYU" else 0) or 0)
        card_sales = sum(functional_value(m) for m in sales if m.get("payment_method") in CARD_PAYMENT_METHODS)
        bank_sales = sum(functional_value(m) for m in sales if m.get("payment_method") in BANK_PAYMENT_METHODS)
        credit_sales = sum(functional_value(m) for m in sales if m.get("category") == "factura_credito")
        cash_sales = sum(functional_value(m) for m in sales if m.get("payment_method") == "efectivo")

        account_changes: dict[str, float] = {}
        for entry in entries:
            for line in entry["lines"]:
                code = line["account_code"]
                account_changes[code] = account_changes.get(code, 0.0) + float(line.get("debit") or 0) - float(line.get("credit") or 0)
        gross_sales = -account_changes.get("4110", 0.0)
        other_income = -account_changes.get("4210", 0.0)
        supplier_costs = account_changes.get("5110", 0.0)
        payroll = sum(account_changes.get(code, 0.0) for code in {"5210", "5220", "5230", "5240"})
        fixed_costs = sum(account_changes.get(code, 0.0) for code in {"5310", "5410", "5420"})
        expenses = sum(change for code, change in account_changes.items() if ACCOUNT_BY_CODE[code]["class"] == "gasto")
        other_costs = expenses - supplier_costs - payroll - fixed_costs
        rows.append({
            "year": year,
            "month": month,
            "gross_sales": round(gross_sales, 2),
            "cash_sales": round(cash_sales, 2),
            "card_sales": round(card_sales, 2),
            "bank_sales": round(bank_sales, 2),
            "credit_sales": round(credit_sales, 2),
            "fixed_costs": round(fixed_costs, 2),
            "other_costs": round(other_costs, 2),
            "payroll": round(payroll, 2),
            "supplier_costs": round(supplier_costs, 2),
            "total_costs": round(expenses, 2),
            "operating_result": round(gross_sales + other_income - expenses, 2),
            "movement_count": len(entries),
        })
    return rows


def annual_result(year: int) -> dict:
    rows = monthly_results(year)
    keys = ["gross_sales", "cash_sales", "card_sales", "bank_sales", "credit_sales", "fixed_costs", "other_costs", "payroll", "supplier_costs", "total_costs", "operating_result"]
    return {"year": year, **{key: round(sum(float(row[key]) for row in rows), 2) for key in keys}}
