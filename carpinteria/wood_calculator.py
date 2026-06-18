from __future__ import annotations

import csv
import math
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from carpinteria.quote_router import WOOD_SPECIES, norm_text
from carpinteria.schemas import Quotation, QuotationLine

try:
    from carpinteria.settings import (
        LABOR_DAY_PRICE_UYU,
        MACHINERY_PERCENT,
        PROFIT_PERCENT,
        WASTE_PERCENT,
    )
except Exception:
    LABOR_DAY_PRICE_UYU = 2500
    MACHINERY_PERCENT = 7.5
    PROFIT_PERCENT = 65
    WASTE_PERCENT = 15


# Excel maestro del dueño (solo existe en su Windows). En Mac/Railway no existe,
# así que se cae al CSV commiteado (fuente de verdad cross-platform), que se
# regenera con scripts/flatten_price_sheets.py.
WOOD_PRICE_PATHS = (
    Path(r"C:\Users\Peluca\Documents\La casa del Carpintero\Cotizador_Madera_V2_Corregido.xlsx"),
    Path(r"C:\Users\Peluca\Downloads\Cotizador_Madera_V2_Corregido (2).xlsx"),
    Path(r"C:\Users\Peluca\Downloads\Cotizador_Madera_V2_Corregido.xlsx"),
)
WOOD_DATA_CSV = Path(__file__).resolve().parent / "data" / "wood_datos.csv"


@dataclass(frozen=True)
class WoodMaterial:
    id: str
    species: str
    features: str
    thickness_in: float
    length_m: float
    width_in: float
    price_uyu: float
    supplier: str = ""

    @property
    def width_cm_for_quote(self) -> float:
        # The existing woodworking sheet uses 1" = 2.25 cm for board coverage.
        return self.width_in * 2.25

    @property
    def price_per_meter_uyu(self) -> float:
        return self.price_uyu / self.length_m if self.length_m else 0.0


FALLBACK_WOOD_MATERIALS = (
    WoodMaterial("Pino Clear 1'", "Pino", "Clear", 1, 3.3, 6, 304.5, "fallback"),
    WoodMaterial("Pino Clear 3'", "Pino", "Clear", 3, 3.3, 3, 409.5, "fallback"),
    WoodMaterial("Euca Clear 1'", "Euca", "Clear", 1, 3.3, 6, 561, "fallback"),
    WoodMaterial("Roble Americano 1'", "Roble", "Americano", 1, 2.5, 11.11111111, 2583, "fallback"),
    WoodMaterial("Cedro Mara Clear 1'", "Cedro Mara", "Clear", 1, 2.4, 8.888888889, 1890, "fallback"),
    WoodMaterial("Abeto Clear 1.5'", "Abeto", "Clear", 1.5, 2.25, 10, 815.85, "fallback"),
)


def _num(value: object) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _build_material(
    *, id: object, species: object, features: object,
    thickness: object, length: object, width: object, price: object, supplier: object,
) -> WoodMaterial | None:
    """Construye un WoodMaterial validando especie y campos. Devuelve None si no aplica."""
    if norm_text(str(species or "")) not in {norm_text(s) for s in WOOD_SPECIES}:
        return None
    material = WoodMaterial(
        id=str(id or "").strip(),
        species=str(species or "").strip(),
        features=str(features or "").strip(),
        thickness_in=_num(thickness),
        length_m=_num(length),
        width_in=_num(width),
        price_uyu=_num(price),
        supplier=str(supplier or "").strip(),
    )
    if material.id and material.thickness_in and material.length_m and material.width_in and material.price_uyu:
        return material
    return None


def _materials_from_xlsx(path: Path) -> list[WoodMaterial]:
    wb = load_workbook(path, data_only=True, read_only=True)
    if "Datos" not in wb.sheetnames:
        return []
    ws = wb["Datos"]
    out: list[WoodMaterial] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        def col(i: int) -> object:
            return row[i] if len(row) > i else ""
        material = _build_material(
            id=col(0), species=col(1), features=col(2),
            thickness=col(3), length=col(5), width=col(7), price=col(9), supplier=col(11),
        )
        if material:
            out.append(material)
    return out


def _materials_from_csv(path: Path) -> list[WoodMaterial]:
    if not path.exists():
        return []
    out: list[WoodMaterial] = []
    with path.open(encoding="utf-8", newline="") as fh:
        for row in csv.DictReader(fh):
            material = _build_material(
                id=row.get("id"), species=row.get("species"), features=row.get("features"),
                thickness=row.get("thickness_in"), length=row.get("length_m"),
                width=row.get("width_in"), price=row.get("price_uyu"), supplier=row.get("supplier"),
            )
            if material:
                out.append(material)
    return out


def load_wood_materials() -> list[WoodMaterial]:
    """Precios de madera maciza. Prioridad: Excel maestro local (Windows del dueño)
    -> CSV commiteado (Mac/Railway) -> fallback hardcodeado."""
    path = next((p for p in WOOD_PRICE_PATHS if p.exists()), None)
    if path is not None:
        materials = _materials_from_xlsx(path)
        if materials:
            return materials
    materials = _materials_from_csv(WOOD_DATA_CSV)
    if materials:
        return materials
    return list(FALLBACK_WOOD_MATERIALS)


def _extract_species(text: str, material: str | None = None) -> str:
    normalized = norm_text(f"{text} {material or ''}")
    for species in ("cedro mara", "eucaliptus", "eucalipto", "euca", "roble", "abeto", "pino"):
        if species in normalized:
            return "euca" if species in {"eucaliptus", "eucalipto"} else species
    return "pino"


def _extract_inches(text: str, default: float = 1.0) -> float:
    normalized = norm_text(text)
    if "pulgada y media" in normalized or "una pulgada y media" in normalized or "1.5" in normalized or "1,5" in normalized:
        return 1.5
    if "una pulgada" in normalized or "1 pulgada" in normalized or "1'" in normalized:
        return 1.0
    match = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:pulgadas?|')", normalized)
    if match:
        return float(match.group(1).replace(",", "."))
    return default


def _extract_leg_section(text: str) -> tuple[float, float] | None:
    normalized = norm_text(text)
    match = re.search(r"patas?.{0,25}?(\d+(?:[.,]\d+)?)\s*x\s*(\d+(?:[.,]\d+)?)\s*(?:pulgadas?|')?", normalized)
    if not match:
        match = re.search(r"(\d+(?:[.,]\d+)?)\s*x\s*(\d+(?:[.,]\d+)?)\s*(?:pulgadas?|').{0,25}?patas?", normalized)
    if not match:
        return None
    return float(match.group(1).replace(",", ".")), float(match.group(2).replace(",", "."))


def _extract_quantity(text: str, default: int = 1) -> int:
    normalized = norm_text(text)
    match = re.search(
        r"\b(\d{1,4})\s*(?:cortes?|discos?|redondos?|redondas?|piezas?|tablas?)\b",
        normalized,
    )
    if match:
        return max(1, int(match.group(1)))
    return max(1, int(default or 1))


def _extract_round_cut_diameter_mm(text: str, width_mm: float | None = None, depth_mm: float | None = None) -> float | None:
    normalized = norm_text(text).replace(",", ".")
    if not any(token in normalized for token in ("redondo", "redonda", "circulo", "circular", "disco")):
        return None
    explicit = re.search(
        r"(?:diametro|diam|redond[oa]s?\s+de|discos?\s+de)\s*(\d+(?:\.\d+)?)\s*(mm|cm|mts?|metros?)?\b",
        normalized,
    )
    if explicit:
        value = float(explicit.group(1))
        unit = explicit.group(2) or "mm"
        if unit.startswith("cm"):
            return value * 10
        if unit.startswith("m"):
            return value * 1000
        return value
    pair = re.search(
        r"\b(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)\s*(mm|cm|mts?|metros?)?\b",
        normalized,
    )
    if pair:
        a = float(pair.group(1))
        b = float(pair.group(2))
        unit = pair.group(3) or "mm"
        value = max(a, b)
        if unit.startswith("cm"):
            return value * 10
        if unit.startswith("m"):
            return value * 1000
        return value
    if width_mm and depth_mm and abs(float(width_mm) - float(depth_mm)) <= max(float(width_mm), float(depth_mm)) * 0.08:
        return max(float(width_mm), float(depth_mm))
    return None


def _dimensions_from_text(text: str) -> tuple[float | None, float | None, float | None]:
    normalized = norm_text(text).replace(",", ".")
    cm_match = re.search(
        r"\b(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)\s*cm\b",
        normalized,
    )
    if cm_match:
        return float(cm_match.group(1)) * 10, float(cm_match.group(3)) * 10, float(cm_match.group(2)) * 10
    nums = [float(n) for n in re.findall(r"\b\d+(?:\.\d+)?\b", normalized)]
    meters = [n for n in nums if 0.1 <= n <= 5]
    if len(meters) >= 3:
        return meters[0] * 1000, meters[2] * 1000, meters[1] * 1000
    return None, None, None


def _match_material(
    materials: list[WoodMaterial],
    *,
    species: str,
    thickness_in: float,
    width_in: float | None = None,
) -> WoodMaterial:
    species_norm = norm_text(species)
    scoped = [m for m in materials if norm_text(m.species) in {species_norm, "euca" if species_norm.startswith("euca") else species_norm}]
    if not scoped:
        scoped = materials
    if width_in:
        exact_width = [m for m in scoped if abs(m.width_in - width_in) < 0.01 and abs(m.thickness_in - thickness_in) < 0.01]
        if exact_width:
            return exact_width[0]
    exact = [m for m in scoped if abs(m.thickness_in - thickness_in) < 0.01]
    if exact:
        return sorted(exact, key=lambda m: abs((width_in or m.width_in) - m.width_in))[0]
    thinner = [m for m in scoped if m.thickness_in < thickness_in]
    thicker = [m for m in scoped if m.thickness_in > thickness_in]
    if thinner and thicker:
        low = max(thinner, key=lambda m: m.thickness_in)
        high = min(thicker, key=lambda m: m.thickness_in)
        span = high.thickness_in - low.thickness_in
        ratio = (thickness_in - low.thickness_in) / span if span else 0
        price = low.price_uyu + (high.price_uyu - low.price_uyu) * ratio
        return WoodMaterial(
            id=f"{low.species or species} estimado {thickness_in:g}'",
            species=low.species or species,
            features="Estimado por interpolacion",
            thickness_in=thickness_in,
            length_m=low.length_m or high.length_m,
            width_in=width_in or low.width_in or high.width_in,
            price_uyu=round(price, 2),
            supplier="estimado",
        )
    return min(scoped, key=lambda m: abs(m.thickness_in - thickness_in))


def quote_solid_wood_table(
    *,
    description: str,
    name: str,
    quantity: int = 1,
    width_mm: float | None = None,
    height_mm: float | None = None,
    depth_mm: float | None = None,
    material: str | None = None,
    thickness_mm: float | None = None,
    waste_percent: float = WASTE_PERCENT,
    machinery_percent: float = MACHINERY_PERCENT,
    profit_percent: float = PROFIT_PERCENT,
    labor_day_price_uyu: float = LABOR_DAY_PRICE_UYU,
) -> Quotation:
    round_diameter_mm = _extract_round_cut_diameter_mm(description, width_mm=width_mm, depth_mm=depth_mm)
    if round_diameter_mm:
        return quote_round_wood_cuts(
            description=description,
            name=name,
            quantity=_extract_quantity(description, quantity),
            diameter_mm=round_diameter_mm,
            material=material,
            thickness_mm=thickness_mm,
            waste_percent=max(waste_percent, 20),
            machinery_percent=machinery_percent,
            profit_percent=profit_percent,
            labor_day_price_uyu=labor_day_price_uyu,
        )

    if not (width_mm and height_mm and depth_mm):
        parsed_w, parsed_h, parsed_d = _dimensions_from_text(description)
        width_mm = width_mm or parsed_w
        height_mm = height_mm or parsed_h
        depth_mm = depth_mm or parsed_d
    if not (width_mm and height_mm and depth_mm):
        return Quotation(notes="Faltan medidas para cotizar madera maciza: largo, ancho y altura.")

    species = _extract_species(description, material)
    top_thickness = _extract_inches(description, (thickness_mm or 25.4) / 25.4 if thickness_mm else 1.0)
    leg_section = _extract_leg_section(description)
    leg_thickness = leg_section[0] if leg_section else 3.0
    leg_width = leg_section[1] if leg_section else leg_thickness

    materials = load_wood_materials()
    top = _match_material(materials, species=species, thickness_in=top_thickness)
    legs = _match_material(materials, species=species, thickness_in=leg_thickness, width_in=leg_width)

    units = max(1, int(quantity or 1))
    top_length_m = width_mm / 1000
    top_width_cm = depth_mm / 10
    leg_length_m = height_mm / 1000
    leg_count = 4
    glue_extra_percent = 15.0

    raw_boards = top_width_cm / top.width_cm_for_quote if top.width_cm_for_quote else 0.0
    board_count = max(1, math.ceil(raw_boards))
    base_linear_m = board_count * top_length_m
    top_total_linear_m = base_linear_m * (1 + glue_extra_percent / 100)
    top_material = round(top_total_linear_m * top.price_per_meter_uyu * units, 2)

    legs_linear_m = leg_count * leg_length_m * units
    legs_material = round(legs_linear_m * legs.price_per_meter_uyu, 2)
    material_total = round(top_material + legs_material, 2)
    waste_amount = round(material_total * waste_percent / 100, 2)
    labor_days = 0.25 * units
    labor_amount = round(labor_days * labor_day_price_uyu, 2)
    machinery_base = material_total + waste_amount + labor_amount
    machinery_amount = round(machinery_base * machinery_percent / 100, 2)
    glue_amount = round((top_length_m * top_width_cm / 100) * 0.25 * 900 * units, 2)
    planing_amount = round(0.25 * labor_day_price_uyu * units, 2)

    lines = [
        QuotationLine(
            concept=f"{top.id} - tablas para tapa encolada ({board_count} tablas x {top_length_m:.2f}m + {glue_extra_percent:.0f}% agregado)",
            quantity=round(top_total_linear_m * units, 2),
            unit="metro",
            unit_price=round(top.price_per_meter_uyu, 2),
            subtotal=top_material,
        ),
        QuotationLine(
            concept=f"{legs.id} - patas {leg_thickness:g}x{leg_width:g} pulgadas ({leg_count} patas x {leg_length_m:.2f}m)",
            quantity=round(legs_linear_m, 2),
            unit="metro",
            unit_price=round(legs.price_per_meter_uyu, 2),
            subtotal=legs_material,
        ),
        QuotationLine(
            concept=f"Merma ({waste_percent:.0f}%)",
            quantity=1,
            unit="recargo",
            unit_price=waste_amount,
            subtotal=waste_amount,
        ),
        QuotationLine(
            concept=f"Mano de obra madera ({labor_days:.2f} dias x UYU {labor_day_price_uyu:.0f})",
            quantity=1,
            unit="recargo",
            unit_price=labor_amount,
            subtotal=labor_amount,
        ),
        QuotationLine(
            concept=f"Maquinaria / cargos fabriles ({machinery_percent:.0f}%)",
            quantity=1,
            unit="recargo",
            unit_price=machinery_amount,
            subtotal=machinery_amount,
        ),
        QuotationLine(
            concept="Encolado - adhesivo para formar tapa",
            quantity=1,
            unit="insumo",
            unit_price=glue_amount,
            subtotal=glue_amount,
        ),
        QuotationLine(
            concept="Cepillado / nivelado posterior al encolado",
            quantity=1,
            unit="proceso",
            unit_price=planing_amount,
            subtotal=planing_amount,
        ),
    ]
    subtotal = round(sum(line.subtotal for line in lines), 2)
    profit = round(subtotal * profit_percent / 100, 2)
    total = round(subtotal + profit, 2)
    return Quotation(
        lines=lines,
        subtotal=subtotal,
        margin_percent=profit_percent,
        margin_amount=profit,
        total=total,
        notes="Cotizacion por madera maciza: se forma la tapa encolando tablas. No usa placas ni cantos.",
        metadata={
            "quote_type": "madera_maciza",
            "subtype": "tablas_encoladas",
            "top_material": top.__dict__,
            "leg_material": legs.__dict__,
            "raw_boards_for_width": raw_boards,
            "board_count_for_width": board_count,
            "base_linear_m_top": base_linear_m,
            "top_total_linear_m": top_total_linear_m,
        },
    )


def quote_round_wood_cuts(
    *,
    description: str,
    name: str,
    quantity: int = 1,
    diameter_mm: float,
    material: str | None = None,
    thickness_mm: float | None = None,
    waste_percent: float = 20,
    machinery_percent: float = MACHINERY_PERCENT,
    profit_percent: float = PROFIT_PERCENT,
    labor_day_price_uyu: float = LABOR_DAY_PRICE_UYU,
) -> Quotation:
    species = _extract_species(description, material)
    thickness_in = _extract_inches(description, (thickness_mm or 38.1) / 25.4 if thickness_mm else 1.5)
    materials = load_wood_materials()
    board = _match_material(materials, species=species, thickness_in=thickness_in)

    units = _extract_quantity(description, quantity)
    diameter_m = diameter_mm / 1000
    diameter_cm = diameter_mm / 10
    strips_per_disk = max(1, math.ceil(diameter_cm / board.width_cm_for_quote))
    base_linear_m = strips_per_disk * diameter_m * units
    total_linear_m = base_linear_m * (1 + waste_percent / 100)
    material_amount = round(total_linear_m * board.price_per_meter_uyu, 2)
    cut_labor_days = max(0.25, 0.025 * units)
    cut_labor = round(cut_labor_days * labor_day_price_uyu, 2)
    machinery_amount = round((material_amount + cut_labor) * machinery_percent / 100, 2)

    lines = [
        QuotationLine(
            concept=f"{board.id} - madera para {units} cortes redondos de {diameter_cm:.1f}cm ({strips_per_disk} tiras por corte + {waste_percent:.0f}% merma)",
            quantity=round(total_linear_m, 2),
            unit="metro",
            unit_price=round(board.price_per_meter_uyu, 2),
            subtotal=material_amount,
        ),
        QuotationLine(
            concept=f"Corte redondo / calado ({units} unidades)",
            quantity=units,
            unit="unidad",
            unit_price=round(cut_labor / units, 2),
            subtotal=cut_labor,
        ),
        QuotationLine(
            concept=f"Maquinaria / cargos fabriles ({machinery_percent:.0f}%)",
            quantity=1,
            unit="recargo",
            unit_price=machinery_amount,
            subtotal=machinery_amount,
        ),
    ]
    subtotal = round(sum(line.subtotal for line in lines), 2)
    profit = round(subtotal * profit_percent / 100, 2)
    total = round(subtotal + profit, 2)
    return Quotation(
        lines=lines,
        subtotal=subtotal,
        margin_percent=profit_percent,
        margin_amount=profit,
        total=total,
        notes=(
            "Cotizacion por cortes redondos de madera maciza. "
            "Se calcula como piezas armadas desde tiras/tablas por ancho util; no usa placas ni cantos."
        ),
        metadata={
            "quote_type": "madera_maciza",
            "subtype": "cortes_redondos",
            "wood_material": board.__dict__,
            "units": units,
            "diameter_mm": diameter_mm,
            "strips_per_disk": strips_per_disk,
            "base_linear_m": base_linear_m,
            "total_linear_m": total_linear_m,
        },
    )
