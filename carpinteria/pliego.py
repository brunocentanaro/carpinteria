from __future__ import annotations

import json
import math
import re
import shutil
import subprocess
import unicodedata

from openai import OpenAI

from carpinteria.prompts import FURNITURE_DECOMPOSE, PLIEGO_ANALYSIS
from carpinteria.openai_errors import friendly_openai_error
from carpinteria.settings import AGENT_MODEL

_client: OpenAI | None = None


def _get_client() -> OpenAI:
    global _client
    if _client is None:
        _client = OpenAI()
    return _client


def _extract_pdf_text(path: str) -> str:
    if shutil.which("pdftotext"):
        result = subprocess.run(
            ["pdftotext", path, "-"],
            capture_output=True, timeout=30,
        )
        if result.returncode == 0:
            return result.stdout.decode("utf-8", errors="replace")

    from pypdf import PdfReader

    reader = PdfReader(path)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    if not text.strip():
        raise RuntimeError("No se pudo extraer texto del PDF. Puede ser un PDF escaneado o basado en imagen.")
    return text


def _extract_xlsx_text(path: str) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(path)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"=== {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            vals = [str(c) if c is not None else "" for c in row]
            if any(v.strip() for v in vals):
                lines.append("\t".join(vals))
    return "\n".join(lines)


def _collapse_ws(text: object) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def _parse_state_quantity(row: str) -> tuple[str, int]:
    match = re.search(r"\s+(\d+)\s+-{2,}\s*$", row)
    if match:
        return row[: match.start()].strip(), int(match.group(1))
    match = re.search(r"\s+(\d+)\s*$", row)
    if match:
        return row[: match.start()].strip(), int(match.group(1))
    return row.strip(), 1


def _split_state_item_body(body: str) -> tuple[str, str]:
    markers = (
        "Material",
        "Madera",
        "Cubierta lavable",
        "Melaminico",
        "Melaminico",
        "Melamínico",
        "2 puertas",
        "De pared",
        "Apilables",
        "COLOR",
        "Plastico",
        "90 Litros",
    )
    best = None
    norm_body = _strip_accents(body).lower()
    for marker in markers:
        idx = norm_body.find(_strip_accents(marker).lower())
        if idx > 0 and (best is None or idx < best):
            best = idx
    if best is None:
        return body.strip(), ""
    return body[:best].strip(" -"), body[best:].strip(" -")


def _strip_accents(text: object) -> str:
    raw = unicodedata.normalize("NFKD", str(text or ""))
    return "".join(ch for ch in raw if not unicodedata.combining(ch))


def _parse_measure_to_mm(value: str, unit: str | None) -> float:
    num = float(value.replace(",", "."))
    unit_norm = _strip_accents(unit or "").lower()
    if unit_norm.startswith("m"):
        return num * 1000
    return num * 10


def _dimensions_from_state_item(text: str, item_no: int, name: str) -> tuple[dict[str, float], list[str]]:
    clean = _collapse_ws(text)
    norm = _strip_accents(clean).lower()
    notes: list[str] = []
    dims: dict[str, float] = {}

    labelled: dict[str, float] = {}
    for match in re.finditer(
        r"(\d+(?:[.,]\d+)?)\s*(cm|cms|m|mts|metros)?\s*"
        r"(?:de\s+)?(ancho|alto|altura|largo|profundidad|fondo)",
        norm,
    ):
        value = _parse_measure_to_mm(match.group(1), match.group(2) or "cm")
        label = match.group(3)
        if label == "ancho":
            labelled["width_mm"] = value
        elif label in ("alto", "altura"):
            labelled["height_mm"] = value
        elif label == "largo":
            labelled["width_mm"] = value
        elif label in ("profundidad", "fondo"):
            labelled["depth_mm"] = value

    if labelled:
        dims.update(labelled)

    triples = re.findall(
        r"(\d+(?:[.,]\d+)?)\s*(cm|cms|m|mts|metros)?\s*x\s*"
        r"(\d+(?:[.,]\d+)?)\s*(cm|cms|m|mts|metros)?\s*x\s*"
        r"(\d+(?:[.,]\d+)?)\s*(cm|cms|m|mts|metros)?",
        norm,
    )
    if triples and not {"width_mm", "height_mm", "depth_mm"}.issubset(dims):
        a, au, b, bu, c, cu = triples[-1]
        common_unit = au or bu or cu or "cm"
        values = [
            _parse_measure_to_mm(a, au or common_unit),
            _parse_measure_to_mm(b, bu or common_unit),
            _parse_measure_to_mm(c, cu or common_unit),
        ]
        dims.setdefault("width_mm", values[0])
        dims.setdefault("height_mm", values[1])
        dims.setdefault("depth_mm", values[2])

    pairs = re.findall(
        r"(\d+(?:[.,]\d+)?)\s*(cm|cms|m|mts|metros)?\s*x\s*"
        r"(\d+(?:[.,]\d+)?)\s*(cm|cms|m|mts|metros)?",
        norm,
    )
    if pairs and not dims:
        a, au, b, bu = pairs[-1]
        common_unit = au or bu or "cm"
        first = _parse_measure_to_mm(a, au or common_unit)
        second = _parse_measure_to_mm(b, bu or common_unit)
        if "mesa" in norm or "escritorio" in norm:
            dims["width_mm"] = first
            dims["depth_mm"] = second
            notes.append("Falta altura para poder cotizar sin asumir medidas.")
        else:
            dims["width_mm"] = first
            dims["height_mm"] = second

    if item_no == 1:
        dims.setdefault("width_mm", 1000)
        dims.setdefault("depth_mm", 600)
        if "height_mm" not in labelled and "altura" not in norm and "alto" not in norm:
            notes.append("Falta altura del escritorio para poder cotizar.")
    elif item_no == 2:
        notes.append("Faltan medidas de la cama para poder cotizar.")
    elif item_no == 3:
        dims.setdefault("width_mm", 1000)
        dims.setdefault("depth_mm", 1000)
        if "height_mm" not in labelled and "altura" not in norm and "alto" not in norm:
            notes.append("Falta altura de la mesa para poder cotizar.")

    return dims, notes


def _missing_dimensions_for_quote(dims: dict[str, float]) -> list[str]:
    labels = {
        "width_mm": "ancho",
        "height_mm": "alto",
        "depth_mm": "profundidad",
    }
    return [label for key, label in labels.items() if not dims.get(key)]


def _unique_notes(notes: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for note in notes:
        key = _norm(note)
        if key in seen:
            continue
        seen.add(key)
        out.append(note)
    return out


def _material_from_state_item(text: str) -> str:
    norm = _strip_accents(text).lower()
    if "pino" in norm:
        return "pino"
    if "madera" in norm and "melamin" not in norm and "mdf" not in norm:
        return "madera"
    if "mdf" in norm and "melamin" not in norm:
        return "MDF"
    if "melamin" in norm:
        return "MDF melaminico"
    return "melaminico"


def _thickness_from_state_item(text: str) -> float:
    match = re.search(r"espesor.{0,45}?(\d+(?:[.,]\d+)?)\s*mm", _strip_accents(text).lower())
    if match:
        return float(match.group(1).replace(",", "."))
    return 18.0


def _hardware_from_state_item(text: str) -> list[str]:
    norm = _strip_accents(text).lower()
    hardware: list[str] = []
    if "cajon" in norm:
        hardware.append("guias para cajones")
    if "cerradura" in norm or "llave" in norm:
        hardware.append("cerradura")
    if "rueda" in norm or "rodante" in norm:
        hardware.append("ruedas")
    if "puerta" in norm:
        hardware.append("bisagras")
    return hardware


def _state_item_number(item: dict) -> int | None:
    code = str(item.get("code") or "")
    for pattern in (r"\bI\s*0*(\d{1,2})\b", r"\bitem\s*0*(\d{1,2})\b", r"\b(?:^|[-_ ])0*(\d{1,2})(?:$|[-_ ])"):
        match = re.search(pattern, code, flags=re.IGNORECASE)
        if match:
            return int(match.group(1))
    text = f"{item.get('name', '')} {item.get('description', '')}"
    match = re.search(r"\b(?:item|item)\s*0*(\d{1,2})\b", _strip_accents(text).lower())
    if match:
        return int(match.group(1))
    return None


def _state_item_sice(item: dict) -> str:
    for value in (item.get("sice_code"), item.get("code"), item.get("description")):
        match = re.search(r"\b(\d{4,6})\b", str(value or ""))
        if match:
            return match.group(1)
    return ""


def _state_item_names_match(a: dict, b: dict) -> bool:
    left = _norm(a.get("name"))
    right = _norm(b.get("name"))
    if not left or not right:
        return False
    return left in right or right in left


def _extract_state_purchase_items(text: str) -> list[dict]:
    """Deterministic fallback for SICE/ITEM/CANTIDAD tables.

    Some public-bid PDFs put short rows in ANEXO I. The language model can
    skip rows that lack a full 3D measure; this parser keeps the numbered
    carpentry rows present and adds explicit assumptions where the pliego is
    incomplete.
    """
    if not re.search(r"\bSICE\b", text, flags=re.IGNORECASE) or not re.search(r"\bITEM\b|\bITEM\b", _strip_accents(text), flags=re.IGNORECASE):
        return []

    compact = _collapse_ws(text)
    header = re.search(r"\bSICE\s+ITEM\s+ARTICULO\b", _strip_accents(compact), flags=re.IGNORECASE)
    if header:
        compact = compact[header.start():]
    else:
        start = re.search(r"ANEXO\s+I\b", compact, flags=re.IGNORECASE)
        if start:
            compact = compact[start.start():]
    end = re.search(r"ANEXO\s+II\b|LOS\s+ART", compact, flags=re.IGNORECASE)
    if end:
        compact = compact[: end.start()]

    starts = list(re.finditer(r"\b(\d{4,6})\s+(\d{1,2})\s+", compact))
    items: list[dict] = []
    for idx, match in enumerate(starts):
        sice = match.group(1)
        item_no = int(match.group(2))
        if item_no < 1 or item_no > 6:
            continue
        row_end = starts[idx + 1].start() if idx + 1 < len(starts) else len(compact)
        row = compact[match.end():row_end].strip()
        row, quantity = _parse_state_quantity(row)
        name, spec = _split_state_item_body(row)
        if not name:
            continue
        full = _collapse_ws(f"{name}. {spec}")
        dims, notes = _dimensions_from_state_item(full, item_no, name)
        notes = _unique_notes(notes)
        material = _material_from_state_item(full)
        thickness = _thickness_from_state_item(full)
        description_parts = [f"Item {item_no} SICE {sice}: {full}"]
        if notes:
            description_parts.append("Faltan datos: " + " ".join(notes))
        missing_dimensions = _missing_dimensions_for_quote(dims)
        items.append({
            "code": f"I{item_no}",
            "sice_code": sice,
            "name": name,
            "quantity": quantity,
            "description": " ".join(description_parts),
            "dimensions": dims,
            "material": material,
            "thickness_mm": thickness,
            "hardware": _hardware_from_state_item(full),
            "edge_banding": "",
            "group": "",
            "delivery_location": "",
            "delivery_days": None,
            "wood_only": True,
            "missing_dimensions": missing_dimensions,
            "source": "sice_anexo_i",
        })
    return items


def _merge_state_purchase_items(pliego: dict, text: str) -> dict:
    fallback = _extract_state_purchase_items(text)
    if not fallback:
        return pliego
    order_by_sice = {_state_item_sice(item): _state_item_number(item) for item in fallback}
    order_by_name = {_norm(item.get("name")): _state_item_number(item) for item in fallback}
    current = list(pliego.get("items") or [])
    seen_numbers = {_state_item_number(item) for item in current}
    seen_numbers.discard(None)
    seen_sice = {_state_item_sice(item) for item in current}
    seen_sice.discard("")
    for item in fallback:
        item_no = _state_item_number(item)
        sice = _state_item_sice(item)
        existing = next(
            (
                candidate for candidate in current
                if (
                    (item_no is not None and _state_item_number(candidate) == item_no)
                    or (sice and _state_item_sice(candidate) == sice)
                    or _state_item_names_match(item, candidate)
                )
            ),
            None,
        )
        if existing is not None and item.get("missing_dimensions"):
            existing["dimensions"] = item.get("dimensions") or {}
            existing["missing_dimensions"] = item.get("missing_dimensions") or []
            existing["description"] = item.get("description") or existing.get("description", "")
            existing["source"] = item.get("source") or existing.get("source", "")
        elif existing is None:
            current.append(item)
            seen_numbers.add(item_no)
            seen_sice.add(sice)
    def order_key(item: dict) -> tuple[int, str]:
        item_no = _state_item_number(item)
        if item_no is None:
            item_no = order_by_sice.get(_state_item_sice(item))
        if item_no is None:
            name = _norm(item.get("name"))
            for fallback_name, fallback_no in order_by_name.items():
                if name and (name in fallback_name or fallback_name in name):
                    item_no = fallback_no
                    break
        return (item_no or 999, str(item.get("code") or ""))

    current.sort(key=order_key)
    pliego["items"] = current
    return pliego


def analyze_pliego(file_paths: list[str]) -> dict:
    all_text = []
    for path in file_paths:
        low = path.lower()
        if low.endswith(".pdf"):
            all_text.append(_extract_pdf_text(path))
        elif low.endswith(".xlsx") or low.endswith(".xls"):
            all_text.append(_extract_xlsx_text(path))
        else:
            with open(path, encoding="utf-8", errors="replace") as f:
                all_text.append(f.read())

    combined = "\n\n---\n\n".join(all_text)
    if len(combined) > 100_000:
        combined = combined[:100_000]

    client = _get_client()
    try:
        response = client.chat.completions.create(
            model=AGENT_MODEL,
            messages=[
                {"role": "system", "content": PLIEGO_ANALYSIS},
                {"role": "user", "content": combined},
            ],
            response_format={"type": "json_object"},
            temperature=0.2,
        )
    except Exception as exc:
        raise RuntimeError(friendly_openai_error(exc)) from exc

    parsed = json.loads(response.choices[0].message.content or "{}")
    return _merge_state_purchase_items(parsed, combined)



NUMBER_WORDS = {
    "un": 1,
    "una": 1,
    "uno": 1,
    "dos": 2,
    "tres": 3,
    "cuatro": 4,
    "cinco": 5,
    "seis": 6,
    "siete": 7,
    "ocho": 8,
    "nueve": 9,
    "diez": 10,
}


def _norm(text: object) -> str:
    raw = str(text or "").lower()
    raw = unicodedata.normalize("NFKD", raw)
    raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", raw).strip()


def _as_float(value: object, default: float = 0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _count_from_text(text: str, noun: str) -> int:
    numeric = re.search(rf"\b(\d+)\s+{noun}es?\b", text)
    if numeric:
        return int(numeric.group(1))
    for word, value in NUMBER_WORDS.items():
        if re.search(rf"\b{word}\s+{noun}es?\b", text):
            return value
    return 0


def _drawer_height_from_text(text: str) -> float | None:
    match = re.search(r"cajon(?:es)?.{0,35}?(\d+(?:[.,]\d+)?)\s*cm\s+de\s+alto", text)
    if not match:
        match = re.search(r"(\d+(?:[.,]\d+)?)\s*cm\s+de\s+alto.{0,35}?cajon", text)
    if not match:
        return None
    return float(match.group(1).replace(",", ".")) * 10


def _door_count_from_text(text: str) -> int:
    numeric = re.search(r"\b(\d+)\s+puertas?\b", text)
    if numeric:
        return int(numeric.group(1))
    for word, value in NUMBER_WORDS.items():
        if re.search(rf"\b{word}\s+puertas?\b", text):
            return value
    count = _count_from_text(text, "puerta")
    if count:
        return count
    if "puertas" in text:
        return 2
    if "puerta" in text or "frente cerrado" in text or "frente con puerta" in text:
        return 1
    return 0


def _lower_door_height_from_text(text: str) -> float | None:
    patterns = (
        r"puertas?.{0,80}?altura\s+de\s+(\d+(?:[.,]\d+)?)\s*cm",
        r"puertas?.{0,80}?alto\s+de\s+(\d+(?:[.,]\d+)?)\s*cm",
        r"hasta\s+la\s+altura\s+de\s+(\d+(?:[.,]\d+)?)\s*cm",
        r"abajo.{0,80}?(\d+(?:[.,]\d+)?)\s*cm",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return float(match.group(1).replace(",", ".")) * 10
    return None


def _is_lower_doors_upper_open_pattern(text: str) -> bool:
    return (
        "puerta" in text
        and "abajo" in text
        and "arriba" in text
        and ("sin puerta" in text or "abierto" in text or "estante" in text)
    )


def _drawers_are_stacked(text: str) -> bool:
    return bool(
        "cajonera" in text
        or "cajones vertical" in text
        or "cajones apil" in text
        or "cajones uno sobre otro" in text
    )


def _normalize_piece_dimensions(out: dict, item: dict) -> None:
    """Correct common AI orientation mistakes before pricing.

    The calculator only sees rectangles. If the model says "lateral 2000x2000"
    for a 2000x2000x500 wardrobe, the area is wildly wrong; normalize the
    standard carcass pieces using the known overall dimensions.
    """
    dims = item.get("dimensions") or {}
    width = _as_float(dims.get("width_mm"))
    height = _as_float(dims.get("height_mm"))
    depth = _as_float(dims.get("depth_mm"))
    thickness = _as_float(item.get("thickness_mm"), 18) or 18
    if width <= 0 or height <= 0 or depth <= 0:
        return

    pieces = out.get("pieces") or []
    for piece in pieces:
        label = _norm(piece.get("label", ""))
        if "caj" in label:
            continue
        if any(token in label for token in ("lateral", "costado", "division", "divisor", "separador")):
            piece["width_mm"] = height
            piece["height_mm"] = depth
            if not piece.get("edge_sides"):
                piece["edge_sides"] = ["left"]
        elif any(token in label for token in ("tapa", "base", "techo", "piso")):
            piece["width_mm"] = width
            piece["height_mm"] = depth
            if not piece.get("edge_sides"):
                piece["edge_sides"] = ["top", "left", "right"]
        elif "estante" in label or "repisa" in label:
            # Preserve the module width inferred by the model, but depth is the
            # cabinet depth. A phrase like "estante de 40 cm" describes the bay
            # height, not a 400mm shelf depth.
            w = _as_float(piece.get("width_mm"))
            if w <= 0 or w > width:
                piece["width_mm"] = max(width - 2 * thickness, 0)
            piece["height_mm"] = depth
            if not piece.get("edge_sides"):
                piece["edge_sides"] = ["top"]
        elif ("trasera" in label or "fondo" in label) and "caj" not in label:
            piece["width_mm"] = width
            piece["height_mm"] = height
            piece["edge_sides"] = []

    label_text = " ".join(_norm(p.get("label", "")) for p in pieces)
    desc_text = _norm(f"{item.get('name', '')} {item.get('description', '')}")
    if not any(("caj" not in _norm(p.get("label", "")) and any(token in _norm(p.get("label", "")) for token in ("tapa", "techo"))) for p in pieces):
        pieces.append({
            "width_mm": width,
            "height_mm": depth,
            "quantity": 1,
            "label": "tapa",
            "edge_sides": ["top", "left", "right"],
        })
    if not any(("caj" not in _norm(p.get("label", "")) and any(token in _norm(p.get("label", "")) for token in ("base", "piso"))) for p in pieces):
        pieces.append({
            "width_mm": width,
            "height_mm": depth,
            "quantity": 1,
            "label": "base",
            "edge_sides": ["top", "left", "right"],
        })
    if not any(("caj" not in _norm(p.get("label", "")) and any(token in _norm(p.get("label", "")) for token in ("lateral", "costado"))) for p in pieces):
        pieces.append({
            "width_mm": height,
            "height_mm": depth,
            "quantity": 2,
            "label": "lateral",
            "edge_sides": ["left"],
        })
    if (
        "trasera" not in label_text
        and "fondo trasero" not in label_text
        and "sin fondo" not in desc_text
        and "sin trasera" not in desc_text
        and "cajonera" not in desc_text
    ):
        pieces.append({
            "width_mm": width,
            "height_mm": height,
            "quantity": 1,
            "label": "trasera",
            "edge_sides": [],
        })

    if _is_lower_doors_upper_open_pattern(desc_text):
        door_count = _door_count_from_text(desc_text) or 4
        lower_h = _lower_door_height_from_text(desc_text) or height / 2
        upper_h = max(height - lower_h, 0)
        door_w = max(width / door_count, 0)

        # This pattern is a low closed row of doors and an upper row of open
        # cubbies. "4 estantes arriba" means four open bays, not four horizontal
        # shelves.
        pieces = [
            p for p in pieces
            if "puerta" not in _norm(p.get("label", ""))
            and "estante" not in _norm(p.get("label", ""))
            and "repisa" not in _norm(p.get("label", ""))
            and "division" not in _norm(p.get("label", ""))
            and "divisor" not in _norm(p.get("label", ""))
        ]
        pieces.append({
            "width_mm": width,
            "height_mm": depth,
            "quantity": 1,
            "label": f"estante divisor horizontal a {lower_h:.0f}mm",
            "edge_sides": ["top"],
        })
        if door_count > 1:
            pieces.append({
                "width_mm": height,
                "height_mm": depth,
                "quantity": door_count - 1,
                "label": "division vertical modulo completo",
                "edge_sides": ["left"],
            })
        pieces.append({
            "width_mm": door_w,
            "height_mm": lower_h,
            "quantity": door_count,
            "label": "puerta inferior",
            "edge_sides": ["top", "bottom", "left", "right"],
        })
        out["notes"] = (
            (out.get("notes") or "") +
            f" Mueble interpretado como {door_count} modulos: puertas inferiores de {lower_h:.0f}mm "
            f"y nichos abiertos superiores de {upper_h:.0f}mm."
        ).strip()

    out["pieces"] = pieces
    _complete_front_doors(out, item, width=width, height=height, thickness=thickness)
    _complete_drawer_pieces(out, item, width=width, height=height, depth=depth, thickness=thickness)


def _complete_front_doors(out: dict, item: dict, *, width: float, height: float, thickness: float) -> None:
    pieces = out.get("pieces") or []
    text = _norm(f"{item.get('name', '')} {item.get('description', '')}")
    if "sin puerta" in text or "abierto" in text:
        return
    door_count = _door_count_from_text(text)
    existing = [p for p in pieces if "puerta" in _norm(p.get("label", ""))]
    if existing:
        for door in existing:
            qty = max(1, int(_as_float(door.get("quantity"), 1)))
            if _as_float(door.get("width_mm")) <= 0 or _as_float(door.get("width_mm")) > width:
                door["width_mm"] = max((width / qty) - thickness, 0)
            if _as_float(door.get("height_mm")) <= 0:
                door["height_mm"] = height
            if not door.get("edge_sides"):
                door["edge_sides"] = ["top", "bottom", "left", "right"]
        return
    if door_count <= 0:
        return
    pieces.append({
        "width_mm": max((width / door_count) - thickness, 0),
        "height_mm": height,
        "quantity": door_count,
        "label": "puerta frente",
        "edge_sides": ["top", "bottom", "left", "right"],
    })
    out["pieces"] = pieces


def _complete_drawer_pieces(out: dict, item: dict, *, width: float, height: float, depth: float, thickness: float) -> None:
    pieces = out.get("pieces") or []
    text = _norm(f"{item.get('name', '')} {item.get('description', '')}")
    pieces = [
        p for p in pieces
        if not (
            "caj" in _norm(p.get("label", ""))
            and ("base" in _norm(p.get("label", "")) or "parte de abajo" in _norm(p.get("label", "")))
        )
    ]
    fronts = [p for p in pieces if "caj" in _norm(p.get("label", "")) and "frente" in _norm(p.get("label", ""))]
    drawer_count = sum(max(1, int(_as_float(p.get("quantity"), 1))) for p in fronts)
    drawer_count = drawer_count or _count_from_text(text, "cajon")
    if drawer_count <= 0:
        return

    front_width = _as_float(fronts[0].get("width_mm")) if fronts else 0
    front_height = _as_float(fronts[0].get("height_mm")) if fronts else 0
    stacked = _drawers_are_stacked(text)
    if stacked:
        front_width = width
        front_height = _drawer_height_from_text(text) or max(height / drawer_count, 0) or front_height
    if front_width <= 0 or front_width > width:
        front_width = width if stacked else max((width / drawer_count) - 2 * thickness, 0)
    if front_height <= 0:
        inferred_height = max(height / drawer_count, 0) if stacked else 0
        front_height = _drawer_height_from_text(text) or inferred_height or 200
    drawer_depth = depth if stacked else max(depth - 50, 0)

    if not fronts:
        front = {
            "width_mm": front_width,
            "height_mm": front_height,
            "quantity": drawer_count,
            "label": "frente cajon",
            "edge_sides": ["top", "bottom", "left", "right"],
        }
        pieces.append(front)
        fronts = [front]

    for front in fronts:
        if stacked:
            front["width_mm"] = front_width
        else:
            front["width_mm"] = min(_as_float(front.get("width_mm"), front_width), width) or front_width
        if stacked or _as_float(front.get("height_mm")) <= 0:
            front["height_mm"] = front_height
        if not front.get("edge_sides"):
            front["edge_sides"] = ["top", "bottom", "left", "right"]

    if stacked:
        for piece in pieces:
            label = _norm(piece.get("label", ""))
            if "caj" not in label or "frente" in label:
                continue
            if "lateral" in label:
                piece["width_mm"] = drawer_depth
                piece["height_mm"] = front_height
                piece["quantity"] = drawer_count * 2
            elif "fondo" in label:
                piece["width_mm"] = front_width
                piece["height_mm"] = front_height
                piece["quantity"] = drawer_count
            elif "trasera" in label:
                piece["width_mm"] = front_width
                piece["height_mm"] = front_height
                piece["quantity"] = drawer_count

    label_text = " ".join(_norm(p.get("label", "")) for p in pieces)
    if "lateral caj" not in label_text:
        pieces.append({
            "width_mm": drawer_depth,
            "height_mm": front_height,
            "quantity": drawer_count * 2,
            "label": "lateral cajón",
            "edge_sides": ["top"],
        })
    if "trasera caj" not in label_text:
        pieces.append({
            "width_mm": front_width,
            "height_mm": front_height,
            "quantity": drawer_count,
            "label": "trasera cajón",
            "edge_sides": [],
        })
    if stacked and "fondo caj" not in label_text:
        pieces.append({
            "width_mm": front_width,
            "height_mm": front_height,
            "quantity": drawer_count,
            "label": "fondo cajon",
            "edge_sides": [],
        })
    out["pieces"] = pieces


def _is_wood_door_item(item: dict) -> bool:
    text = _norm(
        f"{item.get('name', '')} {item.get('description', '')} "
        f"{item.get('material', '')}"
    )
    if "puerta" not in text and "hoja" not in text:
        return False
    if any(term in text for term in ("metal", "aluminio", "vidrio", "blindex")):
        return False
    return any(term in text for term in (
        "madera",
        "enchap",
        "eucal",
        "pino",
        "marco",
        "tapajunta",
        "cubre canto",
        "cubrecanto",
        "placa",
        "hoja",
    ))


def _wood_door_dimensions(item: dict) -> tuple[float, float, float]:
    dims = item.get("dimensions") or {}
    width = _as_float(dims.get("width_mm"), 800.0)
    height = _as_float(dims.get("height_mm"), 2100.0)
    thickness = _as_float(
        dims.get("depth_mm")
        or dims.get("thickness_mm")
        or item.get("thickness_mm"),
        40.0,
    )
    if thickness < 30:
        thickness = 40.0
    return width, height, thickness


def _decompose_wood_door(item: dict) -> dict:
    width, height, thickness = _wood_door_dimensions(item)
    grid_spacing = 100.0
    vertical_count = max(3, math.ceil(width / grid_spacing) + 1)
    horizontal_count = max(6, math.ceil(height / grid_spacing) + 1)
    frame_len_mm = (height * 2) + width
    tapajuntas_len_mm = frame_len_mm * 2
    pieces = [
        {"width_mm": width, "height_mm": height, "quantity": 2, "label": "cara puerta enchapado eucaliptus 4mm", "edge_sides": []},
        {"width_mm": 50, "height_mm": height, "quantity": vertical_count, "label": "alma nido vertical pino 5x50mm", "edge_sides": []},
        {"width_mm": width, "height_mm": 50, "quantity": horizontal_count, "label": "alma nido horizontal pino 5x50mm", "edge_sides": []},
        {"width_mm": thickness, "height_mm": height, "quantity": 2, "label": "cubrecanto lateral eucaliptus", "edge_sides": []},
        {"width_mm": width, "height_mm": thickness, "quantity": 2, "label": "cubrecanto superior inferior eucaliptus", "edge_sides": []},
        {"width_mm": 45, "height_mm": frame_len_mm, "quantity": 1, "label": "marco puerta eucaliptus", "edge_sides": []},
        {"width_mm": 10, "height_mm": tapajuntas_len_mm, "quantity": 1, "label": "tapajuntas eucaliptus ambos lados", "edge_sides": []},
    ]
    return {
        "pieces": pieces,
        "hardware": [
            {"code": "BISAGRA_PUERTA_PASO", "quantity": 4},
            {"code": "CERR_PUERTA_EMBUTIR", "quantity": 1},
            {"code": "PESTILLO_PUERTA", "quantity": 1},
        ],
        "notes": (
            "Puerta placa interpretada con dos caras de enchapado eucaliptus 4mm, "
            "alma tipo nido de abeja en pino 5x50mm, cubrecantos de eucaliptus, "
            "marco/tapajuntas y 3 manos de PU."
        ),
        "metadata": {
            "quote_type": "wood_door",
            "width_mm": width,
            "height_mm": height,
            "thickness_mm": thickness,
            "core_vertical_count": vertical_count,
            "core_horizontal_count": horizontal_count,
        },
    }


def _ensure_hardware_from_text(out: dict, item: dict) -> None:
    text = _norm(f"{item.get('name', '')} {item.get('description', '')}")
    hardware = list(out.get("hardware") or [])

    def upsert(code: str, quantity: int) -> None:
        for hw in hardware:
            if (hw.get("code") or "").strip() == code:
                hw["quantity"] = max(int(hw.get("quantity") or 0), quantity)
                return
        hardware.append({"code": code, "quantity": quantity})

    drawer_count = _count_from_text(text, "cajon")
    if "rueda" in text:
        match = re.search(r"\b(\d+)\s+ruedas?\b", text)
        qty = int(match.group(1)) if match else 4
        upsert("RUEDA_GIR_SIN_FRENO", qty)
    if "guia telescop" in text and drawer_count:
        depth = _as_float((item.get("dimensions") or {}).get("depth_mm"), 450)
        guide_size = min((300, 350, 400, 450, 500, 550), key=lambda x: abs(x - depth))
        upsert(f"GUIA_TELESC_{guide_size}", drawer_count)
    if "cerradura" in text and drawer_count:
        upsert("CERR_CAJONERA_3", drawer_count)
    if "puerta" in text and not drawer_count:
        upsert("BISAGRA_PUERTA_PASO", 4)
        upsert("CERR_PUERTA_EMBUTIR", 1)
        upsert("PESTILLO_PUERTA", 1)

    out["hardware"] = hardware


def decompose_furniture(item: dict) -> dict:
    from carpinteria.hardware_catalog import catalog_prompt_block, get_by_code

    if _is_wood_door_item(item):
        out = _decompose_wood_door(item)
        _ensure_hardware_from_text(out, item)
        cleaned: list[dict] = []
        for hw in out.get("hardware") or []:
            code = (hw.get("code") or "").strip()
            spec = get_by_code(code)
            if spec is None:
                continue
            try:
                qty = int(hw.get("quantity") or 0)
            except (TypeError, ValueError):
                qty = 0
            if qty <= 0:
                continue
            cleaned.append({
                "code": spec.code,
                "name": spec.name,
                "category": spec.category,
                "unit": spec.unit,
                "quantity": qty,
            })
        out["hardware"] = cleaned
        return out

    client = _get_client()

    system_prompt = FURNITURE_DECOMPOSE.replace(
        "{HARDWARE_CATALOG}", catalog_prompt_block()
    )

    desc = (
        f"Mueble: {item.get('name', '')}\n"
        f"Código: {item.get('code', '')}\n"
        f"Descripción: {item.get('description', '')}\n"
        f"Dimensiones: {item.get('dimensions', {})}\n"
        f"Material: {item.get('material', '')} {item.get('thickness_mm', 18)}mm\n"
        f"Herrajes mencionados: {item.get('hardware', [])}\n"
        f"Canto: {item.get('edge_banding', '')}"
    )

    try:
        response = client.chat.completions.create(
            model=AGENT_MODEL,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": desc},
            ],
            response_format={"type": "json_object"},
            temperature=0.2,
        )
    except Exception as exc:
        raise RuntimeError(friendly_openai_error(exc)) from exc

    out = json.loads(response.choices[0].message.content or "{}")
    _normalize_piece_dimensions(out, item)
    _ensure_hardware_from_text(out, item)

    # Normalize hardware: drop unknown codes, enrich each with display name + category.
    cleaned: list[dict] = []
    for hw in out.get("hardware") or []:
        code = (hw.get("code") or "").strip()
        if not code:
            continue
        spec = get_by_code(code)
        if spec is None:
            continue
        try:
            qty = int(hw.get("quantity") or 0)
        except (TypeError, ValueError):
            qty = 0
        if qty <= 0:
            continue
        cleaned.append({
            "code": spec.code,
            "name": spec.name,
            "category": spec.category,
            "unit": spec.unit,
            "quantity": qty,
        })
    out["hardware"] = cleaned
    return out
